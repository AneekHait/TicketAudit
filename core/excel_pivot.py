"""
Native Excel PivotTable and Table (ListObject) construction on top of openpyxl.

openpyxl documents its pivot support as read/preserve only - there is no public
"create a pivot" API. But `Worksheet.add_pivot` exists, and openpyxl's writer
does serialise `ws._pivots` and their caches on save, so a pivot assembled by
hand from the `openpyxl.pivot.*` classes is written out correctly. This module
is that assembly, kept in one place because the schema has three separate index
spaces and several invariants openpyxl does not enforce.

Nothing here imports Qt or touches a DataFrame; it takes plain header/row
sequences so it can be unit tested without Excel.

The invariants that matter, each of which produces a workbook Excel offers to
"repair" (silently dropping every pivot) rather than an exception:

1. One CacheDefinition *object* and one cacheId per distinct source. The writer
   dedupes caches by structural equality, so two identical-but-separately-built
   cache objects collapse into one part while two cacheIds are emitted, leaving
   the second pivot pointing at nothing. `PivotBuilder` hands both out together
   so they cannot drift apart.
2. len(pivotFields) == len(cacheFields), and every rowFields/@x and
   dataFields/@fld must index into that list. -2 is the legal special value
   meaning "the data axis".
3. Location.ref spans the whole rendered block, header through grand total.

`validate()` checks all three against the produced file without needing Excel.
"""
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple
from zipfile import ZipFile

import numpy as np

from openpyxl.pivot.cache import (
    CacheDefinition, CacheField, CacheSource, SharedItems, WorksheetSource,
)
from openpyxl.pivot.fields import Index, Number, Text
from openpyxl.pivot.record import Record, RecordList
from openpyxl.pivot.table import (
    DataField, FieldItem, Location, PivotField, PivotTableStyle, RowColField,
    RowColItem, TableDefinition,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.table import Table, TableStyleInfo

# createdVersion / refreshedVersion / minRefreshableVersion. openpyxl defaults
# these to 0, which describes a pre-2007 pivot; 3 is Excel 2007 and is the most
# permissive value that every modern Excel will agree to refresh.
PIVOT_VERSION = 3

# Excel's own limits, not openpyxl's - openpyxl enforces neither.
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_CELL_CHARS = 32_767


# ---------------------------------------------------------------- Excel Tables

def table_name(label: str, used: Set[str]) -> str:
    """
    A ListObject name Excel will accept, unique within `used`.

    Excel is stricter than openpyxl here: openpyxl only rejects spaces, while
    Excel additionally requires a letter/underscore start, no characters outside
    [A-Za-z0-9._], and a name that is not shaped like a cell reference. Tables
    and defined names also share one namespace, so `used` must be seeded with
    both. Mutates `used`.
    """
    cleaned = re.sub(r"[^A-Za-z0-9_.]", "_", str(label)).strip("_") or "Data"
    # The tbl_ prefix guarantees a letter start and cannot collide with a cell
    # reference like A1 or R1C1, which Excel rejects outright.
    candidate = f"tbl_{cleaned}"[:255]
    base, suffix = candidate, 2
    while candidate.casefold() in used:
        candidate = f"{base}_{suffix}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def register_table(worksheet, name: str, ref: str, *,
                   style: str = "TableStyleMedium2") -> Table:
    """
    Turn a written range into a named Excel Table.

    Beyond the filter buttons, this is what makes `Insert > PivotTable` a
    three-click operation: Excel pre-fills the Table/Range box from the
    ListObject containing the active cell, and a Table name is self-describing
    wherever it appears.

    In write-only mode the caller must set `tableColumns` explicitly - see
    `write_only_table_columns` - or the table ships with Column1..ColumnN, which
    disagrees with the real header row and makes Excel offer a repair.
    """
    table = Table(displayName=name, ref=ref, headerRowCount=1)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)
    return table


def write_only_table_columns(table: Table, headers: Sequence[str]) -> Table:
    """
    Populate a Table's column names by hand.

    Required in write-only mode: openpyxl's table writer reads the header row
    off the worksheet to name the columns, and a write-only sheet has no
    __getitem__, so it silently falls back to Column1..ColumnN.
    """
    from openpyxl.worksheet.table import TableColumn

    table.tableColumns = [
        TableColumn(id=i + 1, name=str(header))
        for i, header in enumerate(headers)
    ]
    return table


def range_ref(first_row: int, first_col: int, last_row: int, last_col: int) -> str:
    """A1-style range from 1-based row/column bounds."""
    return (f"{get_column_letter(first_col)}{first_row}"
            f":{get_column_letter(last_col)}{last_row}")


# ------------------------------------------------------------------ pivot cache

def _py(value: Any) -> Any:
    """
    Unwrap numpy scalars to their Python equivalents.

    A pandas aggregate hands back np.int64, which is NOT an instance of int on
    Windows. Left alone it falls through the numeric branch into Text(), so the
    measure ships as a string and the pivot's Sum silently produces zero - a
    pivot that opens cleanly and is simply wrong.
    """
    if isinstance(value, np.generic):
        return value.item()
    return value


def _is_number(value: Any) -> bool:
    return (isinstance(value, (int, float)) and not isinstance(value, bool))


def _shared_items_for(values: Sequence[Any]) -> Tuple[SharedItems, Optional[List[Any]]]:
    """
    Build a cacheField's sharedItems.

    Axis fields (anything that will sit on a row/column/page axis) need the
    distinct values enumerated, because the pivot's FieldItem.x values index
    into them positionally. Measure fields do not, and declaring their numeric
    range instead keeps the cache small.

    Returns (sharedItems, distinct_values_or_None).
    """
    numbers = [v for v in values if _is_number(v)]
    if numbers and len(numbers) == len(values):
        return SharedItems(
            containsSemiMixedTypes=False, containsString=False,
            containsNumber=True,
            containsInteger=all(float(n).is_integer() for n in numbers),
            minValue=min(numbers), maxValue=max(numbers),
        ), None
    return SharedItems(containsSemiMixedTypes=True, containsString=True), None


class PivotBuilder:
    """
    Builds pivots over one rectangular source, on one target sheet.

    Owns the cache so invariant 1 cannot be violated: every pivot it creates
    shares a single CacheDefinition object *and* a single cacheId. It also
    records each pivot's rendered bounding box so `assert_disjoint` can prove
    they do not overlap, which matters because all pivots share one sheet in
    this report - an overlap is not a corruption but a modal alert on refresh.
    """

    def __init__(self, headers: Sequence[str], rows: Sequence[Sequence[Any]],
                 *, axis_fields: Sequence[int], cache_id: int = 1,
                 source_name: Optional[str] = None,
                 source_sheet: Optional[str] = None,
                 source_ref: Optional[str] = None,
                 save_records: bool = True):
        """
        headers/rows  the source block, exactly as written to the sheet
        axis_fields   column indexes that will sit on a row/col/page axis
        source_name   an Excel Table name; preferred over sheet+ref because it
                      survives a sheet rename and auto-expands
        save_records  write the cache records in full. Keep this True: with
                      records absent the pivot relies on Excel honouring
                      refreshOnLoad, and Excel Online, LibreOffice and mobile
                      Excel generally do not - the pivot then renders empty for
                      anyone the report is forwarded to.
        """
        if source_name is None and (source_sheet is None or source_ref is None):
            raise ValueError("need either source_name or source_sheet+source_ref")

        self.headers = [str(h) for h in headers]
        # Normalised on the way in, so a numpy scalar cannot reach the record
        # writer and turn the measure into text.
        self.rows = [tuple(_py(v) for v in r) for r in rows]
        self.cache_id = cache_id
        self.targets: List[Tuple[str, Tuple[int, int, int, int]]] = []

        # Distinct values per axis field, in first-seen order. FieldItem.x and
        # RowColItem.x both index into this, and keeping one ordering for both
        # removes a whole class of off-by-one bugs.
        self._shared: Dict[int, List[Any]] = {}
        cache_fields = []
        for i, header in enumerate(self.headers):
            column = [r[i] for r in self.rows]
            if i in set(axis_fields):
                order, seen = [], set()
                for value in column:
                    if value not in seen:
                        seen.add(value)
                        order.append(value)
                self._shared[i] = order
                items = SharedItems(_fields=[Text(v=str(v)) for v in order])
            else:
                items, _ = _shared_items_for(column)
            cache_fields.append(CacheField(name=header, sharedItems=items))

        worksheet_source = (
            WorksheetSource(name=source_name) if source_name
            else WorksheetSource(ref=source_ref, sheet=source_sheet))

        self.cache = CacheDefinition(
            # invalid=True tells Excel the cached data cannot be trusted; only
            # meaningful when we omit the records part.
            invalid=not save_records,
            saveData=bool(save_records),
            # Always on: it makes Excel recompute the layout on open, so a
            # mistake in our hand-built items/rowItems/location self-heals
            # instead of shipping.
            refreshOnLoad=True,
            enableRefresh=True, backgroundQuery=False, upgradeOnRefresh=False,
            createdVersion=PIVOT_VERSION, refreshedVersion=PIVOT_VERSION,
            minRefreshableVersion=PIVOT_VERSION,
            # recordCount must be absent when the records part is, or the
            # definition contradicts itself.
            recordCount=len(self.rows) if save_records else None,
            cacheSource=CacheSource(type="worksheet",
                                    worksheetSource=worksheet_source),
            cacheFields=cache_fields,
        )

        if save_records:
            self.cache.records = RecordList(r=[
                Record(_fields=[self._record_field(i, v) for i, v in enumerate(row)])
                for row in self.rows
            ])

    def _record_field(self, index: int, value: Any):
        """One cell of a cache record: an index into sharedItems, or a literal."""
        if index in self._shared:
            return Index(v=self._shared[index].index(value))
        if _is_number(value):
            return Number(v=float(value))
        return Text(v="" if value is None else str(value))

    def add(self, worksheet, name: str, anchor: str, *, row_field: int,
            data_field: int, subtotal: str = "sum",
            style: str = "PivotStyleLight16") -> TableDefinition:
        """
        One pivot: a single row field, a single measure, tabular layout.

        Tabular layout (compact/outline all False) shows the real field name in
        the header instead of "Row Labels", which is what a report wants.
        """
        first_row, first_col = coordinate_to_tuple(anchor)
        items = self._shared.get(row_field, [])
        # header + one row per item + grand total; width = row header + measure
        last_row = first_row + len(items) + 1
        last_col = first_col + 1
        ref = range_ref(first_row, first_col, last_row, last_col)

        pivot_fields = []
        for i in range(len(self.headers)):
            if i == row_field:
                pivot_fields.append(PivotField(
                    axis="axisRow", showAll=False, compact=False, outline=False,
                    # The trailing default item is the subtotal slot, required
                    # because defaultSubtotal is on.
                    items=[FieldItem(x=k) for k in range(len(items))]
                          + [FieldItem(t="default")]))
            elif i == data_field:
                pivot_fields.append(PivotField(
                    dataField=True, showAll=False, compact=False, outline=False))
            else:
                pivot_fields.append(PivotField(
                    showAll=False, compact=False, outline=False))

        caption = f"{subtotal.capitalize()} of {self.headers[data_field]}"
        pivot = TableDefinition(
            name=name, cacheId=self.cache_id, dataCaption="Values",
            createdVersion=PIVOT_VERSION, updatedVersion=PIVOT_VERSION,
            minRefreshableVersion=PIVOT_VERSION,
            useAutoFormatting=True, itemPrintTitles=True, indent=0,
            compact=False, compactData=False, outline=False, outlineData=False,
            multipleFieldFilters=False,
            location=Location(ref=ref, firstHeaderRow=1,
                              # 1 with no column field; 2 once one is added.
                              firstDataRow=1, firstDataCol=1),
            pivotFields=pivot_fields,
            rowFields=[RowColField(x=row_field)],
            rowItems=[RowColItem(x=[Index(v=k)]) for k in range(len(items))]
                     + [RowColItem(t="grand", x=[Index(v=0)])],
            # Required even with no column fields.
            colItems=[RowColItem()],
            dataFields=[DataField(name=caption, fld=data_field,
                                  subtotal=subtotal, baseField=0, baseItem=0)],
            pivotTableStyleInfo=PivotTableStyle(
                name=style, showRowHeaders=True, showColHeaders=True,
                showRowStripes=False, showColStripes=False, showLastColumn=True),
        )
        pivot.cache = self.cache
        worksheet.add_pivot(pivot)
        self.targets.append(
            (name, (first_row, first_col, last_row, last_col)))
        return pivot

    def assert_disjoint(self, gutter: int = 2,
                        occupied: Sequence[Tuple[int, int]] = ()) -> None:
        """
        Prove no two pivots overlap, and that no explicitly written cell falls
        inside one.

        Excel reports an overlap as a modal alert during refresh and aborts,
        leaving the reader a stale block and a scary dialog - so it is worth
        failing the export instead. Fully decidable, because we know each
        pivot's rendered size.

        occupied: (row, col) cells the caller wrote itself, e.g. block titles.
        """
        for i, (name_a, box_a) in enumerate(self.targets):
            r1a, c1a, r2a, c2a = box_a
            for name_b, box_b in self.targets[i + 1:]:
                r1b, c1b, r2b, c2b = box_b
                rows_clear = r2a + gutter < r1b or r2b + gutter < r1a
                cols_clear = c2a + gutter < c1b or c2b + gutter < c1a
                if not (rows_clear or cols_clear):
                    raise ValueError(
                        f"pivots {name_a!r} and {name_b!r} overlap or sit "
                        f"within {gutter} cells: {box_a} vs {box_b}")

        for row, col in occupied:
            for name, (r1, c1, r2, c2) in self.targets:
                if r1 <= row <= r2 and c1 <= col <= c2:
                    raise ValueError(
                        f"cell (row={row}, col={col}) is inside pivot {name!r}; "
                        f"Excel will overwrite it on refresh")


# -------------------------------------------------------------------- validator

def validate(data: bytes) -> List[str]:
    """
    Check a produced .xlsx for the pivot defects Excel answers with a repair
    prompt. Returns a list of findings; empty means clean.

    Deliberately zip+regex rather than a schema validator: it needs no lxml and
    no ECMA-376 schema files, so it can run in the test suite on any machine.
    Pair it with `openpyxl.load_workbook()` on the same file, which raises
    KeyError on the cacheId defect and is the cheapest canary there is.
    """
    archive = ZipFile(BytesIO(data))
    names = set(archive.namelist())
    findings: List[str] = []

    declared_types = set(re.findall(
        r'PartName="([^"]+)"', archive.read("[Content_Types].xml").decode()))
    for name in sorted(names):
        if (name.endswith(".xml") and "/_rels/" not in name
                and ("pivot" in name or "/tables/" in name)):
            if "/" + name not in declared_types:
                findings.append(f"no content-type override for {name}")

    for rels in sorted(n for n in names if n.endswith(".rels")):
        for target in re.findall(r'Target="([^"]+)"', archive.read(rels).decode()):
            if target.startswith("/") and target[1:] not in names:
                findings.append(f"{rels} -> dangling target {target}")

    workbook_xml = archive.read("xl/workbook.xml").decode()
    declared_caches = dict(re.findall(
        r'<pivotCache cacheId="(\d+)" r:id="([^"]+)"', workbook_xml))

    for part in sorted(n for n in names
                       if re.match(r"xl/pivotTables/pivotTable\d+\.xml$", n)):
        xml = archive.read(part).decode()

        cache_id = re.search(r'cacheId="(\d+)"', xml)
        if cache_id is None:
            findings.append(f"{part} has no cacheId")
        elif cache_id.group(1) not in declared_caches:
            findings.append(
                f"{part} cacheId={cache_id.group(1)} missing from workbook "
                f"pivotCaches (the cache-dedupe defect)")

        field_count = int(re.search(r'<pivotFields count="(\d+)"', xml).group(1))
        for match in re.finditer(
                r'<(?:field|dataField)[^>]*?\b(?:x|fld)="(-?\d+)"', xml):
            index = int(match.group(1))
            if index >= field_count or (index < 0 and index != -2):
                findings.append(
                    f"{part} field index {index} out of range "
                    f"(pivotFields={field_count})")

        rels = part.replace("pivotTables/", "pivotTables/_rels/") + ".rels"
        if rels not in names:
            findings.append(f"{part} has no .rels, so no cache link")
            continue
        cache_part = re.search(
            r'Target="([^"]+)"', archive.read(rels).decode()).group(1).lstrip("/")
        cache_xml = archive.read(cache_part).decode()
        cache_field_count = int(
            re.search(r'<cacheFields count="(\d+)"', cache_xml).group(1))
        if cache_field_count != field_count:
            findings.append(
                f"{part} pivotFields={field_count} != "
                f"cacheFields={cache_field_count}")
        if 'saveData="0"' in cache_xml and 'refreshOnLoad="1"' not in cache_xml:
            findings.append(
                f"{cache_part} omits records but does not set refreshOnLoad, "
                f"so the pivot will render empty")

    return findings
