"""
Excel workbook assembly for the sanity-check report.

`core/reporter.py` owns the *data* (every `_build_*_df`); this module owns the
*workbook* — sheet layout, Excel Tables, widths, freeze panes, conditional
formatting and the native charts. Splitting them keeps the builders unit
testable without openpyxl and stops reporter.py from doubling in size.

No Qt imports, like the rest of `core/`.

Sheet layout is six visible tabs plus two hidden helper sheets:

    Overview        front matter, KPIs, and every disclosure
    Dashboard       the six native charts
    Findings        one row per check, severity-ranked
    All Details     every findings table as a titled block, with a jump list
    Pivots          native PivotTables            (wired in a later phase)
    Ticket Detail   one row per ticket            (wired in a later phase)
    PivotData       hidden — the aggregate cube the pivot caches read
    ChartData       hidden — one small immutable block per chart

Why charts read from a hidden ChartData sheet rather than the visible blocks:
the blocks are Excel Tables, and sorting or filtering a Table reorders the cells
a chart points at, so a user tidying the All Details sheet would silently
rewrite the dashboard. A dedicated block per chart also removes all offset
arithmetic against a stack of variable-height tables, which was the most likely
source of a chart quietly plotting the wrong data.

Colour choices are not free-hand: forms and palettes come from the project's
data-visualisation rules. The two ramps used here were checked with the palette
validator (monotone lightness, adjacent ΔL ≥ 0.06, light end clearing the
surface, single hue) before being written down.
"""
import logging
import os
import re
from typing import Any, Dict, List, NamedTuple, Optional, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from core.excel_pivot import (
    EXCEL_MAX_CELL_CHARS, EXCEL_MAX_ROWS, PivotBuilder, range_ref,
    register_table, table_name,
)

log = logging.getLogger("ticketaudit")

# --- Ink and surfaces -----------------------------------------------------
INK = "0B0B0B"
INK_SECONDARY = "52514E"
INK_MUTED = "898781"
GRIDLINE = "E1E0D9"
AXIS_LINE = "C3C2B7"
SURFACE = "FCFCFB"

# --- Series colours -------------------------------------------------------
SERIES_BLUE = "2A78D6"          # categorical slot 1
DEEMPHASIS = "C3C2B7"           # "not the story" grey
STATUS_CRITICAL = "D03B3B"
STATUS_GOOD = "0CA30C"
STATUS_WARNING = "FAB219"

# Ordinal ramps, darkest first. Both pass the palette validator on the light
# surface. Priority and quality tiers are *ordered* scales, so they get a single
# hue stepped by lightness rather than categorical hues — and definitely not a
# ramp keyed to bar length, which would double-encode magnitude as colour.
PRIORITY_RAMP = ("0D366B", "1C5CAB", "3987E5", "86B6EF")
TIER_RAMP = ("0D366B", "1C5CAB", "2A78D6", "5598E7", "86B6EF")

# --- Row tints for verdicts ----------------------------------------------
# "Not checked" is amber, never green: an unexamined dimension must not read as
# a clean one.
VERDICT_FILLS = {
    "Fail": PatternFill("solid", fgColor="FBE4E4"),
    "Not checked": PatternFill("solid", fgColor="FDF3DA"),
    "Review": PatternFill("solid", fgColor="FDF3DA"),
    "Pass": PatternFill("solid", fgColor="E4F4E4"),
}
HEADER_FILL = PatternFill("solid", fgColor="EDEBE5")

# The cube's measure column. Named once so the pivot's data field and the
# builder that produces it cannot disagree about which column holds the count.
CUBE_MEASURE = "Ticket count"
# Blank rows between stacked pivots. Excel refuses to refresh a pivot that would
# overlap another, so this is a correctness margin, not whitespace.
PIVOT_GAP = 2

# Above this many rows the detail leaves the workbook and goes out as a CSV
# sidecar (_write_detail_beside). Below it, the detail stays on its own sheet in
# the one file. The streamed workbook this used to mention no longer exists: the
# report needs native pivots, and write_only mode is a workbook-level flag that
# excludes add_pivot, so the two cannot share a file. core/extract.py is the
# only streamed writer in the codebase, and it needs no pivots.
DEFAULT_INLINE_MAX_ROWS = 100_000
# Excel's hard ceiling, minus the header row. Nothing in openpyxl enforces it:
# a Table ref past this is accepted silently and the file will not open.
DETAIL_MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1
# One progress message per this many rows, for both the detail CSV and the
# extract. Measured at 1.2s apart writing a 199k x 7 .xlsx (20 messages over
# 24.5s), which keeps a status bar moving without flooding the event loop.
DETAIL_PROGRESS_EVERY = 10_000

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 52
CHART_WIDTH_CM = 16
CHART_HEIGHT_CM = 8
CHART_ROW_PITCH = 18
BLOCK_GAP = 2
NULL_CHART_TOP_N = 15


class Block(NamedTuple):
    """Where a written table ended up, so charts and jump links can find it."""
    title: str
    table: str
    title_row: int
    header_row: int
    first_data_row: int
    last_row: int
    first_col: int
    last_col: int

    @property
    def row_count(self) -> int:
        return self.last_row - self.first_data_row + 1


def _sheet_title(worksheet, text: str, subtitle: Optional[str] = None) -> int:
    """Sheet heading. Returns the next free row."""
    worksheet["A1"] = text
    worksheet["A1"].font = Font(bold=True, size=14, color=INK)
    if subtitle:
        worksheet["A2"] = subtitle
        worksheet["A2"].font = Font(size=9, color=INK_MUTED)
        return 4
    return 3


def _autosize(worksheet, block: Block, frame: pd.DataFrame) -> None:
    """
    Width from the header and a sample of values.

    Sampled rather than exhaustive: a 300k-row detail sheet would otherwise pay
    a full string pass per column purely for cosmetics.
    """
    sample = frame.head(200)
    for offset, column in enumerate(frame.columns):
        widest = len(str(column))
        if not sample.empty:
            widest = max(widest, int(sample[column].astype(str).str.len().max() or 0))
        letter = get_column_letter(block.first_col + offset)
        worksheet.column_dimensions[letter].width = max(
            MIN_COL_WIDTH, min(MAX_COL_WIDTH, widest + 2))


def write_block(worksheet, frame: pd.DataFrame, *, title: str, start_row: int,
                used_names: set, first_col: int = 1,
                verdict_column: Optional[str] = None,
                autosize: bool = True) -> Block:
    """
    Write one titled table and register it as an Excel Table.

    Sanitises the frame itself, deliberately. Several of these blocks carry raw
    values out of the source data - sample values, state names, distribution
    values, recurring-issue text - so a single control character in a ticket
    description used to abort the entire export with IllegalCharacterError.
    Doing it here means every cell that reaches a worksheet is cleaned by
    construction, rather than depending on each call site remembering.

    verdict_column: tint each row by its Pass/Fail/Review/Not checked value.
    Applied in the write loop, keyed by column *name* — the old
    _apply_excel_formatting looked cells up by position and broke silently the
    first time a column was inserted before the one it meant.
    """
    frame, formula_columns = sanitize_for_excel(frame)
    title_cell = worksheet.cell(row=start_row, column=first_col, value=title)
    title_cell.font = Font(bold=True, size=11, color=INK)

    header_row = start_row + 1
    for offset, column in enumerate(frame.columns):
        cell = worksheet.cell(row=header_row, column=first_col + offset,
                              value=str(column))
        cell.font = Font(bold=True, color=INK)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    columns = list(frame.columns)
    verdict_index = (columns.index(verdict_column)
                     if verdict_column and verdict_column in columns else None)
    formula_indexes = {columns.index(name) for name in formula_columns
                       if name in columns}

    for row_offset, row in enumerate(
            frame.itertuples(index=False, name=None), start=1):
        tint = None
        if verdict_index is not None:
            tint = VERDICT_FILLS.get(str(row[verdict_index]))
        for col_offset, value in enumerate(row):
            cell = worksheet.cell(row=header_row + row_offset,
                                 column=first_col + col_offset,
                                 value=_excel_safe(value))
            if col_offset in formula_indexes and cell.data_type == "f":
                cell.data_type = "s"      # text, not a formula
            if tint is not None:
                cell.fill = tint

    last_row = header_row + max(len(frame), 1)
    block = Block(
        title=title, table=table_name(title, used_names),
        title_row=start_row, header_row=header_row,
        first_data_row=header_row + 1, last_row=last_row,
        first_col=first_col, last_col=first_col + len(frame.columns) - 1,
    )
    register_table(worksheet, block.table,
                   range_ref(block.header_row, block.first_col,
                             block.last_row, block.last_col))
    if autosize:
        _autosize(worksheet, block, frame)
    return block


# Characters openpyxl refuses outright (IllegalCharacterError). Free-text ITSM
# descriptions pasted out of email carry them, and one bad character at row
# 250,000 destroys a minute of work.
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean_cell_text(value: str, max_chars: Optional[int]) -> str:
    cleaned = _CONTROL_CHARS.sub("", value)
    if max_chars is not None and len(cleaned) > max_chars:
        # openpyxl accepts an over-long string; Excel then offers to repair the
        # file. Truncate with a marker so the loss is visible in the cell.
        # CSV has no such limit, so it passes max_chars=None and keeps the text.
        cleaned = cleaned[:max_chars - 1] + "…"
    return cleaned


def _sanitize_series(series: pd.Series, max_chars: Optional[int]) -> pd.Series:
    """Make one column safe to hand to openpyxl. See sanitize_for_excel."""
    if isinstance(series.dtype, pd.CategoricalDtype):
        categories = series.cat.categories
        if all(isinstance(category, str) for category in categories):
            cleaned = [_clean_cell_text(c, max_chars) for c in categories]
            if len(set(cleaned)) == len(cleaned):
                # Clean the categories, not the rows: a few dozen values instead
                # of hundreds of thousands. rename_categories raises on a
                # collision, hence the length guard and the fallback below.
                series = series.cat.rename_categories(cleaned)
            else:
                series = series.astype(object).map(
                    lambda v: _clean_cell_text(v, max_chars)
                    if isinstance(v, str) else v)
        return series.astype(object).where(series.notna(), None)

    if isinstance(series.dtype, pd.DatetimeTZDtype):
        # Excel has no concept of a timezone; keep the wall-clock time and drop
        # the offset rather than letting the writer decide.
        series = series.dt.tz_localize(None)
        return series.astype(object).where(series.notna(), None)

    if pd.api.types.is_bool_dtype(series):
        # np.bool_ is written as numeric 1/0, not TRUE/FALSE.
        return series.astype(object).map(
            {True: "Yes", False: "No"}).where(series.notna(), None)

    if pd.api.types.is_numeric_dtype(series) or \
            pd.api.types.is_datetime64_any_dtype(series):
        # pandas 3's Int64/Float64 hold pd.NA, which the writer rejects.
        if series.isna().any():
            return series.astype(object).where(series.notna(), None)
        return series

    cleaned = series.astype(object).map(
        lambda v: _clean_cell_text(v, max_chars) if isinstance(v, str) else v)
    return cleaned.where(series.notna(), None)


def sanitize_for_excel(frame: pd.DataFrame, *,
                       max_chars: Optional[int] = EXCEL_MAX_CELL_CHARS):
    """
    Make a frame safe for openpyxl. Returns (frame, formula_columns).

    Every case here was reproduced, and most fail *silently*:

    - **pd.NA aborts the write with no error.** It raises inside openpyxl's row
      writer, which in write-only mode kills the row generator; every later
      append raises StopIteration and `save()` then emits a truncated workbook
      that openpyxl itself cannot reopen. pandas 3's str, Int64 and boolean
      dtypes all produce pd.NA, and optimize_dtypes leaves plenty around.
    - Control characters raise IllegalCharacterError.
    - Strings over 32,767 characters are accepted here and rejected by Excel.
    - np.bool_ writes as 1/0 rather than TRUE/FALSE.
    - Timezone-aware datetimes are accepted but Excel has no timezone concept.

    Category dtypes are cleaned via their *categories*, not their rows — a few
    dozen values instead of hundreds of thousands — and are deliberately not
    converted to str, which roughly halves the frame's memory.

    formula_columns names the columns holding a string that starts with "=":
    openpyxl types those as formulas, so the writer must force them back to
    text. Only "=" matters; +, - and @ are a CSV-import concern, not an xlsx one.
    """
    columns: Dict[str, pd.Series] = {}
    formula_columns = set()
    for name, series in frame.items():
        if series.dtype == object or isinstance(series.dtype, pd.CategoricalDtype) \
                or pd.api.types.is_string_dtype(series):
            as_text = series.astype(object)
            starts = as_text.map(lambda v: isinstance(v, str) and v.startswith("="))
            if bool(starts.any()):
                formula_columns.add(name)
        columns[name] = _sanitize_series(series, max_chars)
    return pd.DataFrame(columns, index=frame.index), formula_columns


def _excel_safe(value: Any) -> Any:
    """
    Minimal write-time guard.

    pd.NA raises inside openpyxl's row writer and, in write-only mode, kills the
    row generator so `save()` emits a truncated workbook with no error at all.
    pandas 3's str/Int64/boolean dtypes all produce it. The full sanitiser
    (control characters, 32k strings, formula-leading text, tz-aware datetimes)
    lands with the streaming writer; this covers the normal-mode path.
    """
    if value is pd.NA or value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (pd.Timestamp,)) and value.tzinfo is not None:
        return value.tz_localize(None)
    return value


def write_detail_csv(path: str, frame: pd.DataFrame, *, progress=None) -> int:
    """
    Write the per-ticket detail as CSV and return the row count.

    CSV rather than a streamed .xlsx above the inline threshold, on measured
    grounds: openpyxl writes about 59k text cells/s and only 18k date cells/s,
    so a 300k x 27 detail sheet is roughly 8 million cells and four minutes.
    The same data is a 70 MB CSV in 9.6s. For a sheet whose whole purpose is to
    be sliced, that trade is worth making - Excel opens it directly, and
    `Data > From Text/CSV` imports it as a real Table for pivoting.

    utf-8-sig, not plain utf-8: without the BOM Excel mis-decodes non-ASCII on
    a double-click, which would mangle exactly the non-English descriptions this
    tool exists to find. core/loader.py reads that encoding first for the same
    reason.

    max_chars=None because CSV has no 32,767-character cell limit; truncating
    here would lose text for no reason. Control characters are still stripped -
    they break parsers - and booleans still render as Yes/No so the CSV and the
    inline sheet describe a ticket the same way.
    """
    if progress:
        progress(f"Writing ticket detail CSV: {len(frame):,} rows…")
    sanitised, _formula_columns = sanitize_for_excel(frame, max_chars=None)
    sanitised.to_csv(path, index=False, encoding="utf-8-sig")
    return len(sanitised)


class ExcelReportBuilder:
    """
    Assembles the workbook from a ReportGenerator's builders.

    Takes the reporter rather than a pile of frames so each sheet can decide
    what it needs, and so a builder that returns nothing simply yields no block
    — which is how "not checked" propagates into the layout.
    """

    def __init__(self, reporter, *, enable_pivots: bool = True,
                 inline_max_rows: int = DEFAULT_INLINE_MAX_ROWS,
                 detail_row_limit: int = DETAIL_MAX_DATA_ROWS,
                 progress=None):
        self.reporter = reporter
        self.analyzer = reporter.analyzer
        self.enable_pivots = enable_pivots
        self.inline_max_rows = inline_max_rows
        self.detail_row_limit = detail_row_limit
        self.progress = progress or (lambda _message: None)
        self.pivot_count = 0
        # Detail-sheet outcome, reported back so the caller can tell the user
        # where the rows went and what was left out.
        self.detail_path: Optional[str] = None
        self.detail_rows_written = 0
        self.detail_truncated = False
        self.used_names: set = set()
        self.blocks: Dict[str, Block] = {}
        self.chart_blocks: Dict[str, Block] = {}
        # The exact frame written for each chart. Charts that colour individual
        # points must read this, not rebuild the frame — a second derivation can
        # filter differently and silently tint the wrong bars.
        self.chart_frames: Dict[str, pd.DataFrame] = {}
        self.chart_sheet = None          # set by _write_chart_data
        self.omissions: List[str] = []
        # Several frames are needed by more than one sheet - column health by
        # three of them - and none of the builders memoise. Without this the
        # findings register alone was rebuilt twice, costing ~13s per call at
        # 300k rows.
        self._frames: Dict[str, pd.DataFrame] = {}

    def frame(self, key: str, build) -> pd.DataFrame:
        """Build a reporter frame once per workbook."""
        if key not in self._frames:
            self._frames[key] = build()
        return self._frames[key]

    # ------------------------------------------------------------------ build

    def build(self, path: str) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)

        # Created up front so the tab order is the reading order, regardless of
        # which sheet is populated when.
        overview = workbook.create_sheet("Overview")
        dashboard = workbook.create_sheet("Dashboard")
        findings = workbook.create_sheet("Findings")
        details = workbook.create_sheet("All Details")
        pivots = workbook.create_sheet("Pivots")

        self.progress("Building findings…")
        self._write_findings(findings)
        self.progress("Building detail tables…")
        self._write_all_details(details)
        self._write_chart_data(workbook)
        self._write_pivot_data(workbook)
        self.progress("Building pivots…")
        self._write_pivots(pivots)
        self.progress("Building charts…")
        self._add_charts(dashboard)
        self._write_detail(workbook, path)
        # Last, so it can report what every other sheet did or could not do.
        self._write_overview(overview)

        workbook.active = 0
        self.progress("Finalising the workbook…")
        workbook.save(path)
        return workbook

    # ----------------------------------------------------------- detail sheet

    def _write_detail(self, workbook, report_path: str) -> None:
        """
        One row per ticket, either inline or in its own streamed workbook.

        The split is forced, not a preference: native pivots need normal mode
        (add_pivot does not exist on a write-only sheet, and write_only is a
        workbook-level flag), while a 300k x 43 detail sheet in normal mode is
        ~13 million Cell objects and about 2.4 GB. So a large detail sheet moves
        to its own streamed file and the pivots stay with the report.
        """
        self.progress("Building ticket detail…")
        try:
            detail = self.reporter._build_ticket_detail_df()
        except Exception:
            log.exception("ticket detail could not be built")
            self.omissions.append(
                "Ticket Detail: could not be built - see the log.")
            return

        if detail.empty:
            self.omissions.append("Ticket Detail: the file has no rows.")
            return

        total_rows = len(detail)
        if total_rows > self.detail_row_limit:
            # Excel's own ceiling. Truncating is the only option, so the only
            # question is whether the reader is told - and every number they
            # derive from a silently shortened sheet would be wrong.
            detail = detail.head(self.detail_row_limit)
            self.detail_truncated = True
            self.omissions.append(
                f"Ticket Detail is TRUNCATED: {self.detail_row_limit:,} of "
                f"{total_rows:,} rows written. An Excel worksheet cannot hold "
                f"more than {EXCEL_MAX_ROWS:,} rows including the header.")

        title = ("Ticket Detail (TRUNCATED)" if self.detail_truncated
                 else "Ticket Detail")

        if len(detail) <= self.inline_max_rows:
            self._write_detail_inline(workbook, detail, title)
        else:
            self._write_detail_beside(report_path, detail, title, total_rows)

    def _write_detail_inline(self, workbook, detail: pd.DataFrame,
                             title: str) -> None:
        worksheet = workbook.create_sheet(title)
        # write_block sanitises; no need to do it twice.
        block = write_block(worksheet, detail, title=title, start_row=1,
                            used_names=self.used_names)
        self.blocks["Ticket Detail"] = block
        worksheet.freeze_panes = worksheet.cell(
            row=block.first_data_row, column=1).coordinate
        self.detail_rows_written = len(detail)

    def _write_detail_beside(self, report_path: str, detail: pd.DataFrame,
                             title: str, total_rows: int) -> None:
        """
        Large detail goes out as a CSV sidecar.

        Two constraints meet here. Native pivots need a non-streamed sheet, so
        the detail cannot share this workbook once it is large. And a streamed
        .xlsx of that size costs minutes - measured at roughly four for
        300k x 27 - against 9.6s for the same data as CSV.
        """
        base, _extension = os.path.splitext(report_path)
        suffix = "_TicketDetail_TRUNCATED" if self.detail_truncated \
            else "_TicketDetail"
        self.detail_path = f"{base}{suffix}.csv"
        try:
            self.detail_rows_written = write_detail_csv(
                self.detail_path, detail, progress=self.progress)
        except Exception:
            log.exception("the detail CSV could not be written")
            self.detail_path = None
            self.omissions.append(
                "Ticket Detail: the companion CSV could not be written - "
                "see the log.")
            return

        self.omissions.append(
            f"Ticket Detail ({self.detail_rows_written:,} rows) is a separate "
            f"CSV: {os.path.basename(self.detail_path)}. Too large for a sheet "
            f"beside the PivotTables, and CSV writes in seconds where .xlsx "
            f"takes minutes. To pivot it: Data > From Text/CSV > Load.")

    # ----------------------------------------------------------------- sheets

    def _write_findings(self, worksheet) -> None:
        frame = self.frame("findings", self.reporter._build_findings_df)
        if frame.empty:
            self.omissions.append("Findings: no checks produced a result")
            return
        row = _sheet_title(
            worksheet, "Findings",
            "Every check, worst first. 'Not checked' means the check could not "
            "run — it is not a pass.")
        self.blocks["Findings"] = write_block(
            worksheet, frame, title="All findings", start_row=row,
            used_names=self.used_names, verdict_column="Result")
        worksheet.freeze_panes = worksheet.cell(
            row=self.blocks["Findings"].first_data_row, column=1).coordinate
        worksheet.sheet_view.showGridLines = False

    def _write_all_details(self, worksheet) -> None:
        """
        Every findings table as a titled block, with a jump list on top.

        A jump list is not decoration: ten stacked tables is a long scroll, and
        without it finding one means hunting. Each block is its own Table so it
        sorts and filters independently.
        """
        reporter = self.reporter
        quality = self.analyzer.check_description_quality(
            reporter.desc_min_length, reporter.desc_max_length)

        planned = [
            ("Required columns", reporter._build_column_check_df, None),
            ("Column health", lambda: self.frame("health", reporter._build_column_health_df), None),
            ("Logic checks", reporter._build_logic_checks_df, "Result"),
            ("State vocabulary", reporter._build_state_values_df, None),
            ("Description quality",
             lambda: reporter._build_description_quality_df(quality), None),
            ("Description quality tiers", lambda: self.frame("tiers", reporter._build_quality_tiers_df), None),
            ("Recurring issues",
             lambda: reporter._build_recurring_issues_df(quality), None),
            ("Language summary", reporter._build_language_summary_df, None),
            ("Language detail", reporter._build_language_detail_df, None),
            ("Distributions", reporter._build_distributions_df, None),
            ("Monthly inflow", lambda: self.frame("inflow", reporter._build_monthly_inflow_df), None),
        ]

        # Leave room for the jump list, whose length is known up front.
        row = _sheet_title(worksheet, "All Details",
                           "Each table below filters and sorts on its own.")
        toc_row = row
        row += len(planned) + 2

        toc_entries: List[tuple] = []
        for title, build, verdict in planned:
            try:
                frame = build()
            except Exception:
                log.exception("report block %r failed to build", title)
                frame = pd.DataFrame()

            if frame.empty:
                # A block that could not be produced says so in place. An
                # omitted sheet is invisible; a stated gap is not.
                worksheet.cell(row=row, column=1, value=title).font = Font(
                    bold=True, size=11, color=INK)
                note = worksheet.cell(
                    row=row + 1, column=1,
                    value="Not checked — the data needed for this table was "
                          "not available in this file.")
                note.font = Font(italic=True, color=INK_SECONDARY)
                toc_entries.append((title, row, "Not checked"))
                self.omissions.append(f"{title}: not available for this file")
                row += 3
                continue

            block = write_block(worksheet, frame, title=title, start_row=row,
                                used_names=self.used_names,
                                verdict_column=verdict)
            self.blocks[title] = block
            toc_entries.append((title, row, f"{len(frame):,} rows"))
            row = block.last_row + 1 + BLOCK_GAP

        self._write_jump_list(worksheet, toc_row, toc_entries)
        self._add_null_data_bars(worksheet)
        worksheet.sheet_view.showGridLines = False

    def _write_jump_list(self, worksheet, start_row: int,
                         entries: Sequence[tuple]) -> None:
        heading = worksheet.cell(row=start_row, column=1, value="Jump to")
        heading.font = Font(bold=True, size=11, color=INK)
        for offset, (title, target_row, note) in enumerate(entries, start=1):
            cell = worksheet.cell(row=start_row + offset, column=1, value=title)
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate,
                location=f"'{worksheet.title}'!A{target_row}")
            cell.font = Font(color=SERIES_BLUE, underline="single")
            worksheet.cell(row=start_row + offset, column=2, value=note).font = \
                Font(color=INK_MUTED, size=9)

    def _add_null_data_bars(self, worksheet) -> None:
        """In-cell bars on Null %, so the worst columns are visible at a glance."""
        block = self.blocks.get("Column health")
        if block is None:
            return
        health = self.frame("health", self.reporter._build_column_health_df)
        if "Null %" not in health.columns:
            return
        index = list(health.columns).index("Null %")
        letter = get_column_letter(block.first_col + index)
        worksheet.conditional_formatting.add(
            f"{letter}{block.first_data_row}:{letter}{block.last_row}",
            DataBarRule(start_type="num", start_value=0,
                        end_type="num", end_value=100, color=SERIES_BLUE))

    def _write_chart_data(self, workbook) -> None:
        """
        One small immutable block per chart, on a hidden sheet.

        Deliberately separate from the visible tables: those are Excel Tables,
        and a user sorting one would reorder the cells a chart reads, silently
        rewriting the dashboard.
        """
        worksheet = workbook.create_sheet("ChartData")
        self.chart_sheet = worksheet
        reporter = self.reporter
        analyzer = self.analyzer

        health = self.frame("health", reporter._build_column_health_df)
        if not health.empty and "Null %" in health.columns:
            # Only columns that actually have nulls. A "worst 15" of mostly
            # zeros is a chart of nothing, and each zero still carries a data
            # label — the reader ends up scanning eight "0"s for the two rows
            # that matter. When no column has nulls the chart is omitted and
            # Overview says why, which is the more useful statement anyway.
            with_nulls = health[health["Null %"] > 0]
            health_frame = (with_nulls.sort_values("Null %", ascending=False)
                                      .head(NULL_CHART_TOP_N)
                                      [["Column", "Null %", "Above Threshold"]]
                                      .reset_index(drop=True))
        else:
            health_frame = pd.DataFrame()

        sources = [
            ("Monthly inflow", self.frame("inflow", reporter._build_monthly_inflow_df)),
            ("Nulls by column", health_frame),
            ("Priority mix",
             reporter._build_dimension_mix_df(analyzer.priority_column)),
            ("State mix",
             reporter._build_dimension_mix_df(analyzer.state_column)),
            ("Quality tiers", self.frame("tiers", reporter._build_quality_tiers_df)),
            ("Language mix", reporter._build_language_mix_df()),
        ]

        row = 1
        for title, frame in sources:
            if frame.empty:
                self.omissions.append(f"Chart '{title}': no data to plot")
                continue
            # "(chart)" in the title so the Table name stays self-describing:
            # several of these duplicate a visible block, and tbl_X_2 tells the
            # reader nothing about which is which.
            block = write_block(worksheet, frame, title=f"{title} (chart)",
                                start_row=row, used_names=self.used_names,
                                autosize=False)
            self.chart_blocks[title] = block
            self.chart_frames[title] = frame
            row = block.last_row + 1 + BLOCK_GAP

        worksheet.sheet_state = "hidden"

    def _write_pivot_data(self, workbook) -> None:
        """The aggregate cube the native pivots read from. Hidden, not secret —
        it is named on Overview."""
        frame = self.frame("cube", self.reporter._build_pivot_cube_df)
        # Which columns the cube's caps excluded, and why. A reader who expected
        # to pivot by assignment group and cannot must be able to tell "the tool
        # decided against it" from "the tool broke".
        self.omissions.extend(
            getattr(self.reporter, "pivot_cube_notes", None) or [])
        if frame.empty:
            return
        worksheet = workbook.create_sheet("PivotData")
        block = write_block(worksheet, frame, title="Pivot cube", start_row=1,
                            used_names=self.used_names, autosize=False)
        self.blocks["PivotData"] = block
        worksheet.sheet_state = "hidden"

    def _write_pivots(self, worksheet) -> None:
        """
        One native PivotTable per cube dimension, stacked down the sheet.

        Wrapped whole: a pivot is the one part of this workbook assembled from
        raw OOXML, and a malformed one makes Excel offer to repair the file,
        silently dropping every pivot. Losing the pivots is survivable; losing
        the report is not. So any failure falls back to Tables-only and says so
        on Overview rather than propagating.
        """
        row = _sheet_title(
            worksheet, "Pivots",
            "Live PivotTables over the aggregated cube. Drag fields to re-slice; "
            "right-click > Refresh after editing the data.")

        if not self.enable_pivots:
            self._pivot_fallback(worksheet, row,
                                 "Native pivots are switched off in settings.")
            return

        cube_block = self.blocks.get("PivotData")
        if cube_block is None:
            self._pivot_fallback(
                worksheet, row,
                "No column had few enough distinct values to pivot on.")
            return

        try:
            self._build_pivots(worksheet, row, cube_block)
        except Exception:
            log.exception("native pivot construction failed")
            self.omissions.append(
                "Pivots: could not be built - see the log. Every table in this "
                "workbook is still an Excel Table, so Insert > PivotTable works.")
            self._pivot_fallback(
                worksheet, row,
                "Native pivots could not be built for this file.")

    def _build_pivots(self, worksheet, row: int, cube_block: Block) -> None:
        cube = self.frame("cube", self.reporter._build_pivot_cube_df)
        columns = list(cube.columns)
        measure = columns.index(CUBE_MEASURE)
        dimensions = [i for i, name in enumerate(columns) if i != measure]

        pivots = PivotBuilder(
            columns,
            list(cube.itertuples(index=False, name=None)),
            axis_fields=dimensions,
            cache_id=1,
            # Named Table rather than a raw range: it survives a sheet rename
            # and auto-expands if rows are added.
            source_name=cube_block.table,
            save_records=True,
        )

        titles: List[tuple] = []
        for index in dimensions:
            label = columns[index]
            title_cell = worksheet.cell(row=row, column=1,
                                        value=f"Tickets by {label}")
            title_cell.font = Font(bold=True, size=11, color=INK)
            titles.append((row, 1))

            pivots.add(worksheet, table_name(f"pt_{label}", self.used_names),
                       f"A{row + 1}", row_field=index, data_field=measure,
                       subtotal="sum")
            row = pivots.targets[-1][1][2] + 1 + PIVOT_GAP

        # Overlaps are not corruption but a modal alert on refresh, and a title
        # inside a pivot's range gets overwritten. Both are decidable here.
        pivots.assert_disjoint(gutter=1, occupied=titles)
        self.pivot_count = len(pivots.targets)

    def _pivot_fallback(self, worksheet, row: int, reason: str) -> None:
        """Say why the sheet is empty, and what to do instead."""
        note = worksheet.cell(row=row, column=1, value=reason)
        note.font = Font(italic=True, color=INK_SECONDARY)
        follow_up = worksheet.cell(
            row=row + 1, column=1,
            value="Every table in this workbook is a named Excel Table: click any "
                  "cell inside one, then Insert > PivotTable > OK.")
        follow_up.font = Font(color=INK_SECONDARY)
        worksheet.column_dimensions["A"].width = 90

    def _write_overview(self, worksheet) -> None:
        reporter = self.reporter
        analyzer = self.analyzer
        summary = reporter._build_summary_df()

        row = _sheet_title(
            worksheet, "TicketAudit — Data Quality Report",
            f"{reporter.filepath or 'data'} · "
            f"{analyzer.total_rows:,} rows × {analyzer.total_columns} columns · "
            f"generated {pd.Timestamp.now():%Y-%m-%d %H:%M}")

        findings = self.frame("findings", reporter._build_findings_df)
        kpis = self._kpis(findings)
        for offset, (label, value) in enumerate(kpis):
            value_cell = worksheet.cell(row=row, column=1 + offset * 2, value=value)
            value_cell.font = Font(bold=True, size=16, color=INK)
            label_cell = worksheet.cell(row=row + 1, column=1 + offset * 2,
                                        value=label)
            label_cell.font = Font(size=9, color=INK_MUTED)
        row += 3

        # A truncated pivot source under-reports every number derived from it,
        # so the count lands in Summary as well as in the notes and the tab name.
        detail_note = (
            f"{self.detail_rows_written:,} of {self.analyzer.total_rows:,} — TRUNCATED"
            if self.detail_truncated
            else f"{self.detail_rows_written:,}" if self.detail_rows_written
            else "not written")
        summary = pd.concat([summary, pd.DataFrame(
            [{"Metric": "Ticket Detail rows written", "Value": detail_note}])],
            ignore_index=True)

        self.blocks["Summary"] = write_block(
            worksheet, summary, title="Summary", start_row=row,
            used_names=self.used_names)
        row = self.blocks["Summary"].last_row + 1 + BLOCK_GAP

        thresholds = pd.DataFrame([
            {"Setting": "Null threshold", "Value": f"{reporter.null_threshold}%"},
            {"Setting": "Description minimum length",
             "Value": f"{reporter.desc_min_length} characters"},
            {"Setting": "Description maximum length",
             "Value": f"{reporter.desc_max_length} characters"},
            {"Setting": "Language detection",
             "Value": (reporter._language_results()[0].get("mode_label", "n/a")
                       if reporter._language_results() else "not run")},
        ])
        self.blocks["Settings"] = write_block(
            worksheet, thresholds, title="Settings used", start_row=row,
            used_names=self.used_names)
        row = self.blocks["Settings"].last_row + 1 + BLOCK_GAP

        # Everything the workbook does NOT contain, and why. A reader must never
        # have to guess whether a missing table means "clean" or "never ran".
        notes = list(self.omissions) + list(
            getattr(reporter, "export_disclosures", []) or [])
        notes.append("PivotData and ChartData are hidden helper sheets: the "
                     "pivot cache and the chart sources.")
        if self.pivot_count:
            notes.append(
                f"Pivots holds {self.pivot_count} live PivotTable(s) over the "
                f"aggregated cube. They refresh on open; right-click > Refresh "
                f"after editing data.")
        disclosure = pd.DataFrame(
            [{"Note": note} for note in notes]
            or [{"Note": "Every check ran; nothing was omitted."}])
        write_block(worksheet, disclosure, title="What was and was not checked",
                    start_row=row, used_names=self.used_names)

        worksheet.sheet_view.showGridLines = False

    def _kpis(self, findings: pd.DataFrame) -> List[tuple]:
        analyzer = self.analyzer
        failed = 0 if findings.empty else int((findings["Result"] == "Fail").sum())
        unchecked = (0 if findings.empty
                     else int((findings["Result"] == "Not checked").sum()))
        dupes = analyzer.check_duplicates()
        missing = len(analyzer.get_missing_columns())
        return [
            (f"{analyzer.total_rows:,}", "rows"),
            (str(failed), "checks failed"),
            (str(unchecked), "not checked"),
            (f"{dupes.get('count', 0):,}", "duplicate rows"),
            (str(missing), "required columns missing"),
        ]

    # ----------------------------------------------------------------- charts

    def _add_charts(self, worksheet) -> None:
        _sheet_title(worksheet, "Data Quality Dashboard",
                     "Charts read from a hidden ChartData sheet, so sorting a "
                     "table elsewhere cannot change them.")

        plan = [
            ("Monthly inflow", self._chart_inflow, "A4"),
            ("Nulls by column", self._chart_nulls, "J4"),
            ("Priority mix", self._chart_priority, f"A{4 + CHART_ROW_PITCH}"),
            ("State mix", self._chart_state, f"J{4 + CHART_ROW_PITCH}"),
            ("Quality tiers", self._chart_quality, f"A{4 + CHART_ROW_PITCH * 2}"),
            ("Language mix", self._chart_language, f"J{4 + CHART_ROW_PITCH * 2}"),
        ]

        for name, factory, anchor in plan:
            block = self.chart_blocks.get(name)
            if block is None:
                continue        # already recorded in self.omissions
            try:
                chart = factory(block)
            except Exception:
                log.exception("chart %r failed to build", name)
                self.omissions.append(f"Chart '{name}': failed to build")
                continue
            chart.width = CHART_WIDTH_CM
            chart.height = CHART_HEIGHT_CM
            worksheet.add_chart(chart, anchor)

        worksheet.sheet_view.showGridLines = False

    def _source(self, block: Block, column_offset: int, *, with_header: bool):
        first = block.header_row if with_header else block.first_data_row
        return Reference(self.chart_sheet, min_col=block.first_col + column_offset,
                         min_row=first, max_row=block.last_row)

    def _categories(self, block: Block, column_offset: int = 0):
        return Reference(self.chart_sheet,
                         min_col=block.first_col + column_offset,
                         min_row=block.first_data_row, max_row=block.last_row)

    def _style(self, chart, *, title: str, legend: bool = False,
               from_zero: bool = True) -> None:
        """
        Uniform chrome: recessive hairline grid, muted ticks, no 3-D, no shadow.

        A single-series chart gets no legend box — its title names the series, and
        a legend would just repeat it.

        from_zero pins the value axis to 0. Excel auto-scales it otherwise, and
        on a magnitude comparison that is actively misleading: with values 20,
        20 and 17 it chose a baseline of 15.5, rendering 17 as a sliver of 20
        when it is 85% of it. Bar length has to stay proportional to value.
        Note openpyxl calls the value axis y_axis even for horizontal bars.
        """
        chart.title = title
        chart.style = None
        chart.roundedCorners = False
        if from_zero:
            chart.y_axis.scaling.min = 0
        if not legend:
            chart.legend = None
        else:
            chart.legend.position = "b"
            chart.legend.overlay = False

        for axis in (chart.x_axis, chart.y_axis):
            # openpyxl omits axes entirely unless delete is explicitly False.
            axis.delete = False
            axis.majorTickMark = "none"
            axis.minorTickMark = "none"
            axis.spPr = GraphicalProperties()
            axis.spPr.line = LineProperties(solidFill=AXIS_LINE, w=9525)
        if chart.y_axis.majorGridlines is not None:
            chart.y_axis.majorGridlines.spPr = GraphicalProperties(
                ln=LineProperties(solidFill=GRIDLINE, w=9525))

    # -- individual charts --------------------------------------------------

    def _chart_inflow(self, block: Block) -> LineChart:
        chart = LineChart()
        chart.add_data(self._source(block, 1, with_header=True), titles_from_data=True)
        chart.set_categories(self._categories(block))
        series = chart.series[0]
        series.graphicalProperties.line = LineProperties(
            solidFill=SERIES_BLUE, w=25400)          # 2pt
        series.smooth = False
        self._style(chart, title="Tickets created per month")
        return chart

    def _chart_nulls(self, block: Block) -> BarChart:
        """
        Emphasis, not a flat series: the columns breaching the threshold are the
        story, so they carry the critical status colour and the rest recede.
        Never colour alone — the bars are labelled and the sheet has an
        'Above Threshold' column.
        """
        chart = BarChart()
        chart.type = "bar"
        chart.add_data(self._source(block, 1, with_header=True), titles_from_data=True)
        chart.set_categories(self._categories(block))
        chart.gapWidth = 60

        written = self.chart_frames["Nulls by column"]
        series = chart.series[0]
        series.graphicalProperties.solidFill = DEEMPHASIS
        for index, above in enumerate(written["Above Threshold"]):
            if above == "Yes":
                point = DataPoint(idx=index)
                point.graphicalProperties = GraphicalProperties(
                    solidFill=STATUS_CRITICAL)
                series.dPt.append(point)
        chart.dataLabels = self._labels()
        # Only claim "worst N" when the list was actually truncated.
        shown = len(written)
        title = ("Null % by column"
                 if shown < NULL_CHART_TOP_N
                 else f"Null % by column (worst {NULL_CHART_TOP_N})")
        self._style(chart, title=title)
        return chart

    def _chart_priority(self, block: Block) -> BarChart:
        chart = BarChart()
        chart.type = "col"
        chart.add_data(self._source(block, 1, with_header=True), titles_from_data=True)
        chart.set_categories(self._categories(block))
        chart.gapWidth = 60
        # Ordered scale, so an ordinal ramp rather than categorical hues.
        self._colour_points(chart.series[0], block.row_count, PRIORITY_RAMP)
        chart.dataLabels = self._labels()
        self._style(chart, title="Tickets by priority")
        return chart

    def _chart_state(self, block: Block) -> BarChart:
        chart = BarChart()
        chart.type = "bar"
        chart.add_data(self._source(block, 1, with_header=True), titles_from_data=True)
        chart.set_categories(self._categories(block))
        chart.gapWidth = 60
        # Nominal categories: one series, one colour. A value ramp here would
        # double-encode bar length as hue.
        chart.series[0].graphicalProperties.solidFill = SERIES_BLUE
        chart.dataLabels = self._labels()
        self._style(chart, title="Tickets by state")
        return chart

    def _chart_quality(self, block: Block) -> BarChart:
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "percentStacked"
        chart.overlap = 100
        # One series per tier: columns 1..5 after the Field name.
        data = Reference(self.chart_sheet, min_col=block.first_col + 1,
                         max_col=block.last_col, min_row=block.header_row,
                         max_row=block.last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(self._categories(block))
        chart.gapWidth = 60
        for series, colour in zip(chart.series, TIER_RAMP):
            series.graphicalProperties.solidFill = colour
        self._style(chart, title="Description quality mix", legend=True)
        return chart

    def _chart_language(self, block: Block) -> BarChart:
        chart = BarChart()
        chart.type = "bar"
        chart.add_data(self._source(block, 1, with_header=True), titles_from_data=True)
        chart.set_categories(self._categories(block))
        chart.gapWidth = 60
        chart.series[0].graphicalProperties.solidFill = SERIES_BLUE
        chart.dataLabels = self._labels()
        self._style(chart, title="Non-English rows by language")
        return chart

    # -- chart helpers ------------------------------------------------------

    @staticmethod
    def _labels():
        from openpyxl.chart.label import DataLabelList
        labels = DataLabelList()
        labels.showVal = True
        labels.showSerName = False
        labels.showCatName = False
        labels.showLegendKey = False
        return labels

    @staticmethod
    def _colour_points(series, count: int, ramp: Sequence[str]) -> None:
        """
        Per-point colours for an ordinal ramp.

        DataPoint's constructor takes `spPr`, not `graphicalProperties`, so the
        properties have to be attached after construction.
        """
        for index in range(count):
            point = DataPoint(idx=index)
            point.graphicalProperties = GraphicalProperties(
                solidFill=ramp[min(index, len(ramp) - 1)])
            series.dPt.append(point)


def build_workbook(reporter, path: str, *, enable_pivots: bool = True,
                   inline_max_rows: int = DEFAULT_INLINE_MAX_ROWS,
                   detail_row_limit: int = DETAIL_MAX_DATA_ROWS,
                   progress=None) -> ExcelReportBuilder:
    """Assemble and save the report workbook. Returns the builder for its notes."""
    builder = ExcelReportBuilder(
        reporter, enable_pivots=enable_pivots,
        inline_max_rows=inline_max_rows, detail_row_limit=detail_row_limit,
        progress=progress)
    builder.build(path)
    return builder
