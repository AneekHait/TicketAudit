"""
Tabular file loading for TicketAudit.

Keeps all file-format knowledge in one GUI-independent place so it can be unit
tested. The GUI only needs two entry points:

    list_sheets(path) -> ["Sheet1", ...]   # [] means "single table, no sheets"
    read_table(path, sheet_name=None) -> pd.DataFrame

Supported: .xlsx (calamine, falling back to openpyxl), .xls (calamine/xlrd), and
delimited text (.csv/.tsv/.txt) with automatic encoding and delimiter detection.
"""
import csv
import functools
import logging
import os
from typing import List, Optional

import pandas as pd

log = logging.getLogger("ticketaudit.loader")

# --- Format detection ---

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
LEGACY_EXCEL_EXTENSIONS = {".xls"}
TEXT_EXTENSIONS = {".csv", ".tsv", ".txt", ".dat"}

# Magic bytes, used when the extension is missing or lying
_ZIP_MAGIC = b"PK\x03\x04"        # .xlsx is a zip container
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0"  # legacy .xls

# A zip keeps its index - the "end of central directory" record - at the
# *end* of the file. Its absence is what separates a half-written archive
# from a valid one, and it is why a file can have a perfect ZIP header, a
# readable first member, and still be unopenable: exactly what an
# interrupted download looks like.
_ZIP_EOCD_MAGIC = b"PK\x05\x06"
# The record is 22 bytes plus a comment of up to 65,535, so it can never
# begin further back than this from the end of the file.
_ZIP_EOCD_SEARCH = 22 + 65535
# Markers for the two things most often named .xlsx while being something
# else. Web apps export HTML tables and SpreadsheetML under that name.
_MARKUP_STARTS = (b"<", b"\xef\xbb\xbf<")

# Tried in order. latin-1 maps all 256 byte values so it can never raise,
# which makes it a guaranteed terminal fallback - we always open the file
# rather than refusing it. utf-8-sig covers plain UTF-8 *and* strips a BOM
# (ServiceNow exports often carry one, which would otherwise corrupt the
# first column name and break column matching).
ENCODINGS = ("utf-8-sig", "cp1252", "latin-1")

# Restricting candidates stops the sniffer being fooled by commas inside
# quoted free-text description fields.
DELIMITERS = ",;\t|"

_SNIFF_BYTES = 64 * 1024


@functools.lru_cache(maxsize=1)
def excel_engine() -> Optional[str]:
    """
    Preferred pandas Excel engine, or None to let pandas choose.

    openpyxl parses the sheet XML in Python and dominates the load time of a
    large export: on a 100k x 25 .xlsx it took ~21s against calamine's ~3.3s
    (6x), for byte-identical output - same shape, columns, dtypes and values.
    Reading the file *is* the load, so this is the single largest win available.

    calamine is an optional accelerator, not a requirement: if it is not
    installed we return None and pandas falls back to openpyxl/xlrd exactly as
    before. That keeps a missing wheel a slow load rather than a broken app.
    """
    try:
        import python_calamine  # noqa: F401
    except ImportError:
        return None
    return "calamine"


def _match_pandas_inference(frame: pd.DataFrame) -> pd.DataFrame:
    """
    Apply the type inference pandas' own Excel readers apply.

    Streaming rows straight from openpyxl skips it, and the difference is
    visible: a text cell holding "007" comes back as the string from openpyxl
    but as the integer 7 from `read_excel` on *either* engine. Reader choice
    must only affect speed and memory, never what the file appears to contain,
    so the streaming path is aligned to pandas rather than the reverse - even
    though keeping the leading zeros is arguably the better answer.
    """
    if frame.empty:
        # pandas' readers leave a header-only sheet's columns as object;
        # building one from empty lists yields float64 instead. There is no
        # data to infer from either way, so match the readers.
        return frame.astype(object) if len(frame.columns) else frame

    for name in frame.columns:
        column = frame[name]
        # pandas 3 infers `str` for a list of strings, so testing `== object`
        # alone silently skips every text column.
        if column.dtype != object and not pd.api.types.is_string_dtype(column.dtype):
            continue
        if column.isna().all():
            # An all-empty column is float64 from pandas' readers, object here
            frame[name] = column.astype("float64")
            continue
        try:
            frame[name] = pd.to_numeric(column)
        except (ValueError, TypeError):
            pass          # genuinely non-numeric - leave it alone
    return frame


def read_excel_streaming(path, sheet_name=None) -> pd.DataFrame:
    """
    Low-memory Excel reader: streams rows instead of building the whole sheet.

    Reserved for when the normal path runs out of memory. Every whole-sheet
    reader holds its own full copy of the data before pandas sees it, which is
    what makes the peak so much larger than the result: on a 300k x 25 sheet
    calamine peaked at ~1050 MB to produce a 40 MB frame. Streaming rows via
    openpyxl's read_only mode peaks at ~500 MB instead, for ~3x the time - a
    bad default, but the difference between opening the file and not.

    Measured alternatives that did *not* work, so they are not tried here:
    reading in column batches peaked *higher* (~1900 MB, the final concat
    doubles it), and declaring category dtypes up front did not lower the peak
    at all because the reader materialises its own copy first.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is None or isinstance(sheet_name, int):
            sheet = workbook[workbook.sheetnames[sheet_name or 0]]
        else:
            sheet = workbook[sheet_name]

        rows = sheet.iter_rows(values_only=True)
        try:
            raw_header = next(rows)
        except StopIteration:
            return pd.DataFrame()

        # Match pandas' naming for blank header cells
        header = [
            str(name) if name is not None else f"Unnamed: {i}"
            for i, name in enumerate(raw_header)
        ]
        width = len(header)
        columns: List[list] = [[] for _ in range(width)]

        # read_only mode can emit trailing all-empty rows from a stale sheet
        # dimension, which pandas' readers do not surface. Only *trailing* ones
        # may be dropped - an all-empty row between populated rows is real data
        # and discarding it silently shortens the frame. Counting deferred
        # empties keeps that distinction without buffering every row.
        pending_empty = 0
        for row in rows:
            if not any(value is not None for value in row):
                pending_empty += 1
                continue
            for _ in range(pending_empty):
                for column in columns:
                    column.append(None)
            pending_empty = 0
            for i in range(width):
                columns[i].append(row[i] if i < len(row) else None)

        # Hand each column over and drop our reference as we go, so the lists
        # and the frame are not both fully resident at the end.
        data = {}
        for i, name in enumerate(header):
            data[name] = columns[i]
            columns[i] = None
        frame = pd.DataFrame(data)

        return _match_pandas_inference(frame)
    finally:
        workbook.close()


def _read_excel_with_fallback(path, sheet_name):
    """
    Read an Excel sheet, preferring the fast engine but never failing because
    of it. A file the accelerator cannot parse is retried on pandas' default
    engine, so engine choice can only affect speed, never whether a file opens.

    Running out of memory is handled separately: retrying another whole-sheet
    reader would just fail the same way, so that goes straight to the streaming
    reader.
    """
    engine = excel_engine()
    if engine:
        try:
            return pd.read_excel(path, sheet_name=sheet_name, engine=engine)
        except MemoryError:
            return _read_excel_low_memory(path, sheet_name)
        except Exception:
            pass
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except MemoryError:
        return _read_excel_low_memory(path, sheet_name)
    except Exception as exc:
        # Reachable without going through list_sheets - read_table is a public
        # entry point, and a file can list its sheets and still fail to read.
        raise ValueError(describe_unreadable_excel(path)) from exc


def _read_excel_low_memory(path, sheet_name):
    """
    Streaming retry after an out-of-memory read, with an actionable message if
    even that will not fit.
    """
    name = os.path.basename(str(path))
    # Worth a warning: this path is several times slower, so without it the
    # only symptom is an unexplained long load.
    log.warning("%s did not fit in memory; retrying in low-memory mode "
                "(slower)", name)
    try:
        return read_excel_streaming(path, sheet_name)
    except MemoryError:
        raise MemoryError(
            f"'{name}' needs more memory than is available, even in low-memory "
            f"mode. Try removing unused columns, splitting the workbook, or "
            f"saving the sheet as CSV - the same data as CSV needs roughly a "
            f"third of the memory of .xlsx."
        ) from None


# --- Memory ---

# A column whose distinct values approach its row count gains nothing from a
# category dtype: the categories array approaches the size of the column, and
# for a unique key like a ticket number it is measurably *worse*. Half is a
# safe cut-off - real ITSM code fields (priority, state, group, category) sit
# far below it, while ticket numbers and free-text descriptions sit far above.
MAX_CATEGORY_FRACTION = 0.5

# Below this the absolute saving is not worth spending time on, and it keeps
# small fixtures - including the whole unit-test suite - on the plain path.
MIN_ROWS_TO_OPTIMIZE = 10_000


def optimize_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert low-cardinality string columns to `category`, in place.

    An ITSM export is mostly repeated code fields, so this is the difference
    between holding a 300k-row file as ~341 MB and as ~109 MB (68%), measured
    with realistically unique free text. It also makes analysis *faster* -
    building the analyzer went 0.17s -> 0.05s - because grouping and counting
    run over integer codes.

    This lowers steady-state memory, not the peak during read: the column is
    materialised as strings before it can be converted. A file too large to
    read at all needs chunking, which is a different change.

    Callers must not assume string dtype afterwards. Two known traps, both
    already handled in core/analyzer.py: `fillna('')` raises on a categorical
    unless '' is a category, and `value_counts()` reports unobserved categories
    with a zero count.
    """
    if df is None or len(df) < MIN_ROWS_TO_OPTIMIZE:
        return df

    rows = len(df)
    for col in df.columns:
        dtype = df[col].dtype
        if isinstance(dtype, pd.CategoricalDtype):
            continue
        if not pd.api.types.is_string_dtype(dtype):
            continue
        try:
            distinct = df[col].nunique(dropna=True)
        except TypeError:
            # Unhashable values (a column of lists) - leave it alone
            continue
        if distinct and distinct / rows < MAX_CATEGORY_FRACTION:
            df[col] = df[col].astype("category")
    return df


def _extension(path) -> str:
    return os.path.splitext(str(path))[1].lower()


def _magic(path) -> bytes:
    try:
        with open(path, "rb") as fh:
            return fh.read(8)
    except OSError:
        return b""


def _has_zip_index(path) -> bool:
    """Whether a zip's end-of-central-directory record is present."""
    try:
        size = os.path.getsize(path)
        with open(path, "rb") as fh:
            fh.seek(max(0, size - _ZIP_EOCD_SEARCH))
            return fh.read().rfind(_ZIP_EOCD_MAGIC) >= 0
    except OSError:
        return False


def describe_unreadable_excel(path) -> str:
    """
    Say what is wrong with a spreadsheet the Excel engines could not open.

    This exists because the raw exception is actively misleading: a truncated
    .xlsx fails with "File is not a zip file", which reads as "wrong format" to
    anyone who does not know that .xlsx *is* a zip. The file is the right
    format; there is just less of it than there should be.

    Returned rather than raised so the caller decides which exception type
    carries it, and so it can be logged without being raised at all.
    """
    name = os.path.basename(str(path))
    try:
        size = os.path.getsize(path)
    except OSError:
        return f"'{name}' could not be read from disk."

    if size == 0:
        return f"'{name}' is empty - the file is 0 bytes."

    head = _magic(path)
    megabytes = size / 1_000_000

    # The case that prompted all of this. Header present, index missing: the
    # write or the download stopped early. Naming the size matters, because
    # the user's own copy of the finished file is the thing to compare it to.
    if head.startswith(_ZIP_MAGIC) and not _has_zip_index(path):
        return (
            f"'{name}' looks like an incomplete file. It is {megabytes:,.1f} MB "
            f"and starts like a valid .xlsx, but the index every .xlsx keeps at "
            f"the end is missing - which normally means the download or the "
            f"save was interrupted. Download it again, and check the finished "
            f"size matches the source. A '(1)' or '(2)' in the name is often a "
            f"browser retry, where only the last attempt completed."
        )

    # Named .xlsx, but actually the old binary format.
    if head.startswith(_OLE2_MAGIC) and _extension(path) in EXCEL_EXTENSIONS:
        return (
            f"'{name}' is in the older Excel format despite its "
            f"{_extension(path)} name. Open it in Excel and use Save As to "
            f"write a real .xlsx, or rename it to .xls."
        )

    # Named .xlsx, but actually HTML or SpreadsheetML - the usual output of a
    # web application's "export to Excel".
    if any(head.startswith(marker) for marker in _MARKUP_STARTS):
        return (
            f"'{name}' contains HTML or XML rather than a spreadsheet, which is "
            f"what some systems produce when they offer an 'Excel' export. Open "
            f"it in Excel and save it as .xlsx or .csv."
        )

    if not head.startswith(_ZIP_MAGIC) and not head.startswith(_OLE2_MAGIC):
        return (
            f"'{name}' is not a spreadsheet - its first bytes match no Excel "
            f"format. If it is really a text export, rename it to .csv and it "
            f"will open."
        )

    # Right container, right index, still unreadable: damaged inside.
    return (
        f"'{name}' is a valid archive but its contents are damaged or are not a "
        f"workbook. Try opening it in Excel; if that also fails, get a fresh "
        f"copy."
    )


def is_excel(path) -> bool:
    """True if the file should be read by pandas' Excel engines."""
    ext = _extension(path)
    if ext in EXCEL_EXTENSIONS or ext in LEGACY_EXCEL_EXTENSIONS:
        return True
    if ext in TEXT_EXTENSIONS:
        return False
    # Unknown/absent extension - trust the content instead
    head = _magic(path)
    return head.startswith(_ZIP_MAGIC) or head.startswith(_OLE2_MAGIC)


def detect_encoding(path) -> str:
    """First encoding from ENCODINGS that decodes the file's leading sample."""
    try:
        with open(path, "rb") as fh:
            sample = fh.read(_SNIFF_BYTES)
    except OSError:
        return ENCODINGS[0]

    for encoding in ENCODINGS:
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError as exc:
            # A multi-byte character split by the sample boundary is not a real
            # decode failure. Without this check a valid UTF-8 file whose 64 KB
            # boundary lands mid-character would fall through to cp1252 and
            # mojibake every accented character in the file.
            if exc.start >= len(sample) - 4:
                return encoding
    return ENCODINGS[-1]


def detect_delimiter(sample: str) -> str:
    """Sniff the field delimiter, defaulting to comma."""
    if not sample.strip():
        return ","
    try:
        return csv.Sniffer().sniff(sample, delimiters=DELIMITERS).delimiter
    except csv.Error:
        # Common for single-column files - comma is the safe default
        return ","


def _read_text_sample(path, encoding: str) -> str:
    try:
        with open(path, "r", encoding=encoding, errors="replace", newline="") as fh:
            return fh.read(_SNIFF_BYTES)
    except OSError:
        return ""


# --- Public API ---

def list_sheets(path) -> List[str]:
    """
    Sheet names for an Excel workbook.

    Returns [] for delimited text, which the caller should read as
    "single table - no sheet picker needed".
    """
    if not is_excel(path):
        return []
    engine = excel_engine()
    if engine:
        try:
            with pd.ExcelFile(path, engine=engine) as workbook:
                return list(workbook.sheet_names)
        except Exception:
            pass
    # Only once the default engine has also failed is the file genuinely
    # unreadable - engine choice may never decide whether a file opens. This is
    # the first call the GUI makes on any workbook, so it is where a corrupt
    # file surfaces, and where the raw "File is not a zip file" used to escape.
    try:
        with pd.ExcelFile(path) as workbook:
            return list(workbook.sheet_names)
    except Exception as exc:
        log.warning("could not list sheets in %s: %s",
                    os.path.basename(str(path)), exc)
        raise ValueError(describe_unreadable_excel(path)) from exc


def read_table(path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Read a tabular file into a DataFrame.

    sheet_name applies to Excel only and is ignored for delimited text.
    Raises ValueError with a user-facing message when the file has no data.

    Large frames come back with low-cardinality string columns as `category`
    (see optimize_dtypes) - do not assume string dtype downstream.
    """
    if is_excel(path):
        # Never pass sheet_name=None to read_excel - that returns a dict of
        # every sheet rather than a DataFrame.
        return optimize_dtypes(
            _read_excel_with_fallback(path, sheet_name if sheet_name else 0))

    encoding = detect_encoding(path)
    delimiter = detect_delimiter(_read_text_sample(path, encoding))
    try:
        return optimize_dtypes(
            pd.read_csv(path, sep=delimiter, encoding=encoding, low_memory=False))
    except pd.errors.EmptyDataError:
        raise ValueError(f"'{os.path.basename(str(path))}' is empty - no columns to analyse.")
