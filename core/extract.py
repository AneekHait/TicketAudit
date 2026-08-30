"""
Pull a smaller, workable file out of one Excel cannot handle.

The users of this tool receive ITSM exports that are wide (25+ columns) and long
(hundreds of thousands of rows), and Excel hangs on exactly the operation they
need: filter to a date window, drop the columns they do not care about. Since
TicketAudit opens those files comfortably, it is the only thing in their hands
that can do the filtering.

Measured on a 400k x 23 fixture, taking Feb-2025..Jul-2026 with 7 of 23 columns:

    source            400,000 x 23 = 9.2M cells
    filter                    14 ms
    result            199,129 x  7 = 1.39M cells   (6.6x smaller)
    write CSV               0.43 s
    write .xlsx            28-42 s

Two facts from that shape the module:

1. The filter is free and the write is everything, so `plan_extract` can be
   called on every keystroke to drive a live count, while `build_extract` and
   `write_extract` belong on a worker thread.
2. The column reduction does most of the work, not the date range. An 18-month
   window out of 36 months is still half the rows; going from 23 columns to 7 is
   what turned 9.2M cells into 1.4M.

No Qt imports, like the rest of `core/`.
"""
import logging
import os
from typing import Any, Dict, List, NamedTuple, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from core.excel_pivot import EXCEL_MAX_ROWS
from core.excel_report import (
    DETAIL_MAX_DATA_ROWS, DETAIL_PROGRESS_EVERY, MAX_COL_WIDTH, MIN_COL_WIDTH,
    sanitize_for_excel,
)

log = logging.getLogger("ticketaudit")

FORMAT_XLSX = "xlsx"
FORMAT_CSV = "csv"

# A soft caution shown next to the cell count, NOT a measured Excel threshold -
# labelled as a rule of thumb wherever it surfaces. The basis: our own write rate
# is ~59k text cells/s, so a million cells takes ~17s just to serialise, which is
# a reasonable proxy for "Excel will feel this too".
EXCEL_SLOW_CELLS = 1_000_000


class ExtractCancelled(Exception):
    """
    Raised when the caller asked to stop mid-write.

    A distinct type rather than a bool return, because a cancelled write must
    not be mistaken for a finished one anywhere up the stack - and the partial
    file is deleted before it propagates, so there is nothing left to mistake.
    """


class ExtractTooLarge(Exception):
    """Raised when an .xlsx extract would exceed Excel's row ceiling.

    Deliberately an error rather than a silent truncation. The report truncates
    because its detail sheet is a by-product the user did not size; here the user
    chose the range, so narrowing it or switching to CSV is a real choice they
    can make, and quietly dropping rows off the end of their own selection would
    be worse than refusing.
    """


class ExtractPlan(NamedTuple):
    """What a selection would produce, without producing it."""
    rows: int
    columns: int
    undated_rows: int              # rows the date filter cannot place either way
    undated_included: bool
    months: List[Tuple[str, int]]  # (YYYY-MM, count) inside the range
    excluded_rows: int = 0         # rows dropped by `exclude`, already counted out

    @property
    def cells(self) -> int:
        return self.rows * self.columns

    @property
    def exceeds_excel_rows(self) -> bool:
        return self.rows > DETAIL_MAX_DATA_ROWS

    @property
    def may_be_slow_in_excel(self) -> bool:
        return self.cells > EXCEL_SLOW_CELLS


# Month abbreviations spelled out rather than taken from strftime("%b"), which
# follows the C locale and would render "Feb-25" differently on a non-English
# Windows install - the same label then means two things across two machines.
MONTH_ABBR = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")

# Days missing from the start or end of a month before it is called partial.
# One or two quiet days at an edge is ordinary - a public holiday, or a Sunday
# with nothing raised - and flagging those would cry wolf on healthy data.
MIN_PARTIAL_EDGE_DAYS = 3

# A run of consecutive days with nothing at all, inside a month whose edges are
# covered. A whole week of silence in a ticket queue is not a quiet spell; it is
# missing data. This one is a heuristic and is labelled as one in the UI.
MIN_INTERIOR_GAP_DAYS = 7

# Beyond this many hand-picked months, they are summarised rather than listed.
MAX_MONTHS_SPELLED_OUT = 4


def _month_key(period: str) -> int:
    """"2025-02" -> a sortable integer, without building a Period object."""
    year, month = str(period).split("-")[:2]
    return int(year) * 12 + int(month) - 1


def month_label(period: str) -> str:
    """
    "2025-02" -> "Feb-25", the form the people using this name their ranges in.

    The ISO code stays the identifier everywhere else - it sorts, it is
    unambiguous, and it is what goes in filenames and provenance.
    """
    try:
        year, month = period.split("-")
        return f"{MONTH_ABBR[int(month) - 1]}-{year[-2:]}"
    except (ValueError, IndexError, AttributeError):
        return period


def range_label(start: str, end: str) -> str:
    """A human range: "Feb-25 to Jul-26". Empty when there is no filter."""
    if not start or not end:
        return "all dates"
    if start == end:
        return month_label(start)
    return f"{month_label(start)} to {month_label(end)}"


def is_contiguous(months: Sequence[str]) -> bool:
    """Whether these months form an unbroken run, so a range can describe them."""
    keys = sorted(_month_key(m) for m in months)
    return bool(keys) and keys == list(range(keys[0], keys[0] + len(keys)))


def months_label(months: Sequence[str]) -> str:
    """
    Name a hand-picked set of months.

    Spelled out while the list is short, because "Feb-26, May-26 and Jul-26" is
    the thing the reader needs to know and a range would be a lie about it.
    Beyond that it is summarised with its span, since a wall of month names in a
    dialog is not read.
    """
    ordered = sorted(months, key=_month_key)
    if not ordered:
        return "no months"
    if is_contiguous(ordered):
        return range_label(ordered[0], ordered[-1])
    names = [month_label(m) for m in ordered]
    if len(names) <= MAX_MONTHS_SPELLED_OUT:
        return ", ".join(names[:-1]) + f" and {names[-1]}"
    return (f"{len(names)} selected months between "
            f"{names[0]} and {names[-1]}")


class MonthCoverage(NamedTuple):
    """
    How much of a calendar month the file actually has rows for.

    Judged on the **span** of days present, not the count of distinct days: a
    weekday-only queue is missing eight or nine days of every month and is
    perfectly complete. What marks a month as partial is rows stopping before
    the month does - an export taken mid-month, or starting mid-month.
    """
    period: str
    rows: int
    days_in_month: int
    first_day: int
    last_day: int
    longest_gap: int          # longest interior run of days with no rows

    @property
    def missing_start(self) -> int:
        return self.first_day - 1

    @property
    def missing_end(self) -> int:
        return self.days_in_month - self.last_day

    @property
    def truncated_start(self) -> bool:
        return self.missing_start >= MIN_PARTIAL_EDGE_DAYS

    @property
    def truncated_end(self) -> bool:
        return self.missing_end >= MIN_PARTIAL_EDGE_DAYS

    @property
    def has_interior_gap(self) -> bool:
        return self.longest_gap >= MIN_INTERIOR_GAP_DAYS

    @property
    def partial(self) -> bool:
        return (self.truncated_start or self.truncated_end
                or self.has_interior_gap)

    @property
    def covered_days(self) -> int:
        """Days between the first and last row, inclusive."""
        return max(0, self.last_day - self.first_day + 1)

    def describe(self) -> str:
        """
        Why this month is partial, in days rather than adjectives.

        A count nobody can trace is not actionable: "Aug-26 has only 412 rows"
        invites the reader to believe August was quiet, where "Aug-26 stops on
        the 12th of 31" tells them the month is not over.
        """
        label = month_label(self.period)
        if self.truncated_start and self.truncated_end:
            return (f"{label} only has rows from day {self.first_day} to day "
                    f"{self.last_day} of {self.days_in_month}")
        if self.truncated_end:
            return (f"{label} stops on day {self.last_day} of "
                    f"{self.days_in_month}")
        if self.truncated_start:
            return (f"{label} starts on day {self.first_day}, so days 1-"
                    f"{self.first_day - 1} are missing")
        return (f"{label} has a {self.longest_gap}-day stretch with no rows "
                f"at all")


def month_coverage(dates: pd.Series) -> List[MonthCoverage]:
    """
    Day coverage per month, ascending.

    Vectorised because it runs on the same path as the live count. The obvious
    form - to_period("M").astype(str) then groupby - spends its time building
    400k strings to produce 36 groups, and measured 167ms at 400k rows against
    18ms for this. Month identity is carried as year*12+month instead, and the
    days present per month as one boolean matrix filled by a single fancy-index
    assignment.
    """
    if dates is None:
        return []
    valid = dates.dropna()
    if valid.empty:
        return []

    years = valid.dt.year.to_numpy()
    months = valid.dt.month.to_numpy()
    days = valid.dt.day.to_numpy()
    keys = years.astype("int64") * 12 + (months - 1)

    order = np.unique(keys)
    index = np.searchsorted(order, keys)
    rows = np.bincount(index, minlength=len(order))

    # One row per month, one column per day-of-month (1..31).
    present = np.zeros((len(order), 32), dtype=bool)
    present[index, days] = True

    out: List[MonthCoverage] = []
    for slot, key in enumerate(order):
        year, month = divmod(int(key), 12)
        period = f"{year:04d}-{month + 1:02d}"
        flags = present[slot]
        seen = np.flatnonzero(flags)
        first, last = int(seen[0]), int(seen[-1])
        # Longest run of absent days strictly between the first and last.
        longest = run = 0
        for day in range(first + 1, last):
            if flags[day]:
                run = 0
            else:
                run += 1
                longest = max(longest, run)
        out.append(MonthCoverage(
            period=period, rows=int(rows[slot]),
            days_in_month=int(pd.Period(period, freq="M").days_in_month),
            first_day=first, last_day=last, longest_gap=longest))
    return out


def partial_months(coverage: Sequence[MonthCoverage], start: str = "",
                   end: str = "") -> List[MonthCoverage]:
    """The partial months inside a range - all of them when no range is given."""
    inside = [c for c in coverage
              if (not start or c.period >= start) and (not end or c.period <= end)]
    return [c for c in inside if c.partial]


def column_fill(df: pd.DataFrame,
                null_counts: Optional[pd.Series] = None) -> Dict[str, float]:
    """
    The fraction of each column that actually holds a value, 0.0 to 1.0.

    Blank means null *or* the empty string. Counting only nulls would report a
    column of "" as completely full, and on an ITSM export the columns nobody
    fills are exactly the ones worth dropping - so the distinction matters here
    more than anywhere else.

    `null_counts` accepts the analyzer's cached df.isnull().sum() so a wide file
    is not scanned twice. The empty-string pass is an elementwise comparison
    against a scalar, not a string conversion: astype(str) on a category column
    would materialise the whole thing (see the loader notes on dtypes).
    """
    total = len(df)
    if total == 0:
        return {str(column): 0.0 for column in df.columns}

    if null_counts is None:
        null_counts = df.isnull().sum()

    fill: Dict[str, float] = {}
    for column in df.columns:
        blank = int(null_counts.get(column, 0))
        series = df[column]
        # Anything that could hold text. Not `dtype == object`: under pandas 3 a
        # string column arrives as the `str` dtype, and testing for object
        # silently skipped every text column - reporting a sheet full of ""
        # as completely filled.
        if not (pd.api.types.is_numeric_dtype(series)
                or pd.api.types.is_datetime64_any_dtype(series)
                or pd.api.types.is_timedelta64_dtype(series)
                or pd.api.types.is_bool_dtype(series)):
            try:
                blank += int((series == "").sum())
            except (TypeError, ValueError):
                # A column of mixed types can refuse the comparison; the null
                # count alone is still a truthful lower bound on blankness.
                log.debug("could not count empty strings in %s", column,
                          exc_info=True)
        fill[str(column)] = max(0.0, min(1.0, 1.0 - blank / total))
    return fill


def month_counts(dates: pd.Series) -> List[Tuple[str, int]]:
    """
    Every month present in a parsed date column, with its row count, ascending.

    The counts are the part Excel cannot give you without filtering first, which
    is why they belong in the picker rather than being computed only on demand.
    """
    if dates is None:
        return []
    valid = dates.dropna()
    if valid.empty:
        return []
    counts = valid.dt.to_period("M").value_counts().sort_index()
    return [(str(period), int(count)) for period, count in counts.items()]


def month_bounds(start: str, end: str) -> Tuple[pd.Timestamp, pd.Timestamp]:
    """
    Half-open timestamp bounds for an inclusive month range.

    The end is the START of the month after `end`, so the whole of the closing
    month is included. Writing the upper bound as `<= start_of(end)` is the
    off-by-a-month that silently keeps only the 1st of July from a range ending
    in Jul-2026 - the single most likely bug in this feature, so it is fixed by
    construction here rather than left to each call site.
    """
    first = pd.Period(start, freq="M")
    last = pd.Period(end, freq="M")
    if last < first:
        first, last = last, first
    return first.start_time, (last + 1).start_time


def _month_keys(dates: pd.Series) -> "np.ndarray":
    """
    Each row's month as year*12+month, for set membership.

    Not to_period("M").astype(str): that builds one string per row to compare
    against a handful of values, and is what made month_coverage 167ms at 400k
    rows before it was rewritten this way.
    """
    return (dates.dt.year.to_numpy().astype("int64") * 12
            + dates.dt.month.to_numpy() - 1)


def _apply_exclusion(mask: pd.Series, exclude) -> pd.Series:
    """
    Remove the excluded rows from a selection mask.

    `exclude` is a boolean mask of rows to drop, decided by the caller: which
    values mean "cancelled" is vocabulary, and vocabulary belongs with the
    analyzer rather than in here. This module only needs to know which rows are
    out.
    """
    if exclude is None:
        return mask
    dropped = pd.Series(exclude, index=mask.index).fillna(False).astype(bool)
    return mask & ~dropped


def _selection_mask(dates: pd.Series, start: str, end: str,
                    months: Optional[Sequence[str]],
                    include_undated: bool) -> pd.Series:
    """
    Rows inside the chosen months, whether that is a range or a picked set.

    A comparison against NaT is False and isin on a NaN key is False, so undated
    rows fall out of either mask without a separate guard - and are added back
    only when asked for.
    """
    if months is not None:
        wanted = {_month_key(m) for m in months}
        if not wanted:
            mask = pd.Series(False, index=dates.index)
        else:
            dated = dates.notna()
            keys = np.full(len(dates), -1, dtype="int64")
            keys[dated.to_numpy()] = _month_keys(dates[dated])
            mask = pd.Series(np.isin(keys, list(wanted)), index=dates.index)
    else:
        lower, upper = month_bounds(start, end)
        mask = (dates >= lower) & (dates < upper)
    if include_undated:
        mask = mask | dates.isna()
    return mask


def _ordered_columns(df: pd.DataFrame, columns: Sequence[str]) -> List[str]:
    """
    Selected columns, in source order.

    Tick order is not a meaningful order, and a reader comparing the extract
    against the original wants the columns where they expect them.
    """
    wanted = set(columns)
    return [column for column in df.columns if column in wanted]


def plan_extract(df: pd.DataFrame, dates: pd.Series, *, start: str = "",
                 end: str = "", columns: Sequence[str],
                 months: Optional[Sequence[str]] = None,
                 include_undated: bool = False,
                 exclude=None) -> ExtractPlan:
    """
    Count what a selection would produce, without materialising it.

    `months` picks an arbitrary set - "February, May and July" is a real request
    and is not a range - and takes precedence over start/end when given. This is
    what the live count calls, so it must stay cheap: ~15 ms at 400k rows. It
    counts rather than slices for that reason.
    """
    selected = _ordered_columns(df, columns)
    if dates is None:
        # No usable date column, so no month filter can narrow anything; the
        # column selection and the exclusion still can.
        everything = pd.Series(True, index=df.index)
        kept = _apply_exclusion(everything, exclude)
        return ExtractPlan(rows=int(kept.sum()), columns=len(selected),
                           undated_rows=0, undated_included=False, months=[],
                           excluded_rows=int(len(df) - kept.sum()))

    undated = dates.isna()
    chosen = _selection_mask(dates, start, end, months, include_undated=False)
    if include_undated:
        chosen = chosen | undated

    kept = _apply_exclusion(chosen, exclude)
    # Counted against the selection, not the file: "1,204 excluded" has to mean
    # 1,204 fewer rows in this extract, or the arithmetic on screen does not add
    # up for anyone checking it.
    excluded = int(chosen.sum() - kept.sum())

    inside = dates[kept & ~undated]
    counts = (inside.dt.to_period("M").value_counts().sort_index()
              if not inside.empty else pd.Series(dtype="int64"))

    return ExtractPlan(
        rows=int(kept.sum()),
        columns=len(selected),
        # The number ticking the undated box would *add*, which is not the same
        # as the number of undated rows in the file once an exclusion is on. It
        # must not change meaning with the flag, or the label beside the box
        # would say one thing when ticked and another when not.
        undated_rows=int(_apply_exclusion(undated, exclude).sum()),
        undated_included=include_undated,
        months=[(str(period), int(count)) for period, count in counts.items()],
        excluded_rows=excluded,
    )


def build_extract(df: pd.DataFrame, dates: pd.Series, *, start: str = "",
                  end: str = "", columns: Sequence[str],
                  months: Optional[Sequence[str]] = None,
                  include_undated: bool = False,
                  exclude=None,
                  limit: Optional[int] = None) -> pd.DataFrame:
    """
    The subset itself. Column dtypes are preserved, categories included - see
    loader.optimize_dtypes; converting to str here would roughly double the
    memory for no gain, since the writers sanitise separately.

    `limit` takes only the first n matching rows, for the preview. It is the
    same function with the same mask and the same column order, deliberately: a
    preview built by a second code path could differ from the file that follows
    it, and then it would be worse than no preview at all.
    """
    selected = _ordered_columns(df, columns)
    if not selected:
        raise ValueError("no columns selected for the extract")
    if dates is None:
        mask = _apply_exclusion(pd.Series(True, index=df.index), exclude)
        frame = df.loc[mask, selected]
        return frame.head(limit) if limit is not None else frame

    mask = _apply_exclusion(
        _selection_mask(dates, start, end, months, include_undated), exclude)
    if limit is None:
        return df.loc[mask, selected]
    # Positional, so the cost is the first n rows rather than the whole subset.
    keep = np.flatnonzero(mask.to_numpy())[:limit]
    return df.iloc[keep].loc[:, selected]


def suggest_filename(source_path: Optional[str], date_column: str,
                     start: str, end: str, fmt: str,
                     months: Optional[Sequence[str]] = None) -> str:
    """
    A filename that records what the file contains.

    The months and the date column go in the name because a CSV carries no
    metadata sheet: "September's tickets" means different rows depending on
    whether September was read from the opened or the resolved date, and the
    filename is the only place a CSV can say so.

    A hand-picked set is named by its span and its count - "2026-02_to_2026-07"
    alone would claim the months in between, which is the one thing the name
    must not do. The full list lives in the Extract Info sheet or the sidecar.
    """
    stem = "extract"
    if source_path:
        stem = os.path.splitext(os.path.basename(source_path))[0]
    safe_column = "".join(c if c.isalnum() else "_" for c in date_column or "date")

    if months is not None:
        ordered = sorted(months, key=_month_key)
        if not ordered:
            span = "no_months"
        elif len(ordered) == 1:
            span = ordered[0]
        elif is_contiguous(ordered):
            span = f"{ordered[0]}_to_{ordered[-1]}"
        else:
            span = f"{len(ordered)}_months_{ordered[0]}_to_{ordered[-1]}"
    else:
        span = f"{start}_to_{end}"
    return f"{stem}_{safe_column}_{span}.{fmt}"


def _provenance(date_column: str, start: str, end: str, plan: ExtractPlan,
                source_path: Optional[str],
                incomplete: Sequence[str] = (),
                months: Optional[Sequence[str]] = None,
                excluded_states: Sequence[str] = ()) -> List[Tuple[str, Any]]:
    """
    The facts a reader needs to interpret the extract, as label/value pairs.

    The incomplete months are here for the same reason the undated count is: a
    month the file only partly covers has a real row count that is not
    comparable with a whole month, and a reader who cannot see which months
    those were will read the shortfall as a drop in volume.
    """
    return [
        ("Source file", os.path.basename(source_path) if source_path else "n/a"),
        ("Date column used", date_column or "none"),
        # Named exactly, because a picked set is not a range and describing it as
        # one would claim months the extract does not contain.
        ("Months included",
         months_label(months) if months is not None
         else f"{range_label(start, end)}"
              f"{f'  ({start} to {end})' if start and end else ''}"),
        ("Months, in full",
         ", ".join(month_label(m) for m in sorted(months, key=_month_key))
         if months is not None else "a continuous range"),
        ("Rows written", plan.rows),
        ("Columns written", plan.columns),
        ("Rows with no date", plan.undated_rows),
        ("Undated rows included", "Yes" if plan.undated_included else "No"),
        # Rows that are missing from this file for a reason the reader cannot
        # infer from the data in it. Without this the extract is simply short.
        ("States excluded",
         ", ".join(excluded_states) if excluded_states else "none"),
        ("Rows excluded by state", plan.excluded_rows),
        ("Incomplete months included",
         ", ".join(month_label(p) for p in incomplete) if incomplete else "none"),
        ("Generated", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")),
    ]


def _write_rows(worksheet, frame: pd.DataFrame, formula_columns: set,
                progress=None, cancelled=None) -> int:
    """
    Stream the rows, wrapping only the cells openpyxl would treat as formulas.

    A write-only sheet cannot have a cell mutated after assignment the way
    write_block does, so a formula-leading string has to arrive as a
    WriteOnlyCell with its type already set. The wrap is per-column, not
    per-cell: sanitize_for_excel has already said which columns can contain one,
    and testing every cell of millions would cost more than it saves.
    """
    columns = list(frame.columns)
    formula_indexes = [i for i, name in enumerate(columns)
                       if name in formula_columns]
    total = len(frame)
    written = 0

    for row in frame.itertuples(index=False, name=None):
        if formula_indexes:
            values = list(row)
            for index in formula_indexes:
                value = values[index]
                if isinstance(value, str) and value.startswith("="):
                    cell = WriteOnlyCell(worksheet, value=value)
                    cell.data_type = "s"
                    values[index] = cell
            worksheet.append(values)
        else:
            worksheet.append(list(row))

        written += 1
        if written % DETAIL_PROGRESS_EVERY == 0:
            if progress:
                progress(f"Writing extract: {written:,} of {total:,} rows "
                         f"({written / total:.0%})")
            # Checked on the same beat as the progress message: often enough to
            # feel immediate on a write measured in tens of seconds, rarely
            # enough that the check costs nothing against millions of rows.
            if cancelled is not None and cancelled():
                raise ExtractCancelled(
                    f"cancelled after {written:,} of {total:,} rows")
    return written


def info_sidecar_path(path: str) -> str:
    """Where the CSV's provenance goes: beside it, never inside it."""
    base, _extension = os.path.splitext(path)
    return f"{base}_ExtractInfo.txt"


def write_extract(path: str, frame: pd.DataFrame, *, fmt: str = FORMAT_XLSX,
                  date_column: str = "", start: str = "", end: str = "",
                  plan: Optional[ExtractPlan] = None,
                  source_path: Optional[str] = None,
                  incomplete: Sequence[str] = (),
                  months: Optional[Sequence[str]] = None,
                  excluded_states: Sequence[str] = (),
                  progress=None, cancelled=None) -> int:
    """
    Write the extract and return the row count.

    Deliberately lean: header row, values, a frozen header, sized columns and
    nothing else. No Excel Table, no conditional formatting, no charts - all of
    which the report adds and all of which cost Excel real time at scale. The
    whole point of this file is that Excel does not choke on it.

    Being lean is also what makes it fast. With no PivotTable to host, the
    workbook can use write_only mode, which the report cannot: add_pivot does not
    exist on a write-only sheet and write_only is a workbook-level flag. Normal
    mode would hold every Cell object for the same 1.4M cells.
    """
    if frame.empty:
        raise ValueError("the extract is empty; widen the range or the columns")

    if cancelled is not None and cancelled():
        raise ExtractCancelled("cancelled before writing started")

    if fmt == FORMAT_CSV:
        if progress:
            progress(f"Writing extract CSV: {len(frame):,} rows…")
        # max_chars=None: CSV has no 32,767-character cell limit, so truncating
        # would lose text for nothing. Control characters are still stripped.
        sanitised, _formula = sanitize_for_excel(frame, max_chars=None)
        # utf-8-sig, because without the BOM Excel mis-decodes non-ASCII on a
        # double-click and mangles exactly the non-English text this tool exists
        # to find.
        sanitised.to_csv(path, index=False, encoding="utf-8-sig")

        # Beside the file, not above the data: provenance lines inside the CSV
        # would break read_csv and any re-import, which is the same reason the
        # .xlsx puts it on a second sheet. A sidecar keeps the promise that the
        # date column used travels with the extract.
        rows = _provenance(date_column, start, end,
                           plan or ExtractPlan(len(sanitised),
                                               len(sanitised.columns),
                                               0, False, []),
                           source_path, incomplete, months, excluded_states)
        try:
            with open(info_sidecar_path(path), "w", encoding="utf-8-sig") as handle:
                print("TicketAudit extract", file=handle)
                for label, value in rows:
                    print(f"{label}: {value}", file=handle)
        except OSError:
            # The extract itself is written and valid; losing the note beside it
            # is not worth failing the export over.
            log.warning("could not write the extract info file beside %s", path,
                        exc_info=True)
        return len(sanitised)

    if len(frame) > DETAIL_MAX_DATA_ROWS:
        raise ExtractTooLarge(
            f"{len(frame):,} rows will not fit an Excel worksheet, which holds "
            f"{EXCEL_MAX_ROWS:,} including the header. Narrow the range, or "
            f"choose CSV, which has no limit.")

    sanitised, formula_columns = sanitize_for_excel(frame)
    headers = [str(column) for column in sanitised.columns]

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Extract")
    # Both must be set before the first append: openpyxl emits the sheet
    # properties and column definitions when the first row is written, so doing
    # this afterwards silently does nothing.
    worksheet.freeze_panes = "A2"
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            MIN_COL_WIDTH, min(MAX_COL_WIDTH, len(header) + 2))

    try:
        worksheet.append(headers)
        written = _write_rows(worksheet, sanitised, formula_columns, progress,
                              cancelled)

        # A second sheet rather than a header block on the first: anything above
        # the data would break a copy-paste or a re-import of the extract.
        info = workbook.create_sheet("Extract Info")
        info.append(["Setting", "Value"])
        for label, value in _provenance(date_column, start, end,
                                        plan or ExtractPlan(
                                            written, len(headers), 0, False, []),
                                        source_path, incomplete, months,
                                        excluded_states):
            info.append([label, value])

        if progress:
            progress("Finalising the extract…")
        workbook.save(path)
        return written
    except Exception:
        # A streamed sheet writes into a temp file that openpyxl only removes
        # during save(). If an append raises, save() never runs and the temp file
        # survives until process exit.
        writer = getattr(worksheet, "_writer", None)
        if writer is not None:
            # Close the row generator *before* the temp file, in that order.
            # openpyxl drives the rows through a generator that is inside a
            # `with xf.element("sheetData")` block; if the file goes first, the
            # generator's own finalizer tries to write the closing tag to a
            # closed file and Python prints an ignored-exception traceback from
            # the garbage collector - on every cancel, long after the event.
            rows = getattr(worksheet, "_rows", None)
            if rows is not None:
                try:
                    rows.close()
                except Exception:
                    log.debug("could not close the streamed row generator",
                              exc_info=True)
            try:
                writer.cleanup()
            except Exception:
                log.debug("could not clean up the streamed sheet's temp file",
                          exc_info=True)
        # Nothing partial is left behind. save() has either not run or not
        # finished, and a truncated .xlsx that Excel refuses to open - or worse,
        # opens short - must not be left sitting where the user asked for their
        # extract.
        if os.path.exists(path):
            try:
                os.remove(path)
            except OSError:
                log.warning("could not remove the incomplete extract at %s",
                            path, exc_info=True)
        raise
