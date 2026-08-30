"""
Tests for core/excel_pivot.py — native Excel PivotTable construction.

openpyxl has no supported API for *creating* pivot tables, so this module hand
assembles the XML. The failure mode is not an exception: Excel offers to
"repair" the workbook and silently drops every pivot, which no unit test would
notice unless it looks for the specific structural defects. Hence two canaries
that need no Excel:

  * validate() — content types, dangling rels, cacheId wiring, field indexes
  * load_workbook() on the file just written — raises KeyError on the
    cache-dedupe defect, which is the highest-probability corruption

Verified against real Excel (Office 16) during development: 3 pivots survived
the open, one shared cache, grand totals matched the independently computed
values.
"""
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from core.excel_pivot import (
    PivotBuilder, range_ref, register_table, table_name, validate,
)


HEADERS = ["Priority", "State", "Ticket count"]
ROWS = [
    ("P1", "Open", 12), ("P1", "Closed", 30), ("P2", "Open", 5),
    ("P2", "Closed", 44), ("P3", "Open", 2), ("P4", "Closed", 1),
]


def _workbook(save_records=True, cache_ids=(1, 1), separate_builders=False):
    """A cube sheet plus a Pivots sheet carrying two pivots."""
    wb = Workbook()
    cube = wb.active
    cube.title = "PivotData"
    cube.append(HEADERS)
    for row in ROWS:
        cube.append(list(row))

    used = set()
    name = table_name("PivotData", used)
    register_table(cube, name, range_ref(1, 1, len(ROWS) + 1, len(HEADERS)))

    sheet = wb.create_sheet("Pivots")

    def make(cache_id):
        return PivotBuilder(HEADERS, ROWS, axis_fields=[0, 1],
                            cache_id=cache_id, source_name=name,
                            save_records=save_records)

    first = make(cache_ids[0])
    first.add(sheet, "pt_Priority", "A3", row_field=0, data_field=2)
    if separate_builders:
        # Two structurally identical caches built separately: the writer dedupes
        # them by equality, so this is the corruption case.
        second = make(cache_ids[1])
        second.add(sheet, "pt_State", "E3", row_field=1, data_field=2)
    else:
        first.add(sheet, "pt_State", "E3", row_field=1, data_field=2)
    return wb, first


def _saved(wb) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestProducesAValidWorkbook:
    def test_validator_reports_nothing(self):
        data = _saved(_workbook()[0])
        assert validate(data) == []

    def test_openpyxl_can_reload_it(self):
        """The cacheId canary — KeyError here is the cache-dedupe defect."""
        data = _saved(_workbook()[0])
        reloaded = load_workbook(BytesIO(data))
        assert len(reloaded["Pivots"]._pivots) == 2

    def test_pivots_share_exactly_one_cache(self):
        reloaded = load_workbook(BytesIO(_saved(_workbook()[0])))
        caches = {id(p.cache) for p in reloaded["Pivots"]._pivots}
        assert len(caches) == 1, "one source must mean one cache part"

    def test_location_covers_items_plus_header_and_grand_total(self):
        reloaded = load_workbook(BytesIO(_saved(_workbook()[0])))
        by_name = {p.name: p for p in reloaded["Pivots"]._pivots}
        # 4 priorities + header + grand total = rows 3..8
        assert by_name["pt_Priority"].location.ref == "A3:B8"
        # 2 states + header + grand total = rows 3..6
        assert by_name["pt_State"].location.ref == "E3:F6"

    def test_records_are_written_by_default(self):
        """Records make the pivot render in Excel Online and LibreOffice too."""
        reloaded = load_workbook(BytesIO(_saved(_workbook()[0])))
        cache = reloaded["Pivots"]._pivots[0].cache
        assert cache.recordCount == len(ROWS)
        assert cache.refreshOnLoad is True

    def test_omitting_records_still_sets_refresh_on_load(self):
        """Without records the pivot depends entirely on Excel refreshing."""
        data = _saved(_workbook(save_records=False)[0])
        assert validate(data) == []
        cache = load_workbook(BytesIO(data))["Pivots"]._pivots[0].cache
        assert cache.refreshOnLoad is True
        assert cache.recordCount is None, \
            "a record count with no records part is self-contradictory"


class TestTheCacheDedupeTrap:
    """
    The writer dedupes caches by structural equality, not identity. Two
    separately built but identical caches collapse to one part while two
    cacheIds are emitted, leaving the second pivot pointing at nothing.
    """

    def test_distinct_cache_ids_on_one_source_corrupt_the_file(self):
        data = _saved(_workbook(cache_ids=(1, 2), separate_builders=True)[0])
        with pytest.raises(KeyError):
            load_workbook(BytesIO(data))

    def test_the_validator_catches_it_too(self):
        data = _saved(_workbook(cache_ids=(1, 2), separate_builders=True)[0])
        findings = validate(data)
        assert findings, "validate() must not pass a dangling cacheId"
        assert any("cacheId" in f for f in findings)

    def test_one_builder_per_source_is_the_fix(self):
        """PivotBuilder owns the cache, so its pivots cannot drift apart."""
        data = _saved(_workbook(cache_ids=(1, 1), separate_builders=False)[0])
        assert validate(data) == []
        load_workbook(BytesIO(data))


class TestGeometry:
    def test_overlapping_pivots_are_rejected(self):
        wb, builder = _workbook()
        sheet = wb["Pivots"]
        # pt_Priority already occupies A3:B8
        builder.add(sheet, "pt_Clash", "A5", row_field=1, data_field=2)
        with pytest.raises(ValueError, match="overlap"):
            builder.assert_disjoint()

    def test_side_by_side_pivots_pass(self):
        _, builder = _workbook()
        builder.assert_disjoint()

    def test_a_written_cell_inside_a_pivot_is_rejected(self):
        """Excel overwrites its own pivot range on refresh."""
        _, builder = _workbook()
        with pytest.raises(ValueError, match="inside pivot"):
            builder.assert_disjoint(occupied=[(4, 1)])

    def test_a_title_above_the_pivots_is_fine(self):
        _, builder = _workbook()
        builder.assert_disjoint(occupied=[(1, 1)])


class TestTableNames:
    def test_spaces_and_punctuation_are_replaced(self):
        assert table_name("All Details", set()) == "tbl_All_Details"

    def test_names_are_made_unique(self):
        used = set()
        first = table_name("Pivot Data", used)
        second = table_name("Pivot Data", used)
        assert first != second
        assert second.endswith("_2")

    def test_uniqueness_is_case_insensitive(self):
        """Excel's table/defined-name namespace ignores case."""
        used = set()
        table_name("data", used)
        assert table_name("DATA", used) != "tbl_DATA"

    def test_a_blank_label_still_yields_a_valid_name(self):
        assert table_name("!!!", set()) == "tbl_Data"

    def test_names_never_start_with_a_digit(self):
        """A name shaped like a cell reference or starting with a digit is
        rejected by Excel even though openpyxl accepts it."""
        assert table_name("2024 tickets", set()).startswith("tbl_")
        assert table_name("A1", set()).startswith("tbl_")


class TestRangeRef:
    def test_single_cell(self):
        assert range_ref(1, 1, 1, 1) == "A1:A1"

    def test_multi_column(self):
        assert range_ref(1, 1, 8, 4) == "A1:D8"

    def test_past_column_z(self):
        assert range_ref(2, 27, 5, 28) == "AA2:AB5"


class TestNumpyScalars:
    """
    A pandas aggregate yields np.int64, which is not an instance of int on
    Windows. Unhandled it falls through to Text(), so the measure ships as a
    string and the pivot's Sum quietly produces zero - a file that opens
    perfectly and is simply wrong.
    """

    def test_a_numpy_measure_is_written_as_a_number(self):
        import numpy as np
        from openpyxl.pivot.fields import Number

        rows = [("P1", np.int64(12)), ("P2", np.int64(30))]
        builder = PivotBuilder(["Priority", "Ticket count"], rows,
                              axis_fields=[0], source_name="tbl_x")
        for record in builder.cache.records.r:
            measure = record._fields[1]
            assert isinstance(measure, Number), \
                "a numpy int must not be written as text"

    def test_numpy_values_survive_the_round_trip(self):
        import numpy as np
        wb = Workbook()
        sheet = wb.active
        rows = [("P1", np.int64(12)), ("P2", np.int64(30))]
        builder = PivotBuilder(["Priority", "Ticket count"], rows,
                              axis_fields=[0], source_name="tbl_x")
        builder.add(sheet, "pt", "A3", row_field=0, data_field=1)
        assert validate(_saved(wb)) == []

    def test_numpy_axis_values_are_unwrapped(self):
        import numpy as np
        rows = [(np.int64(1), 10), (np.int64(2), 20)]
        builder = PivotBuilder(["Sprint", "Ticket count"], rows,
                              axis_fields=[0], source_name="tbl_x")
        # stored as Python ints, so the sharedItems index lookup works
        assert all(isinstance(value, int) and not isinstance(value, np.generic)
                   for value in builder._shared[0])


class TestBuilderGuards:
    def test_a_source_must_be_named_or_located(self):
        with pytest.raises(ValueError, match="source_name or source_sheet"):
            PivotBuilder(HEADERS, ROWS, axis_fields=[0])

    def test_a_sheet_and_ref_source_also_works(self):
        wb = Workbook()
        sheet = wb.active
        builder = PivotBuilder(HEADERS, ROWS, axis_fields=[0],
                              source_sheet="PivotData", source_ref="A1:C7")
        builder.add(sheet, "pt", "A3", row_field=0, data_field=2)
        assert validate(_saved(wb)) == []
