"""
Tests for core/reporter.py — the logic-check report surfaces.

Focused on one property that is easy to break and expensive to get wrong: a rule
that could not run must never be reported as a rule that passed. A "0 / Pass"
against a check the tool skipped tells the reader their data is clean on a
dimension nobody looked at.
"""
import pytest
import pandas as pd

from core.analyzer import SanityAnalyzer
from core.reporter import ReportGenerator


def _report(df, **kwargs):
    return ReportGenerator(df, SanityAnalyzer(df), filepath="tickets.xlsx", **kwargs)


@pytest.fixture
def messy_df():
    """One instance of every cross-field defect, with exact counts."""
    n = 60
    return pd.DataFrame({
        "number": [f"INC{i:04d}" for i in range(n)],
        "state": ["Closed"] * 40 + ["New"] * 18 + ["Closd"] + ["Stage 9"],
        "priority": ["P1"] * 57 + [None] * 3,
        "opened_at": pd.to_datetime(["2024-01-01"] * n),
        "resolved_at": pd.to_datetime(
            ["2024-02-01"] * 38 + [None] * 2      # 2 closed with no date
            + ["2024-02-01"] + [None] * 19),      # 1 open with a date
        "assignment_group": ["Team"] * n,
        "cmdb_ci": ["CI"] * n,
        "short_description": ["Login issue for the user account"] * n,
        "description": ["A sufficiently detailed description of the problem"] * n,
    })


# ──────────────────────────────────────────────────────────────
# Logic Checks sheet
# ──────────────────────────────────────────────────────────────

class TestLogicChecksSheet:
    def _verdicts(self, df):
        frame = _report(df)._build_logic_checks_df()
        return {(r['Check'], r['Rule']): r['Result']
                for _, r in frame.iterrows()}

    def test_failures_reported_with_counts(self, messy_df):
        frame = _report(messy_df)._build_logic_checks_df()
        rows = {r['Rule']: r for _, r in frame.iterrows()}
        closed_no_date = next(r for k, r in rows.items() if "no resolved_at" in k)
        assert closed_no_date['Rows Affected'] == 2
        assert closed_no_date['Result'] == "Fail"

    def test_clean_data_reports_pass(self, standard_df):
        assert set(self._verdicts(standard_df).values()) <= {"Pass"}

    def test_missing_date_columns_are_not_checked_not_passed(self, messy_df):
        df = messy_df.drop(columns=["opened_at", "resolved_at"])
        verdicts = self._verdicts(df)
        date_rules = [v for (check, _), v in verdicts.items()
                      if check in ("Date sequence", "Future dates")]
        assert date_rules, "date rules missing from the sheet entirely"
        assert all(v == "Not checked" for v in date_rules)
        assert "Pass" not in date_rules

    def test_missing_state_column_is_not_checked_not_passed(self, messy_df):
        df = messy_df.drop(columns=["state"])
        verdicts = self._verdicts(df)
        state_rules = [v for (check, _), v in verdicts.items()
                       if check in ("State / resolution", "Open ticket priority")]
        assert state_rules
        assert all(v == "Not checked" for v in state_rules)

    def test_missing_priority_column_is_not_checked_not_passed(self, messy_df):
        df = messy_df.drop(columns=["priority"])
        verdicts = self._verdicts(df)
        priority = [v for (check, _), v in verdicts.items()
                    if check == "Open ticket priority"]
        assert priority == ["Not checked"]

    def test_unclassified_states_are_review_not_fail(self, messy_df):
        """An unreadable state value is a limit of the tool's vocabulary, not a
        defect in the data, so it must not be reported as a failure."""
        verdicts = self._verdicts(messy_df)
        classification = [v for (check, _), v in verdicts.items()
                          if check == "State classification"]
        assert classification == ["Review"]

    def test_rows_affected_never_zero_for_a_skipped_rule(self, messy_df):
        """A 0 next to a skipped rule reads as 'nothing wrong here'."""
        df = messy_df.drop(columns=["state"])
        frame = _report(df)._build_logic_checks_df()
        for _, row in frame.iterrows():
            if row['Result'] == "Not checked":
                assert row['Rows Affected'] == "Not checked"


# ──────────────────────────────────────────────────────────────
# State Values sheet
# ──────────────────────────────────────────────────────────────

class TestStateValuesSheet:
    def test_suspects_listed_first_with_a_reason(self, messy_df):
        frame = _report(messy_df)._build_state_values_df()
        assert frame.iloc[0]['State Value'] == "Closd"
        assert "Closed" in frame.iloc[0]['Assessment']

    def test_classification_exposed_per_value(self, messy_df):
        frame = _report(messy_df)._build_state_values_df()
        reads = dict(zip(frame['State Value'], frame['Reads As']))
        assert reads["Closed"] == "Closed"
        assert reads["New"] == "Open"
        assert reads["Stage 9"] == "Unrecognised"

    def test_counts_match_the_data(self, messy_df):
        frame = _report(messy_df)._build_state_values_df()
        counts = dict(zip(frame['State Value'], frame['Count']))
        assert counts["Closed"] == 40
        assert counts["New"] == 18

    def test_empty_without_a_state_column(self, messy_df):
        frame = _report(messy_df.drop(columns=["state"]))._build_state_values_df()
        assert frame.empty


# ──────────────────────────────────────────────────────────────
# Text report
# ──────────────────────────────────────────────────────────────

class TestTextReport:
    def test_cross_field_section_present(self, messy_df):
        text = _report(messy_df).generate_text_report()
        assert "CROSS-FIELD CHECKS" in text

    def test_findings_named_with_the_real_columns(self, messy_df):
        text = _report(messy_df).generate_text_report()
        assert "resolved_at" in text
        assert "Closd" in text

    def test_singular_counts_read_correctly(self, messy_df):
        """A count of 1 must not render as '1 tickets have'."""
        text = _report(messy_df).generate_text_report()
        assert "1 tickets" not in text
        assert "1 rows" not in text

    def test_skips_are_stated_not_silent(self, messy_df):
        text = _report(messy_df.drop(columns=["priority"])).generate_text_report()
        assert "skipped" in text.lower()

    def test_no_state_column_does_not_crash(self, messy_df):
        text = _report(messy_df.drop(columns=["state"])).generate_text_report()
        assert "CROSS-FIELD CHECKS" in text


# ──────────────────────────────────────────────────────────────
# Summary report
# ──────────────────────────────────────────────────────────────

class TestSummaryReport:
    """
    generate_summary_report is the other model-free generator: the same checks
    as the text report, sorted into critical / warning / passed with the actions
    they imply. It is written to be pasted into a mail client, so the two things
    that can break it silently are markup creeping into the output and a check
    that could not run being credited as a pass.
    """

    def test_output_is_plain_text(self, messy_df):
        text = _report(messy_df).generate_summary_report()
        assert isinstance(text, str) and text.strip()
        # No HTML, no Markdown - the destination is a mail body.
        for banned in ("<p>", "<br", "<div", "<table", "<b>", "</", "**",
                       "##", "|---", "```"):
            assert banned not in text, f"markup in a plain-text report: {banned}"

    def test_it_leads_with_a_subject_line(self, messy_df):
        """The first line is what a mail client wants, so it goes first."""
        text = _report(messy_df).generate_summary_report()
        assert text.splitlines()[0].startswith("Subject: ")

    def test_findings_are_named_with_the_real_columns_and_values(self, messy_df):
        text = _report(messy_df).generate_summary_report()
        assert "CRITICAL ISSUES" in text
        assert "WARNINGS" in text
        assert "resolved_at" in text          # closed tickets with no date
        assert "Closd" in text                # the state typo, by value
        assert "REQUESTED ACTIONS" in text

    def test_a_clean_file_says_every_check_passed(self, standard_df):
        text = _report(standard_df).generate_summary_report()
        assert "All Checks Passed" in text
        assert "CRITICAL ISSUES" not in text
        assert "WARNINGS" not in text
        assert "No further action is required" in text

    def test_singular_counts_read_correctly(self, messy_df):
        """A count of 1 must not render as '1 tickets have'."""
        text = _report(messy_df).generate_summary_report()
        assert "1 tickets" not in text
        assert "1 rows" not in text

    def test_a_missing_required_column_is_critical(self, messy_df):
        text = _report(messy_df.drop(columns=["state"])).generate_summary_report()
        assert 'Required column "State/Status" is missing' in text

    def test_no_data_is_said_rather_than_crashed(self):
        assert ReportGenerator(None, None).generate_summary_report() == \
            "No data loaded."


class TestSummaryReportDoesNotPassAnUnrunCheck:
    """
    The same rule the logic-check sheet follows: crediting the data for
    validation nobody performed is worse than saying nothing, because a reader
    cannot tell the difference from the outside.
    """

    def _passed_block(self, text):
        """The PASSED bullets only - a finding elsewhere is not a pass."""
        if "PASSED:" not in text:
            return ""
        return text.split("PASSED:", 1)[1].split("\n\n", 1)[0]

    def test_missing_state_column_earns_no_state_pass(self, messy_df):
        text = _report(messy_df.drop(columns=["state"])).generate_summary_report()
        passed = self._passed_block(text)
        assert "consistent with" not in passed
        assert "consistent vocabulary" not in passed

    def test_missing_date_columns_earn_no_date_pass(self, messy_df):
        df = messy_df.drop(columns=["opened_at", "resolved_at"])
        passed = self._passed_block(_report(df).generate_summary_report())
        assert "Date sequence is valid" not in passed

    def test_missing_priority_column_earns_no_priority_claim(self, messy_df):
        df = messy_df.drop(columns=["priority"])
        text = _report(df).generate_summary_report()
        assert "no \"priority\" value" not in text
        assert "Assign \"priority\"" not in text


class TestSummaryReportLanguageSection:
    """
    The one place the summary and the text report deliberately disagree. The
    text report is a document about the whole file, so an absent language
    section there would read as "nothing found" and it says "Not checked"
    instead. The summary is a mail to a stakeholder: saying nothing about a
    check nobody ran is not a false claim, while "language not checked" is
    noise about a view the reader may never open.
    """

    def test_nothing_is_said_when_no_check_has_run(self, messy_df):
        text = _report(messy_df).generate_summary_report()
        assert "language" not in text.lower()
        assert "English" not in text
        assert "Not checked" not in text, \
            "the summary must not raise a check nobody asked for"

    def test_handed_over_findings_are_reported(self, messy_df, lang_result):
        text = _report(messy_df,
                       lang_results=[lang_result]).generate_summary_report()
        assert "not in English" in text
        assert "description" in text
        # The languages themselves, so the reader knows what they are dealing with
        assert "fr (" in text or "French" in text

    def test_a_check_that_could_not_run_is_a_warning_not_a_pass(self, messy_df):
        """
        check_column returns a `warning` when every detection tier was disabled.
        Reporting that column as English would be the false pass this whole
        rule exists to prevent.
        """
        blocked = {'column': 'description',
                   'warning': 'Lingua is not installed, so lingua_only '
                              'detected nothing'}
        text = _report(messy_df,
                       lang_results=[blocked]).generate_summary_report()
        assert 'could not be checked' in text
        assert 'entirely English' not in text
        assert 'Re-run the language check' in text


# ──────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────

def _sheet_text(path, sheet):
    """All non-empty cell text on one sheet, for locating a stacked block."""
    from openpyxl import load_workbook
    worksheet = load_workbook(path)[sheet]
    return "\n".join(str(c.value) for row in worksheet.iter_rows()
                     for c in row if c.value is not None)


class TestExcelExport:
    """
    The export used to be one sheet per table. It is now six tabs with the
    findings tables stacked as blocks on All Details, so these assert the
    *content at its new location* rather than a sheet name.
    """

    def test_the_consolidated_tabs_are_written(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        assert _report(messy_df).export_to_excel(str(out)) is True
        sheets = pd.ExcelFile(out).sheet_names
        for tab in ("Overview", "Dashboard", "Findings", "All Details", "Pivots"):
            assert tab in sheets

    def test_the_findings_tables_are_blocks_on_all_details(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        _report(messy_df).export_to_excel(str(out))
        text = _sheet_text(out, "All Details")
        for block in ("Required columns", "Column health", "Logic checks",
                      "State vocabulary", "Description quality",
                      "Recurring issues", "Distributions", "Monthly inflow"):
            assert block in text, f"{block} block is missing"

    def test_a_missing_state_column_yields_a_not_checked_block(self, messy_df,
                                                              tmp_path):
        """
        Previously the sheet was omitted entirely. A stated gap is better than an
        invisible one, so the block heading stays and says why it is empty.
        """
        out = tmp_path / "report.xlsx"
        df = messy_df.drop(columns=["state"])
        assert _report(df).export_to_excel(str(out)) is True
        text = _sheet_text(out, "All Details")
        assert "State vocabulary" in text
        assert "Not checked" in text

    def test_disclosures_are_reported_back_to_the_caller(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        reporter = _report(messy_df.drop(columns=["opened_at", "resolved_at"]))
        reporter.export_to_excel(str(out))
        assert reporter.export_disclosures
        assert any("Monthly inflow" in note
                   for note in reporter.export_disclosures)

    def test_native_pivots_can_be_switched_off(self, messy_df, tmp_path):
        from openpyxl import load_workbook
        out = tmp_path / "report.xlsx"
        reporter = _report(messy_df)
        reporter.native_pivots = False
        reporter.export_to_excel(str(out))
        assert load_workbook(out)["Pivots"]._pivots == []


# ──────────────────────────────────────────────────────────────
# Language findings reach the report
# ──────────────────────────────────────────────────────────────

@pytest.fixture
def lang_result():
    """A check_column() result as the GUI would hand it over."""
    from core.language import LanguageChecker
    df = pd.DataFrame({"description": [
        "Le serveur ne repond plus depuis ce matin merci",
        "Der Drucker funktioniert nicht mehr richtig heute",
        "The printer is offline and refusing all print jobs",
        "Привет мир это русский текст здесь для теста",
        "tiny",
    ]})
    return LanguageChecker().check_column(df, "description", mode="all")


class TestLanguageInTextReport:
    """
    Language analysis never reached either report: the GUI did not pass
    lang_columns, so the Excel sheet was skipped and the text report had no
    language section at all.
    """

    def test_section_is_present(self, messy_df, lang_result):
        text = _report(messy_df, lang_results=[lang_result]).generate_text_report()
        assert "LANGUAGE CHECK" in text

    def test_languages_are_named_with_percentages(self, messy_df, lang_result):
        text = _report(messy_df, lang_results=[lang_result]).generate_text_report()
        assert "French" in text
        assert "% of analyzed" in text

    def test_detection_method_is_stated(self, messy_df, lang_result):
        """A count means nothing without knowing which methods produced it."""
        text = _report(messy_df, lang_results=[lang_result]).generate_text_report()
        assert "Detection method:" in text

    def test_an_unrun_check_says_so_rather_than_being_omitted(self, messy_df):
        text = _report(messy_df).generate_text_report()
        assert "LANGUAGE CHECK" in text, "the section must exist even with no result"
        assert "Not checked" in text
        # and it must tell the reader how to include it
        assert "Run Check" in text

    def test_a_check_that_could_not_run_is_not_reported_clean(self, messy_df):
        from core.language import LanguageChecker
        checker = LanguageChecker()
        checker._lingua_available = False
        result = checker.check_column(
            pd.DataFrame({"d": ["Le serveur ne repond plus du tout"]}), "d",
            mode="lingua_only")
        text = _report(messy_df, lang_results=[result]).generate_text_report()
        assert "NOT CHECKED" in text
        assert "appear to be English" not in text

    def test_skipped_short_rows_are_disclosed(self, messy_df, lang_result):
        text = _report(messy_df, lang_results=[lang_result]).generate_text_report()
        assert "skipped as too short" in text

    def test_examples_are_given_so_the_finding_is_actionable(self, messy_df, lang_result):
        """A count with no examples leaves the reader nothing to act on."""
        text = _report(messy_df, lang_results=[lang_result]).generate_text_report()
        assert "Examples:" in text
        assert "Le serveur" in text or "Der Drucker" in text

    def test_examples_carry_ticket_ids(self, lang_result):
        """The examples must be findable in the source file."""
        df = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(5)],
            "description": [
                "Le serveur ne repond plus depuis ce matin merci",
                "Der Drucker funktioniert nicht mehr richtig heute",
                "The printer is offline and refusing all print jobs",
                "Привет мир это русский текст здесь для теста",
                "tiny",
            ],
        })
        from core.language import LanguageChecker
        result = LanguageChecker().check_column(df, "description", mode="all")
        text = _report(df, lang_results=[result]).generate_text_report()
        assert "INC000" in text.split("Examples:")[1]

    def test_detection_methods_state_their_reliability(self, messy_df, lang_result):
        """A keyword guess and a script match are not equally trustworthy."""
        text = _report(messy_df, lang_results=[lang_result]).generate_text_report()
        assert "definitive" in text or "heuristic" in text

    def test_a_missing_id_column_does_not_break_examples(self, lang_result):
        df = pd.DataFrame({"description": [
            "Le serveur ne repond plus depuis ce matin merci",
            "The printer is offline and refusing all print jobs",
        ]})
        from core.language import LanguageChecker
        result = LanguageChecker().check_column(df, "description", mode="all")
        text = _report(df, lang_results=[result]).generate_text_report()
        assert "Examples:" in text


class TestLanguageReachesTheSummary:
    """
    The SUMMARY block is built from analyzer.get_summary(), which has no
    language awareness - so a file whose only defect was non-English text
    printed "No major issues found" directly below the section reporting it.
    """

    @pytest.fixture
    def clean_but_foreign(self):
        n = 6
        return pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": ["Closed"] * n,
            "priority": ["P1"] * n,
            "opened_at": pd.to_datetime(["2024-01-01"] * n),
            "resolved_at": pd.to_datetime(["2024-02-01"] * n),
            "assignment_group": ["Team"] * n,
            "cmdb_ci": ["CI"] * n,
            "short_description": ["Login issue for the user account"] * n,
            "description": [
                "Le serveur ne repond plus depuis ce matin merci beaucoup",
                "Der Drucker funktioniert nicht mehr richtig seit heute",
                "Impresora no funciona correctamente hoy dia gracias",
                "The printer is offline and refusing all print jobs now",
                "Network connectivity dropped on the third floor switch",
                "User cannot log in to the finance application at all",
            ],
        })

    def _summary(self, text):
        return text.split("           SUMMARY")[1]

    def test_language_issue_appears_in_the_summary(self, clean_but_foreign):
        from core.language import LanguageChecker
        result = LanguageChecker().check_column(clean_but_foreign, "description", mode="all")
        assert result['non_english_count'] > 0, "fixture must contain non-English text"

        text = _report(clean_but_foreign, lang_results=[result]).generate_text_report()
        summary = self._summary(text)
        assert "No major issues found" not in summary
        assert "not in English" in summary

    def test_clean_english_leaves_the_summary_clean(self, clean_but_foreign):
        df = clean_but_foreign.copy()
        df["description"] = [
            "The printer is offline and refusing all print jobs now",
            "Network connectivity dropped on the third floor switch",
            "User cannot log in to the finance application at all",
            "Laptop battery drains within an hour of unplugging it",
            "Email client crashes when opening a large attachment",
            "Shared drive mapping disappears after every restart",
        ]
        from core.language import LanguageChecker
        result = LanguageChecker().check_column(df, "description", mode="all")
        summary = self._summary(_report(df, lang_results=[result]).generate_text_report())
        assert "not in English" not in summary

    def test_unrunnable_check_is_flagged_in_the_summary(self, clean_but_foreign):
        from core.language import LanguageChecker
        checker = LanguageChecker()
        checker._lingua_available = False
        result = checker.check_column(clean_but_foreign, "description", mode="lingua_only")
        summary = self._summary(
            _report(clean_but_foreign, lang_results=[result]).generate_text_report())
        assert "NOT checked" in summary
        assert "No major issues found" not in summary


class TestLanguageSheets:
    """
    The two language tables are now blocks on All Details. Their *builders* are
    still asserted directly, since those carry the contracts that matter.
    """

    def test_both_blocks_appear_when_a_check_was_run(self, messy_df, tmp_path,
                                                    lang_result):
        out = tmp_path / "report.xlsx"
        _report(messy_df, lang_results=[lang_result]).export_to_excel(str(out))
        text = _sheet_text(out, "All Details")
        assert "Language summary" in text
        assert "Language detail" in text

    def test_blocks_say_not_checked_when_nothing_ran(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        _report(messy_df).export_to_excel(str(out))
        text = _sheet_text(out, "All Details")
        # The heading stays and states the gap; it is not silently dropped.
        assert "Language summary" in text
        assert "Not checked" in text

    def test_detail_carries_per_language_percentages(self, messy_df, lang_result):
        detail = _report(messy_df,
                         lang_results=[lang_result])._build_language_detail_df()
        assert "% of Analyzed" in detail.columns
        assert "Rows" in detail.columns
        assert "French" in set(detail["Language"])
        # percentages must be a real share, not a copy of the count
        row = detail[detail["Language"] == "French"].iloc[0]
        assert 0 < row["% of Analyzed"] <= 100

    def test_summary_states_the_detection_method(self, messy_df, lang_result):
        summary = _report(messy_df,
                          lang_results=[lang_result])._build_language_summary_df()
        assert "Detection Method" in summary.columns
        assert summary.iloc[0]["Status"] == "Checked"

    def test_unrunnable_check_is_not_checked_not_zero(self, messy_df):
        from core.language import LanguageChecker
        checker = LanguageChecker()
        checker._lingua_available = False
        result = checker.check_column(
            pd.DataFrame({"d": ["Le serveur ne repond plus du tout"]}), "d",
            mode="lingua_only")
        summary = _report(messy_df,
                          lang_results=[result])._build_language_summary_df()
        assert summary.iloc[0]["Status"] == "Not checked"
        # Must survive as the string, so this column may never go numeric.
        assert summary.iloc[0]["Non-English Rows"] == "Not checked"

    def test_results_are_reused_not_recomputed(self, messy_df, mocker):
        """
        Detection costs ~26s per 50k rows with Lingua, so an export must not
        redo work the user already paid for.
        """
        reporter = _report(messy_df, lang_results=[{
            'column': 'description', 'total_analyzed': 5, 'total_original': 5,
            'skipped_short': 0, 'non_english_count': 1, 'percentage': 20.0,
            'languages': {'French': 1}, 'methods': {'keywords': 1},
            'samples': [], 'mode': 'all', 'mode_label': 'characters + keywords + Lingua',
            'rules_used': True, 'using_lingua': True, 'lingua_available': True,
            'min_length': 10,
        }])
        spy = mocker.spy(reporter.lang_checker, "check_column")
        reporter.generate_text_report()
        assert spy.call_count == 0

    def test_explicit_lang_columns_still_computes(self, messy_df):
        """The opt-in path keeps working for callers that want a fresh run."""
        reporter = _report(messy_df, lang_columns=["short_description"],
                           lang_mode="rules_only")
        results = reporter._language_results()
        assert len(results) == 1
        assert results[0]['column'] == "short_description"
        assert results[0]['using_lingua'] is False


# ──────────────────────────────────────────────────────────────
# Findings register
# ──────────────────────────────────────────────────────────────

class TestFindingsRegister:
    """
    One row per check, so a reader can see failures and coverage gaps in one
    place. The ordering matters as much as the content: "Not checked" outranks
    "Review" and "Pass" because an unexamined dimension is the easiest thing to
    miss.
    """

    HEADERS = ['Area', 'Check', 'Severity', 'Rows Affected', 'Detail', 'Result']

    def test_headers(self, messy_df):
        assert list(_report(messy_df)._build_findings_df().columns) == self.HEADERS

    def test_failures_sort_above_passes(self, messy_df):
        frame = _report(messy_df)._build_findings_df()
        results = list(frame['Result'])
        assert results[0] == "Fail"
        assert results.index("Fail") < results.index("Pass")

    def test_not_checked_sorts_above_review_and_pass(self, messy_df):
        """A gap in coverage must not be buried under a wall of passes."""
        df = messy_df.drop(columns=["opened_at", "resolved_at"])
        results = list(_report(df)._build_findings_df()['Result'])
        assert "Not checked" in results
        first_unchecked = results.index("Not checked")
        for later in ("Review", "Pass"):
            if later in results:
                assert first_unchecked < results.index(later)

    def test_a_skipped_rule_reports_not_checked_not_zero(self, messy_df):
        df = messy_df.drop(columns=["opened_at", "resolved_at"])
        frame = _report(df)._build_findings_df()
        unchecked = frame[frame['Result'] == "Not checked"]
        assert not unchecked.empty
        assert (unchecked['Rows Affected'] == "Not checked").all()

    def test_missing_required_column_is_critical(self, messy_df):
        frame = _report(messy_df.drop(columns=["priority"]))._build_findings_df()
        structure = frame[frame['Area'] == "Structure"]
        assert (structure['Severity'] == "Critical").any()
        assert (structure['Result'] == "Fail").any()

    def test_passes_carry_no_severity(self, messy_df):
        frame = _report(messy_df)._build_findings_df()
        passes = frame[frame['Result'] == "Pass"]
        assert not passes.empty
        assert set(passes['Severity']) == {"-"}

    def test_logic_rows_agree_with_the_logic_block(self, messy_df):
        """The register must not re-derive what _build_logic_checks_df says."""
        reporter = _report(messy_df)
        logic = reporter._build_logic_checks_df()
        register = reporter._build_findings_df()
        register_logic = register[register['Area'] == "Logic"]

        assert len(register_logic) == len(logic)
        by_rule = dict(zip(logic['Rule'], logic['Result']))
        for _, row in register_logic.iterrows():
            assert by_rule[row['Detail']] == row['Result']

    def test_an_unrun_language_check_is_declared(self, messy_df):
        frame = _report(messy_df)._build_findings_df()
        language = frame[frame['Area'] == "Language"]
        assert (language['Result'] == "Not checked").all()

    def test_a_language_result_is_reported(self, messy_df, lang_result):
        frame = _report(messy_df, lang_results=[lang_result])._build_findings_df()
        language = frame[frame['Area'] == "Language"]
        assert (language['Result'] == "Fail").any()

    def test_duplicates_reach_the_register(self, messy_df):
        df = messy_df.copy()
        df.loc[df.index[1], "number"] = df.loc[df.index[0], "number"]
        frame = _report(df)._build_findings_df()
        dupes = frame[frame['Area'] == "Uniqueness"]
        assert dupes.iloc[0]['Result'] == "Fail"

    def test_quality_score_row_counts_rows_not_the_score(self, messy_df):
        """'Rows Affected' of 1 next to a score of 50 would read as one bad row."""
        reporter = _report(messy_df)
        frame = reporter._build_findings_df()
        score_rows = frame[frame['Check'].str.contains("quality score")]
        quality = reporter.analyzer.check_description_quality(20, 5000)
        expected = {
            info['column']: sum(info['score_distribution'][t]
                                for t in ('Critical', 'Poor', 'Fair'))
            for info in quality['columns']
        }
        for _, row in score_rows.iterrows():
            column = row['Check'].replace(" quality score", "")
            assert row['Rows Affected'] == expected[column]


class TestColumnHealth:
    def test_one_row_per_column(self, messy_df):
        frame = _report(messy_df)._build_column_health_df()
        assert len(frame) == len(messy_df.columns)

    def test_merges_nulls_and_profile(self, messy_df):
        frame = _report(messy_df)._build_column_health_df()
        for header in ('Type', 'Sample Value', 'Nulls', 'Null %',
                       'Above Threshold', 'Padded Whitespace'):
            assert header in frame.columns

    def test_null_counts_match_check_nulls(self, messy_df):
        reporter = _report(messy_df)
        nulls = reporter.analyzer.check_nulls(reporter.null_threshold)
        frame = reporter._build_column_health_df().set_index('Column')
        for col, info in nulls.items():
            assert frame.loc[col, 'Nulls'] == info['count']
            assert frame.loc[col, 'Above Threshold'] == (
                'Yes' if info['above_threshold'] else 'No')

    def test_matched_requirement_is_named_for_detected_columns(self, messy_df):
        frame = _report(messy_df)._build_column_health_df().set_index('Column')
        assert frame.loc['state', 'Matched Requirement'] == 'State/Status'

    def test_an_unmatched_column_is_blank_not_guessed(self, messy_df):
        df = messy_df.copy()
        df['unrelated_note'] = "text"
        frame = _report(df)._build_column_health_df().set_index('Column')
        assert frame.loc['unrelated_note', 'Matched Requirement'] == ''


class TestMonthlyInflow:
    def test_shares_sum_to_about_one_hundred(self, messy_df):
        frame = _report(messy_df)._build_monthly_inflow_df()
        assert not frame.empty
        assert frame['Share %'].sum() == pytest.approx(100.0, abs=0.5)

    def test_tickets_sum_to_the_dated_rows(self, messy_df):
        frame = _report(messy_df)._build_monthly_inflow_df()
        assert frame['Tickets'].sum() == len(messy_df)

    def test_no_date_column_yields_an_empty_frame(self, messy_df):
        df = messy_df.drop(columns=["opened_at", "resolved_at"])
        assert _report(df)._build_monthly_inflow_df().empty


class TestDistributions:
    def test_covers_more_than_the_old_four_keywords(self, messy_df):
        frame = _report(messy_df)._build_distributions_df()
        assert set(frame['Column']) > {"state", "priority"}

    def test_high_cardinality_column_is_declared_not_expanded(self, messy_df):
        df = messy_df.copy()
        df['free_text'] = [f"unique note {i}" for i in range(len(df))]
        frame = _report(df)._build_distributions_df()
        free_text = frame[frame['Column'] == 'free_text']
        assert len(free_text) == 1, "a free-text column must not be expanded"
        assert "not summarised" in free_text.iloc[0]['Value']

    def test_share_never_exceeds_one_hundred(self, messy_df):
        """Summing individually rounded percentages produced 100.2%."""
        df = messy_df.copy()
        df['free_text'] = [f"unique note {i}" for i in range(len(df))]
        frame = _report(df)._build_distributions_df()
        assert (frame['Share %'] <= 100.0).all()

    def test_values_are_capped_per_column_with_the_remainder_kept(self):
        n = 120
        df = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": ["Closed"] * n,
            "priority": ["P1"] * n,
            "opened_at": pd.to_datetime(["2024-01-01"] * n),
            "resolved_at": pd.to_datetime(["2024-02-01"] * n),
            "assignment_group": [f"Team {i % 40}" for i in range(n)],
            "cmdb_ci": ["CI"] * n,
            "short_description": ["Login issue for the user account"] * n,
            "description": ["A sufficiently detailed description"] * n,
        })
        frame = _report(df)._build_distributions_df()
        groups = frame[frame['Column'] == 'assignment_group']
        # 40 distinct: 20 shown plus one remainder row
        assert len(groups) == 21
        assert "more values" in groups.iloc[-1]['Value']
        assert groups['Count'].sum() == n, "the remainder must keep its real count"


class TestPivotCube:
    def test_counts_sum_to_the_frame(self, messy_df):
        frame = _report(messy_df)._build_pivot_cube_df()
        assert frame['Ticket count'].sum() == len(messy_df)

    def test_dimension_count_is_capped(self, messy_df):
        from core.reporter import MAX_CUBE_DIMENSIONS
        frame = _report(messy_df)._build_pivot_cube_df()
        dimensions = [c for c in frame.columns if c != 'Ticket count']
        assert 0 < len(dimensions) <= MAX_CUBE_DIMENSIONS

    def test_free_text_is_never_a_dimension(self, messy_df):
        df = messy_df.copy()
        df['free_text'] = [f"unique note {i}" for i in range(len(df))]
        frame = _report(df)._build_pivot_cube_df()
        assert 'free_text' not in frame.columns

    def test_dimension_values_are_strings(self, messy_df):
        """A cacheField name must match its header, and NaN would be blank."""
        frame = _report(messy_df)._build_pivot_cube_df()
        for col in (c for c in frame.columns if c != 'Ticket count'):
            assert frame[col].map(type).eq(str).all()


class TestTicketDetail:
    def test_one_row_per_ticket_and_all_original_columns(self, messy_df):
        frame = _report(messy_df)._build_ticket_detail_df()
        assert len(frame) == len(messy_df)
        for col in messy_df.columns:
            assert col in frame.columns

    def test_flags_agree_with_the_checks(self, messy_df):
        """The reason get_row_flags exists: one source for count and row list."""
        reporter = _report(messy_df)
        frame = reporter._build_ticket_detail_df()
        xfield = reporter.analyzer.check_cross_field_logic()

        assert (frame['Closed with no resolution date'] == 'Yes').sum() == \
            xfield['closed_no_date']
        assert (frame['Open with resolution date'] == 'Yes').sum() == \
            xfield['open_with_date']
        assert (frame['Open with no priority'] == 'Yes').sum() == \
            xfield['open_missing_priority']

    def test_a_skipped_rule_says_not_checked_not_blank(self, messy_df):
        """A pivot on 'Not checked' tells the truth; a pivot on blanks lies."""
        df = messy_df.drop(columns=["priority"])
        frame = _report(df)._build_ticket_detail_df()
        assert set(frame['Open with no priority']) == {'Not checked'}

    def test_qa_flag_count_matches_the_yes_columns(self, messy_df):
        from core.reporter import _FLAG_LABELS
        frame = _report(messy_df)._build_ticket_detail_df()
        labels = [l for l in _FLAG_LABELS.values() if l in frame.columns]
        expected = sum((frame[l] == 'Yes').astype(int) for l in labels)
        assert list(frame['QA Flag Count']) == list(expected)

    def test_qa_findings_names_every_flag_it_counts(self, messy_df):
        frame = _report(messy_df)._build_ticket_detail_df()
        for _, row in frame.iterrows():
            if row['QA Flag Count'] == 0:
                assert row['QA Findings'] == 'None'
            else:
                assert len(row['QA Findings'].split(', ')) == row['QA Flag Count']

    def test_description_scores_are_included_and_agree(self, messy_df):
        reporter = _report(messy_df)
        frame = reporter._build_ticket_detail_df()
        scores = reporter.analyzer.get_description_row_scores(20, 5000)
        for col, entry in scores.items():
            assert list(frame[f'{col} Score']) == list(entry['score'])
            assert list(frame[f'{col} Tier']) == list(entry['tier'])

    def test_resolution_days_keeps_negatives(self):
        """A negative resolution time IS the closed-before-created defect."""
        df = pd.DataFrame({
            "number": ["INC1", "INC2"],
            "state": ["Closed", "Closed"],
            "priority": ["P1", "P1"],
            "opened_at": pd.to_datetime(["2024-03-01", "2024-01-01"]),
            "resolved_at": pd.to_datetime(["2024-01-01", "2024-01-05"]),
            "assignment_group": ["Team", "Team"],
            "cmdb_ci": ["CI", "CI"],
            "short_description": ["Login issue for the user"] * 2,
            "description": ["A sufficiently detailed description"] * 2,
        })
        frame = _report(df)._build_ticket_detail_df()
        assert frame['Resolution Days'].iloc[0] < 0
        assert frame['Resolution Days'].iloc[1] > 0

    def test_state_reads_as_uses_one_vocabulary(self, messy_df):
        frame = _report(messy_df)._build_ticket_detail_df()
        assert set(frame['State Reads As']) <= {'Open', 'Closed', 'Unrecognised'}


class TestPerRowLanguage:
    """
    Four distinct values, so nothing is inferred: a row nobody looked at must
    never read as English.
    """

    @pytest.fixture
    def mixed_df(self):
        n = 8
        return pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": ["Closed"] * n,
            "priority": ["P1"] * n,
            "opened_at": pd.to_datetime(["2024-01-01"] * n),
            "resolved_at": pd.to_datetime(["2024-02-01"] * n),
            "assignment_group": ["Team"] * n,
            "cmdb_ci": ["CI"] * n,
            "short_description": ["Login issue for the user account"] * n,
            "description": [
                "Le serveur ne repond plus depuis ce matin merci",
                "Der Drucker funktioniert nicht mehr richtig heute",
                "The printer is offline and refusing all print jobs",
                "Impresora no funciona correctamente hoy dia gracias",
                "tiny",
                "abc",
                "Laptop battery drains within an hour of unplugging",
                "El sistema no responde cuando intento abrir archivo",
            ],
        })

    def _checked(self, df, **kwargs):
        from core.language import LanguageChecker
        checker = LanguageChecker()
        if kwargs.pop("no_lingua", False):
            checker._lingua_available = False
        return checker.check_column(df, "description", **kwargs)

    def test_the_column_appears_only_when_a_check_ran(self, mixed_df):
        without = _report(mixed_df)._build_ticket_detail_df()
        assert "description Language" not in without.columns

        result = self._checked(mixed_df, mode="all")
        with_check = _report(mixed_df,
                             lang_results=[result])._build_ticket_detail_df()
        assert "description Language" in with_check.columns

    def test_detected_languages_land_on_the_right_rows(self, mixed_df):
        result = self._checked(mixed_df, mode="all")
        frame = _report(mixed_df, lang_results=[result])._build_ticket_detail_df()
        languages = list(frame["description Language"])
        assert languages[0] == "French"
        assert languages[1] == "German"
        assert languages[2] == "English"

    def test_short_rows_say_so_rather_than_english(self, mixed_df):
        result = self._checked(mixed_df, mode="all")
        frame = _report(mixed_df, lang_results=[result])._build_ticket_detail_df()
        languages = list(frame["description Language"])
        assert languages[4] == "Too short to check"
        assert languages[5] == "Too short to check"

    def test_never_blank(self, mixed_df):
        result = self._checked(mixed_df, mode="all")
        frame = _report(mixed_df, lang_results=[result])._build_ticket_detail_df()
        assert frame["description Language"].notna().all()
        assert (frame["description Language"].astype(str).str.strip() != "").all()

    def test_an_unrunnable_check_marks_every_row_not_checked(self, mixed_df):
        result = self._checked(mixed_df, mode="lingua_only", no_lingua=True)
        frame = _report(mixed_df, lang_results=[result])._build_ticket_detail_df()
        assert set(frame["description Language"]) == {"Not checked"}

    def test_a_handbuilt_result_without_the_keys_is_tolerated(self, mixed_df):
        """test_reporter and older callers construct these dicts by hand."""
        minimal = {
            'column': 'description', 'total_analyzed': 6, 'total_original': 8,
            'skipped_short': 2, 'non_english_count': 1, 'percentage': 16.67,
            'languages': {'French': 1}, 'methods': {'keywords': 1},
            'samples': [], 'mode': 'all', 'mode_label': 'all methods',
            'rules_used': True, 'using_lingua': True, 'lingua_available': True,
            'min_length': 10,
        }
        frame = _report(mixed_df, lang_results=[minimal])._build_ticket_detail_df()
        assert "description Language" not in frame.columns

    def test_it_costs_no_extra_detection(self, mixed_df, mocker):
        """The per-row answer is a by-product, never a second pass."""
        result = self._checked(mixed_df, mode="all")
        reporter = _report(mixed_df, lang_results=[result])
        spy = mocker.spy(reporter.lang_checker, "detect_language_series")
        reporter._build_ticket_detail_df()
        assert spy.call_count == 0


class TestNewContentReachesTheWorkbook:
    def test_the_findings_register_gets_its_own_tab(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        _report(messy_df).export_to_excel(str(out))
        assert "Findings" in pd.ExcelFile(out).sheet_names

    def test_the_new_tables_are_blocks_on_all_details(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        _report(messy_df).export_to_excel(str(out))
        text = _sheet_text(out, "All Details")
        for block in ("Column health", "Monthly inflow", "Distributions"):
            assert block in text

    def test_the_cube_lives_on_a_hidden_sheet(self, messy_df, tmp_path):
        from openpyxl import load_workbook
        out = tmp_path / "report.xlsx"
        _report(messy_df).export_to_excel(str(out))
        wb = load_workbook(out)
        assert wb["PivotData"].sheet_state == "hidden"

    def test_export_survives_a_file_with_no_dates(self, messy_df, tmp_path):
        out = tmp_path / "report.xlsx"
        df = messy_df.drop(columns=["opened_at", "resolved_at"])
        reporter = _report(df)
        assert reporter.export_to_excel(str(out)) is True
        # Stated rather than silently absent
        assert any("Monthly inflow" in note
                   for note in reporter.export_disclosures)
