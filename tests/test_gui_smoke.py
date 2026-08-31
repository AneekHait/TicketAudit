"""
Structural smoke test for the GUI.

Runs headless via QT_QPA_PLATFORM=offscreen. It exists because most GUI
breakage in this app is silent - a view that fails to build, a table with no
model, a stylesheet property Qt quietly drops. Visiting every view and
asserting it populates catches that in one place.
"""
import os
import sys

import pandas as pd
import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

pytest.importorskip("PySide6")

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication, QAbstractItemView, QTableView

from config import ConfigManager
from core.analyzer import SanityAnalyzer
import gui.app_pyside as ui


@pytest.fixture(scope="module")
def qapp():
    app = QApplication.instance() or QApplication(sys.argv)
    app.setStyle("Fusion")
    ui.apply_dark_palette(app)
    return app


@pytest.fixture
def ticket_df():
    n = 8
    return pd.DataFrame({
        "number": [f"INC{i:04d}" for i in range(1, n + 1)],
        "assignment_group": ["Service Desk", "Network"] * 4,
        "cmdb_ci": ["SRV-01", "LAP-02"] * 4,
        "priority": ["P1", "P2", "P3", "P2"] * 2,
        "state": ["Open", "Closed"] * 4,
        "opened_at": pd.to_datetime([f"2024-0{(i % 6) + 1}-15" for i in range(n)]),
        "resolved_at": pd.to_datetime([f"2024-0{(i % 6) + 1}-20" for i in range(n)]),
        "short_description": ["Password reset needed for the account"] * 4
                             + ["Printer offline on floor three refusing jobs"] * 4,
        "description": ["A sufficiently long description for this row"] * n,
        "mostly_empty": [None] * 6 + ["x", "y"],
    })


@pytest.fixture
def window(qapp, tmp_path, monkeypatch, ticket_df):
    """A window whose config is isolated to tmp_path, never the real one."""
    cfg = str(tmp_path / "config.json")
    monkeypatch.setattr(ui, "ConfigManager", lambda *a, **k: ConfigManager(cfg))

    win = ui.SanityCheckApp()
    win.df = ticket_df
    win.analyzer = SanityAnalyzer(ticket_df)
    win.filepath = str(tmp_path / "tickets.xlsx")
    win._set_content_enabled(True)

    yield win

    win.threadpool.waitForDone(10000)
    win.deleteLater()


def _view_rows(win):
    return [(i, win.nav_list.item(i).data(Qt.ItemDataRole.UserRole))
            for i in range(win.nav_list.count())
            if win.nav_list.item(i).data(Qt.ItemDataRole.UserRole)]


def _settle(qapp, win, rounds=15):
    for _ in range(rounds):
        qapp.processEvents()
        win.threadpool.waitForDone(5000)
        qapp.processEvents()


class TestShell:
    def test_every_registered_view_is_reachable(self, window):
        registered = {n for n, _ in window.tab_defs}
        navigable = {n for _, n in _view_rows(window)}
        assert navigable == registered - ui.NON_NAV_VIEWS

    def test_nav_group_headers_are_not_selectable(self, window):
        headers = [window.nav_list.item(i)
                   for i in range(window.nav_list.count())
                   if window.nav_list.item(i).data(Qt.ItemDataRole.UserRole) is None]
        assert headers, "expected group headers"
        for item in headers:
            assert item.flags() == Qt.ItemFlag.NoItemFlags

    def test_view_pages_occupy_stack_indices_after_empty_state(self, window):
        # Page 0 is reserved for the empty state
        assert set(window.view_indices.values()) == set(
            range(1, len(window.view_indices) + 1))

    def test_empty_state_stays_interactive_when_content_disabled(self, window):
        window.df = None
        window._set_content_enabled(False)
        assert window.stack.widget(0).isEnabled(), \
            "empty state must stay live or its Browse button is dead"

    def test_export_is_gated_on_data(self, window):
        window.df = None
        window._set_content_enabled(False)
        assert not window.btn_export.isEnabled()
        assert not window.export_action.isEnabled()


class TestEveryViewBuildsAndPopulates:
    def test_visit_all_views(self, qapp, window):
        failures = []
        for row, name in _view_rows(window):
            window.nav_list.setCurrentRow(row)
            _settle(qapp, window)

            page = window.tab_widgets[name]
            if page.layout() is None:
                failures.append(f"{name}: never built")
            elif window.stack.currentIndex() != window.view_indices[name]:
                failures.append(f"{name}: not the visible page")
            elif window.view_title.text() != name:
                failures.append(f"{name}: title not updated")
        assert not failures, failures

    def test_tables_have_models_with_rows(self, qapp, window):
        for row, name in _view_rows(window):
            window.nav_list.setCurrentRow(row)
            _settle(qapp, window)

        # Overview's two tables are behind its Run button - the register runs
        # every check in the app, so opening a file does not start it. Pressed
        # here rather than added to on_demand below, so they stay covered.
        window._select_view("Overview")
        _settle(qapp, window)
        window.btn_overview_run.click()
        _settle(qapp, window, rounds=8)

        # Tables that only fill on an explicit user action, so an empty model
        # after merely visiting the view is correct, not a bug:
        #   dup_rows_table/dup_ids_table       - no duplicates in the fixture
        #   lang_breakdown_table/samples_table - need a "Run Check" click, and
        #                                        stay hidden when all-English
        #   lang_detail_table                  - needs a "Run Check" click
        on_demand = {"dup_rows_table", "dup_ids_table", "lang_breakdown_table",
                     "lang_samples_table", "lang_detail_table"}

        empty = []
        for attr in dir(window):
            if not attr.endswith("_table") or attr in on_demand:
                continue
            view = getattr(window, attr, None)
            if not isinstance(view, QTableView):
                continue
            model = view.model()
            if model is None or model.rowCount() == 0:
                empty.append(attr)
        assert not empty, f"tables with no data: {empty}"


class TestTableConventions:
    def test_all_tables_read_only_and_consistent(self, qapp, window):
        for row, _ in _view_rows(window):
            window.nav_list.setCurrentRow(row)
            _settle(qapp, window, rounds=5)

        checked = 0
        for attr in dir(window):
            view = getattr(window, attr, None)
            if not attr.endswith("_table") or not isinstance(view, QTableView):
                continue
            checked += 1
            assert view.editTriggers() == QAbstractItemView.EditTrigger.NoEditTriggers, \
                f"{attr} is editable"
            assert view.alternatingRowColors(), f"{attr} missing alternating rows"
        assert checked >= 6, f"only found {checked} tables"


class TestRefreshGuards:
    """
    _refresh_async exists because a slow analysis that finishes after the user
    has switched sheets used to paint the previous sheet's numbers under the
    current sheet's name. Reproduced before the fix: sheet B's "3 empty" was
    overwritten by sheet A's "40 empty".
    """

    def _run(self, window, key="View", result="payload", seq_bump=0):
        """Dispatch through _refresh_async, optionally invalidating mid-flight."""
        seen = []
        got_analyzer = []

        def task(analyzer):
            got_analyzer.append(analyzer)
            return result

        window._refresh_async(key, task, seen.append)
        window._data_seq += seq_bump
        window.threadpool.waitForDone(5000)
        QApplication.instance().processEvents()
        return seen, got_analyzer

    def test_result_is_delivered_normally(self, qapp, window):
        seen, _ = self._run(window)
        assert seen == ["payload"]

    def test_task_receives_the_bound_analyzer(self, qapp, window):
        """A closure over self.analyzer would resolve when the worker runs."""
        _, got = self._run(window)
        assert got and got[0] is window.analyzer

    def test_stale_result_is_discarded(self, qapp, window):
        """The sequence bump stands in for a sheet load completing mid-flight."""
        seen, _ = self._run(window, seq_bump=1)
        assert seen == [], "a result from a previous sheet must not be painted"

    def test_duplicate_dispatch_is_dropped_while_in_flight(self, qapp, window):
        calls = []

        def task(analyzer):
            calls.append(1)
            return "x"

        window._refresh_inflight.add("Busy View")
        window._refresh_async("Busy View", task, lambda r: None)
        window.threadpool.waitForDone(5000)
        assert calls == [], "already-running view should not be queued again"

    def test_inflight_is_released_after_completion(self, qapp, window):
        self._run(window, key="Released")
        assert "Released" not in window._refresh_inflight

    def test_inflight_is_released_after_error(self, qapp, window):
        errors = []

        def boom(analyzer):
            raise RuntimeError("nope")

        window._refresh_async("Broken", boom, lambda r: None,
                              err_callback=errors.append)
        window.threadpool.waitForDone(5000)
        qapp.processEvents()
        assert "Broken" not in window._refresh_inflight, \
            "a failed refresh must not block the view forever"
        assert errors, "error callback should have run"

    def test_no_dispatch_without_data(self, qapp, window):
        calls = []
        window.df = None
        window.analyzer = None
        window._refresh_async("View", lambda a: calls.append(1), lambda r: None)
        window.threadpool.waitForDone(2000)
        assert calls == []

    def test_lock_is_per_analyzer_not_global(self, qapp, window):
        """
        A shared lock made the new sheet's refresh queue behind the old sheet's
        unfinished work, leaving the view blank until it finished.
        """
        import threading
        first = window._analyzer_lock
        assert isinstance(first, threading.Lock().__class__)
        # Simulate what load_sheet's on_loaded does when a new sheet arrives
        window._data_seq += 1
        window._analyzer_lock = threading.Lock()
        assert window._analyzer_lock is not first


class TestNoHandBuiltMarkup:
    """
    Phase 11 converted four views from hand-built HTML/ASCII to real widgets.
    These guard against a regression back to string-built markup, which is what
    made those views ignore the theme and the design tokens.
    """

    def test_views_expose_real_models_not_text(self, qapp, window):
        for row, name in _view_rows(window):
            if name in ("Monthly Inflow", "Data Profile", "Logic Checks"):
                window.nav_list.setCurrentRow(row)
                _settle(qapp, window)

        # Monthly Inflow: table replaced the fixed-width ASCII QLabel
        assert window.monthly_table.model().rowCount() > 0
        assert window.monthly_table.model().columnCount() == 3

        # Data Profile: table replaced the <table border='1'> HTML
        prof = window.profile_table.model()
        assert prof.rowCount() == len(window.df.columns)
        assert [prof.horizontalHeaderItem(c).text() for c in range(prof.columnCount())] == \
            ["Column", "Type", "Whitespace", "Sample"]

        # Logic Checks: the rules are a table, not a stack of prose bands. This
        # replaced first a 78-line HTML builder and then five full-width tinted
        # cards, so the guard is that the rules and their evidence are both
        # model-backed and nothing here is assembled as a string.
        rules = window.lc_findings_table.model()
        assert rules is not None and rules.rowCount() > 0
        assert [rules.headerData(c, Qt.Orientation.Horizontal,
                                 Qt.ItemDataRole.DisplayRole)
                for c in range(rules.columnCount())] == ui.SanityCheckApp.LOGIC_COLUMNS

        # Selecting a rule fills the inspector, whose evidence sample is also a
        # real model rather than a formatted block of text.
        assert window.lc_ins_title.text()
        assert window.lc_evidence_table.model() is not None

    def test_source_has_no_html_or_ascii_table_building(self):
        src = (ui.__file__).replace(".pyc", ".py")
        with open(src, encoding="utf-8") as fh:
            body = fh.read()

        # setHtml on a results pane, and the ASCII/HTML table idioms
        for banned in ("<table border=", "html.append(", "'-' * 36", '"-" * 36'):
            assert banned not in body, f"hand-built markup is back: {banned}"

    def test_language_check_does_not_parse_emoji_prefixes(self):
        src = (ui.__file__).replace(".pyc", ".py")
        with open(src, encoding="utf-8") as fh:
            body = fh.read()
        # Colour used to be chosen with line.startswith("âœ…") on format_result output
        assert 'startswith("âœ…")' not in body
        assert "format_result" not in body, \
            "Language Check must render from check_column()'s dict, not formatted text"


class TestLanguageModeIsReportedAccurately:
    """
    The Language Check summary claimed "Detection: Lingua + rules" whatever the
    Mode combo said, because it read is_lingua_available() - installed, not
    used. Selecting "Rules only" still credited the ML model.
    """

    def _run(self, qapp, window, mode_id):
        window.show_view("Language Check")
        _settle(qapp, window)
        combo = window.lang_mode_combo
        combo.setCurrentIndex(combo.findData(mode_id))
        window.lang_col_combo.setCurrentText("short_description")
        window.run_language_check()
        _settle(qapp, window)
        return window.lang_summary_label.text()

    def test_mode_ids_are_data_not_display_text(self, qapp, window):
        """The combo shows readable labels; the mode id lives in item data."""
        window.show_view("Language Check")   # views are built lazily
        _settle(qapp, window)
        combo = window.lang_mode_combo
        ids = {combo.itemData(i) for i in range(combo.count())}
        assert ids == {"all", "lingua_only", "rules_only"}
        for i in range(combo.count()):
            assert combo.itemData(i) != combo.itemText(i), \
                "raw mode ids must not be the visible label"

    def test_rules_only_summary_does_not_credit_lingua(self, qapp, window):
        text = self._run(qapp, window, "rules_only")
        assert "no Lingua" in text, f"summary must say Lingua was skipped: {text!r}"

    def test_lingua_only_summary_does_not_credit_rules(self, qapp, window):
        text = self._run(qapp, window, "lingua_only")
        assert "Lingua only" in text, f"summary must name Lingua-only: {text!r}"

    def test_detail_table_is_shown_for_a_clean_result(self, qapp, window):
        """A clean result still has to disclose how narrow the run was."""
        self._run(qapp, window, "rules_only")
        model = window.lang_detail_table.model()
        # isHidden(), not isVisible(): the top-level window is never shown in a
        # headless run, so every descendant reports isVisible() False regardless
        # of its own setVisible call.
        assert not window.lang_detail_table.isHidden()
        assert model is not None and model.rowCount() > 0
        labels = [model.item(r, 0).text() for r in range(model.rowCount())]
        assert "Lingua ML model" in labels
        assert any("skipped" in l.lower() for l in labels)

    def test_detail_table_marks_skipped_tiers(self, qapp, window):
        self._run(qapp, window, "lingua_only")
        model = window.lang_detail_table.model()
        values = {model.item(r, 0).text(): model.item(r, 1).text()
                  for r in range(model.rowCount())}
        assert values["Keyword pattern matching"] == "Skipped by mode"
        assert values["Lingua ML model"] == "Used"


class TestColumnMappingOverride:
    """
    A wrong automatic match makes every check measure the wrong field. The view
    has to show why a column was chosen, and the mapping has to be correctable
    and to survive a sheet reload.
    """

    def _ambiguous(self):
        n = 4
        return pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": ["New"] * n,
            "urgency": ["P1", "P2", None, None],
            "impact": ["1", None, None, None],
            "assignment_group": ["Team"] * n,
            "cmdb_ci": ["CI"] * n,
            "opened_at": pd.to_datetime(["2024-01-01"] * n),
            "resolved_at": pd.to_datetime(["2024-02-01"] * n),
            "short_description": ["Login issue for the user account"] * n,
            "description": ["A sufficiently detailed description"] * n,
        })

    def _load(self, qapp, window, frame):
        window.df = frame
        window.analyzer = SanityAnalyzer(frame)
        window.show_view("Column Check")
        _settle(qapp, window)

    def test_the_view_shows_scores_and_alternatives(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        model = window.col_check_table.model()
        headers = [model.horizontalHeaderItem(c).text()
                   for c in range(model.columnCount())]
        assert headers == ["Required Column", "Status", "Matched Column",
                           "Score", "Other candidates"]

    def test_a_tie_is_shown_with_both_scores(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        model = window.col_check_table.model()
        rows = {model.item(r, 0).text(): model.item(r, 4).text()
                for r in range(model.rowCount())}
        assert "impact" in rows["Priority"]
        assert "(" in rows["Priority"], "the runner-up's score must be visible"

    def test_applying_an_override_redirects_the_checks(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        assert window.analyzer.priority_column == "urgency"
        window.apply_column_overrides({"Priority": "impact"})
        _settle(qapp, window)
        assert window.analyzer.priority_column == "impact"

    def test_the_override_is_persisted(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        window.apply_column_overrides({"Priority": "impact"})
        assert window.config["column_overrides"] == {"Priority": "impact"}
        # and written, not just held in memory
        reloaded = ConfigManager(window.config_mgr.config_file).config
        assert reloaded["column_overrides"] == {"Priority": "impact"}

    def test_the_view_marks_a_hand_set_mapping(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        window.apply_column_overrides({"Priority": "impact"})
        _settle(qapp, window)
        model = window.col_check_table.model()
        statuses = {model.item(r, 0).text(): model.item(r, 1).text()
                    for r in range(model.rowCount())}
        assert "Set by you" in statuses["Priority"]

    def test_applying_invalidates_in_flight_work(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        before = window._data_seq
        window.apply_column_overrides({"Priority": "impact"})
        assert window._data_seq > before

    def test_applying_clears_stale_language_findings(self, qapp, window):
        """A remapping can change which column a finding was about."""
        self._load(qapp, window, self._ambiguous())
        window.language_results["description"] = {"column": "description"}
        window.apply_column_overrides({"Priority": "impact"})
        assert window.language_results == {}

    def test_re_applying_the_same_mapping_is_a_no_op(self, qapp, window):
        self._load(qapp, window, self._ambiguous())
        window.apply_column_overrides({"Priority": "impact"})
        _settle(qapp, window)
        analyzer = window.analyzer
        window.apply_column_overrides({"Priority": "impact"})
        assert window.analyzer is analyzer

    def test_an_ignored_override_is_surfaced(self, qapp, window):
        """A setting from another export shape must not fail silently."""
        frame = self._ambiguous()
        window.df = frame
        window.analyzer = SanityAnalyzer(
            frame, column_overrides={"Priority": "not_a_column"})
        window.show_view("Column Check")
        _settle(qapp, window)
        assert "ignored" in window.col_override_label.text()
        assert window.col_override_label.property("state") == "warn"

    def test_a_saved_override_survives_a_sheet_load(self, qapp, window,
                                                    tmp_path, monkeypatch):
        """
        Loading a sheet rebuilt the analyzer from scratch, which would revert
        the correction the user just made.
        """
        frame = self._ambiguous()
        window.config["column_overrides"] = {"Priority": "impact"}
        csv = tmp_path / "t.csv"
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        assert window.analyzer.priority_column == "impact"

    def test_a_saved_date_order_survives_a_sheet_load(self, qapp, window,
                                                      tmp_path):
        frame = self._ambiguous()
        window.config["date_dayfirst"] = True
        csv = tmp_path / "t.csv"
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        assert window.analyzer.dayfirst is True


class TestDateFormatChoice:
    """
    An ambiguous date column is the one finding that changes every other number
    in the app, so it gets a control rather than only a warning.
    """

    def _ambiguous_df(self):
        rows = [("05/03/2024", "08/03/2024"), ("02/04/2024", "09/04/2024"),
                ("10/05/2024", "12/05/2024")]
        n = len(rows)
        return pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": ["Closed"] * n,
            "priority": ["P1"] * n,
            "assignment_group": ["Team"] * n,
            "cmdb_ci": ["CI"] * n,
            "opened_at": [o for o, _ in rows],
            "resolved_at": [r for _, r in rows],
            "short_description": ["Login issue for the user account"] * n,
            "description": ["A sufficiently detailed description"] * n,
        })

    def _load(self, qapp, window, frame):
        window.df = frame
        window.analyzer = SanityAnalyzer(frame)
        window.show_view("Logic Checks")
        _settle(qapp, window)

    def test_the_card_is_hidden_when_dates_are_unambiguous(self, qapp, window):
        """The fixture's dates are real datetimes, so there is nothing to ask."""
        window.show_view("Logic Checks")
        _settle(qapp, window)
        assert window.lc_format_card.isHidden()

    def test_an_ambiguous_column_raises_the_card(self, qapp, window):
        self._load(qapp, window, self._ambiguous_df())
        assert not window.lc_format_card.isHidden()
        text = window.lc_format_label.text()
        assert "either way" in text

    def test_the_message_prices_the_other_reading(self, qapp, window):
        self._load(qapp, window, self._ambiguous_df())
        text = window.lc_format_label.text()
        assert "changes 3 rows" in text or "3 rows" in text
        assert "resolution time" in text

    def test_applying_day_first_rebuilds_the_analyzer(self, qapp, window):
        self._load(qapp, window, self._ambiguous_df())
        combo = window.lc_format_combo
        combo.setCurrentIndex(combo.findData(True))
        window.apply_date_format()
        _settle(qapp, window)
        assert window.analyzer.dayfirst is True
        assert window.analyzer._get_created_dates()[0] == pd.Timestamp("2024-03-05")

    def test_applying_fixes_the_derived_numbers(self, qapp, window):
        self._load(qapp, window, self._ambiguous_df())
        before = (window.analyzer._get_closed_dates()
                  - window.analyzer._get_created_dates()).dt.days
        assert list(before) == [92, 213, 61]

        combo = window.lc_format_combo
        combo.setCurrentIndex(combo.findData(True))
        window.apply_date_format()
        _settle(qapp, window)
        after = (window.analyzer._get_closed_dates()
                 - window.analyzer._get_created_dates()).dt.days
        assert list(after) == [3, 7, 2]

    def test_applying_invalidates_in_flight_work(self, qapp, window):
        """Anything still computing against the old parse must be discarded."""
        self._load(qapp, window, self._ambiguous_df())
        before = window._data_seq
        combo = window.lc_format_combo
        combo.setCurrentIndex(combo.findData(True))
        window.apply_date_format()
        assert window._data_seq > before

    def test_applying_clears_stale_language_findings(self, qapp, window):
        self._load(qapp, window, self._ambiguous_df())
        window.language_results["description"] = {"column": "description"}
        combo = window.lc_format_combo
        combo.setCurrentIndex(combo.findData(True))
        window.apply_date_format()
        assert window.language_results == {}

    def test_re_applying_the_same_choice_is_a_no_op(self, qapp, window):
        self._load(qapp, window, self._ambiguous_df())
        analyzer = window.analyzer
        window.lc_format_combo.setCurrentIndex(
            window.lc_format_combo.findData(None))
        window.apply_date_format()
        assert window.analyzer is analyzer, "no needless rebuild"

    def test_the_combo_reflects_the_analyzer(self, qapp, window):
        frame = self._ambiguous_df()
        window.df = frame
        window.analyzer = SanityAnalyzer(frame, dayfirst=True)
        window.show_view("Logic Checks")
        _settle(qapp, window)
        assert window.lc_format_combo.currentData() is True

    def test_a_conflicting_column_is_an_error_not_a_choice(self, qapp, window):
        frame = self._ambiguous_df()
        frame["opened_at"] = ["13/01/2024", "01/13/2024", "05/03/2024"]
        self._load(qapp, window, frame)
        assert not window.lc_format_card.isHidden()
        assert window.lc_format_card.property("severity") == "error"
        assert "No single reading is correct" in window.lc_format_label.text()


class TestExportReport:
    """
    The export is the app's only long-running action, and it used to run with no
    progress, no button gating, a bare "Done!" dialog, and a closure over
    self.df that let a mid-export sheet switch redirect the remaining sheets.
    """

    def _prepare(self, window, tmp_path, monkeypatch, filename="r.xlsx"):
        out = tmp_path / filename
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k: shown.append((title, text))))
        return out, shown

    def test_the_export_writes_the_workbook(self, qapp, window, tmp_path,
                                            monkeypatch):
        out, shown = self._prepare(window, tmp_path, monkeypatch)
        window.export_report()
        _settle(qapp, window)
        assert out.exists()
        assert shown, "the user must be told it finished"

    def test_the_dialog_names_the_path_not_just_done(self, qapp, window, tmp_path,
                                                    monkeypatch):
        out, shown = self._prepare(window, tmp_path, monkeypatch)
        window.export_report()
        _settle(qapp, window)
        _title, text = shown[-1]
        assert str(out) in text

    def test_the_dialog_lists_what_was_not_checked(self, qapp, window, tmp_path,
                                                  monkeypatch):
        """A skipped check the user never sees is the failure this guards."""
        out, shown = self._prepare(window, tmp_path, monkeypatch)
        window.export_report()
        _settle(qapp, window)
        _title, text = shown[-1]
        # no language check has been run on the fixture
        assert "Notes:" in text
        assert "Language" in text

    def test_export_buttons_are_gated_during_the_run(self, qapp, window, tmp_path,
                                                    monkeypatch):
        """Both were live, so two exports could target the same path at once."""
        self._prepare(window, tmp_path, monkeypatch)
        window._begin_export()
        assert not window.btn_export.isEnabled()
        assert not window.export_action.isEnabled()
        window._finish_export()
        assert window.btn_export.isEnabled()
        assert window.export_action.isEnabled()

    def test_progress_messages_reach_the_status_bar(self, qapp, window, tmp_path,
                                                   monkeypatch):
        """
        WorkerSignals.progress existed and was connected, but nothing could emit
        it because the task was never handed the emitter.
        """
        self._prepare(window, tmp_path, monkeypatch)
        seen = []
        monkeypatch.setattr(window, "update_status", seen.append)
        window.export_report()
        _settle(qapp, window)
        assert seen, "the export must report progress"
        assert any("Building" in message for message in seen)

    def test_the_task_does_not_read_self_df(self, qapp, window, tmp_path,
                                            monkeypatch):
        """
        A sheet switch mid-export must not redirect the remaining sheets onto
        the new sheet's data. The frame is snapshotted at dispatch, so clearing
        self.df afterwards cannot affect the file being written.
        """
        out, _shown = self._prepare(window, tmp_path, monkeypatch)
        original = window.df
        window.export_report()
        window.df = None                 # simulate the switch landing mid-flight
        _settle(qapp, window)
        window.df = original
        assert out.exists(), "the export must survive self.df changing"

    def test_a_stale_sheet_is_disclosed_not_hidden(self, qapp, window, tmp_path,
                                                  monkeypatch):
        out, _shown = self._prepare(window, tmp_path, monkeypatch)
        seen = []
        monkeypatch.setattr(window, "update_status", seen.append)
        window.export_report()
        window._data_seq += 1            # a different sheet is now loaded
        _settle(qapp, window)
        assert any("previous sheet" in message for message in seen)

    def test_failure_re_enables_the_buttons(self, qapp, window, tmp_path,
                                            monkeypatch):
        """A failed export must not leave the app unable to export again."""
        self._prepare(window, tmp_path, monkeypatch)
        errors = []
        monkeypatch.setattr(window, "show_error", errors.append)
        monkeypatch.setattr(
            ui.ReportGenerator, "export_to_excel",
            lambda self, path, progress=None: (_ for _ in ()).throw(
                RuntimeError("disk full")))
        window.export_report()
        _settle(qapp, window)
        assert errors and "disk full" in errors[0]
        assert window.btn_export.isEnabled()
        assert not window.progress_bar.isVisible()

    def test_an_oversized_frame_asks_before_exporting(self, qapp, window,
                                                     tmp_path, monkeypatch):
        asked = []
        monkeypatch.setattr(
            ui.QMessageBox, "warning",
            staticmethod(lambda *a, **k: asked.append(a[2])
                         or ui.QMessageBox.StandardButton.Cancel))
        monkeypatch.setattr(ui, "DETAIL_MAX_DATA_ROWS", 2)
        window.export_report()
        assert asked, "the row ceiling must be raised before the save dialog"
        assert "worksheet holds" in asked[0]

    def test_declining_the_warning_cancels_the_export(self, qapp, window,
                                                     tmp_path, monkeypatch):
        out, _shown = self._prepare(window, tmp_path, monkeypatch)
        monkeypatch.setattr(
            ui.QMessageBox, "warning",
            staticmethod(lambda *a, **k: ui.QMessageBox.StandardButton.Cancel))
        monkeypatch.setattr(ui, "DETAIL_MAX_DATA_ROWS", 2)
        window.export_report()
        _settle(qapp, window)
        assert not out.exists()


class TestTablesAreCopyable:
    """
    Results are read-only, which is not the same as unextractable - they exist
    to be pasted into a ticket or a report. Copy is installed by
    configure_table so no table can be left out by omission.
    """

    def _run_language_check(self, qapp, window):
        """The shared fixture is all-English, so the breakdown would be empty."""
        df = window.df.copy()
        df["description"] = [
            "Le serveur ne repond plus depuis ce matin merci",
            "Der Drucker funktioniert nicht mehr richtig heute",
            "The printer is offline and refusing all print jobs",
            "Impresora no funciona correctamente hoy dia gracias",
            "Привет мир это русский текст здесь для теста",
            "Network connectivity dropped on the third floor switch",
            "Bonjour je ne peux pas acceder a mon compte aidez moi",
            "Das System ist sehr langsam geworden seit gestern",
        ]
        window.df = df
        window.analyzer = SanityAnalyzer(df)

        window.show_view("Language Check")
        _settle(qapp, window)
        combo = window.lang_mode_combo
        combo.setCurrentIndex(combo.findData("all"))
        window.lang_col_combo.setCurrentText("description")
        window.run_language_check()
        _settle(qapp, window)
        assert window.lang_breakdown_table.model() is not None, \
            "fixture must produce detections for a copy test to mean anything"

    def test_every_table_has_a_copy_action(self, qapp, window):
        for row, _ in _view_rows(window):
            window.nav_list.setCurrentRow(row)
            _settle(qapp, window)

        without = []
        for attr in dir(window):
            if not attr.endswith("_table"):
                continue
            view = getattr(window, attr, None)
            if not isinstance(view, QTableView):
                continue
            if "Copy" not in [a.text() for a in view.actions()]:
                without.append(attr)
        assert not without, f"tables with no copy action: {without}"

    def test_copy_is_tab_separated_for_excel(self, qapp, window):
        window.show_view("Column Check")
        _settle(qapp, window)
        ui._copy_table(window.col_check_table, selected_only=False,
                       with_headers=True, with_context=False)
        text = qapp.clipboard().text()
        assert "\t" in text, "must be TSV so it pastes into Excel as columns"
        first = text.splitlines()[0].split("\t")
        assert len(first) == window.col_check_table.model().columnCount()

    def test_report_copy_names_the_dataset_and_sheet(self, qapp, window):
        # setVisible matters: the sheet name is omitted from the provenance
        # block when the picker is hidden, because a hidden picker means the
        # file has no sheets. load_file_async shows it for a workbook, so this
        # has to as well.
        window.combo_sheets.blockSignals(True)
        window.combo_sheets.addItems(["Incidents"])
        window.combo_sheets.setCurrentText("Incidents")
        window.combo_sheets.setEnabled(True)
        window.combo_sheets.setVisible(True)
        window.combo_sheets.blockSignals(False)

        window.show_view("Column Check")
        _settle(qapp, window)
        ui._copy_table(window.col_check_table, selected_only=False,
                       with_headers=True, with_context=True)
        text = qapp.clipboard().text()
        assert "tickets.xlsx" in text, "a pasted table must say which file it came from"
        assert "Sheet: Incidents" in text
        assert "Generated:" in text

    def test_report_copy_carries_the_percentage_denominator(self, qapp, window):
        """A '% of Analyzed' column is unreadable without the row counts."""
        self._run_language_check(qapp, window)
        ui._copy_table(window.lang_breakdown_table, selected_only=False,
                       with_headers=True, with_context=True)
        text = qapp.clipboard().text()
        assert "Rows analyzed:" in text
        assert "Detection method:" in text
        assert "Column checked:" in text

    def test_language_breakdown_reports_percentages(self, qapp, window):
        self._run_language_check(qapp, window)
        headers = ui._table_headers(window.lang_breakdown_table)
        assert headers == ["Language", "Rows", "% of Analyzed"], \
            "counts alone cannot be compared between columns or files"

    def test_copy_follows_the_displayed_sort_order(self, qapp, window):
        """Copy reads DisplayRole, so what was on screen is what is pasted."""
        window.show_view("Null Analysis")
        _settle(qapp, window)
        table = window.null_table
        table.sortByColumn(0, Qt.SortOrder.AscendingOrder)
        _settle(qapp, window)
        model = table.model()
        on_screen = [str(model.data(model.index(r, 0))) for r in range(model.rowCount())]

        ui._copy_table(table, selected_only=False, with_headers=False, with_context=False)
        pasted = [line.split("\t")[0] for line in qapp.clipboard().text().splitlines()]
        assert pasted == on_screen

    def test_selection_copy_excludes_headers_and_context(self, qapp, window):
        window.show_view("Column Check")
        _settle(qapp, window)
        table = window.col_check_table
        table.selectRow(0)
        ui._copy_table(table, selected_only=True, with_headers=False, with_context=False)
        text = qapp.clipboard().text()
        assert len(text.splitlines()) == 1, "one selected row copies as one line"
        assert "DataProbe" not in text

    def test_empty_table_copies_nothing(self, qapp, window):
        """Guards against clobbering the clipboard with a blank string."""
        window.show_view("Language Check")
        _settle(qapp, window)
        qapp.clipboard().setText("untouched")
        copied = ui._copy_table(window.lang_samples_table, selected_only=False,
                                with_headers=True, with_context=True)
        assert copied == 0
        assert qapp.clipboard().text() == "untouched"


class TestViewsDoNotOverlap:
    """
    A QVBoxLayout given less room than its size hints need takes the space from
    whatever can shrink - so a table pinned with setFixedHeight kept its height
    while the section headers below it were laid out *inside* it, and the views
    drew on top of each other. Reported on a maximized window in Language
    Check, which stacks three pinned tables.

    The fix is a scroll area, and the guard is structural: any view stacking
    pinned content must scroll, or it will overlap at some window size.
    """

    def _overlaps(self, root):
        """Overlapping visible siblings sharing a parent that has a layout."""
        found = []
        stack = [root]
        while stack:
            node = stack.pop()
            kids = [c for c in node.children()
                    if isinstance(c, ui.QWidget) and c.isVisible()
                    and c.width() > 1 and c.height() > 1]
            if node.layout() is not None:
                for i in range(len(kids)):
                    for j in range(i + 1, len(kids)):
                        a, b = kids[i].geometry(), kids[j].geometry()
                        overlap = a.intersected(b)
                        if overlap.height() > 2 and overlap.width() > 2:
                            found.append(
                                f"{type(kids[i]).__name__} {a.top()}..{a.bottom()}"
                                f" X {type(kids[j]).__name__} "
                                f"{b.top()}..{b.bottom()}")
            stack.extend(kids)
        return found

    def _run_language_check(self, qapp, window, size):
        """
        The view that broke: three pinned tables, all populated.

        show() before resize() is load-bearing. An unshown window ignores
        resize, leaves the page at its default height, and the deficit that
        causes the overlap never happens - which is how the first version of
        this test passed against the unfixed code.
        """
        window.show()
        window.resize(*size)
        _settle(qapp, window)

        frame = window.df.copy()
        frame["description"] = [
            "Le serveur ne repond plus depuis ce matin merci",
            "Der Drucker funktioniert nicht mehr richtig heute",
            "Il portale non risponde piu da stamattina adesso",
            "Impresora no funciona correctamente hoy dia gracias",
            "Skriveren fungerer ikke i dag dessverre takk",
            "The printer is offline and refusing all print jobs",
            "Network connectivity dropped on the third floor switch",
            "User cannot log in to the finance application at all",
        ]
        window.df = frame
        window.analyzer = SanityAnalyzer(frame)
        window.show_view("Language Check")
        _settle(qapp, window)
        window.lang_col_combo.setCurrentText("description")
        window.run_language_check()
        _settle(qapp, window)

    # 1280x720 is a 14-inch 1080p laptop at Windows' default 150% scaling,
    # maximised - the machine this tool is actually used on, and the one size
    # missing here. It is *shorter* than the 1000x700 case while being wider,
    # which is the shape that hides a pinned table's last rows.
    @pytest.mark.parametrize("size", [(1000, 700), (1280, 720), (1280, 900),
                                      (1920, 1005)])
    def test_no_view_overlaps_itself(self, qapp, window, size):
        self._run_language_check(qapp, window, size)

        broken = {}
        for row, name in _view_rows(window):
            window.nav_list.setCurrentRow(row)
            _settle(qapp, window)
            found = self._overlaps(window.tab_widgets[name])
            if found:
                broken[name] = found
        assert not broken, f"overlapping widgets at {size}: {broken}"

    def test_every_view_is_reachable_without_scrolling_on_a_14_inch_laptop(
            self, qapp, window):
        """
        All views fit the nav column at 1280x720.

        This is a real regression that had shipped: at 32px rows and 34px group
        captions the column needed 584px and had 465px, so some views sat below
        the fold with nothing on screen saying so - a scrollbar in a sidebar is
        not somewhere people look. Six captions were spending 192px of it.

        Pinned as a *measurement* rather than a pixel budget, because the thing
        that matters is the relationship: a fifteenth view, a taller row or a
        roomier group caption each push something off again, and all three are
        caught here. A larger nav *font* is not - the row height is a fixed
        sizeHint that does not track the font, so bigger text clips inside the
        row instead of growing it. That is worth knowing rather than assuming
        this test covers it. A taller window is not asserted either: 1100x700
        legitimately still scrolls.
        """
        window.show()
        window.resize(1280, 720)
        _settle(qapp, window)

        nav = window.nav_list
        needed = sum(nav.sizeHintForRow(r) for r in range(nav.count()))
        available = nav.viewport().height()
        assert needed <= available, (
            f"the nav column needs {needed}px and has {available}px, so "
            f"{sum(1 for _ in _view_rows(window))} views do not all fit - "
            f"something is below the fold on a 14-inch laptop")

    def test_language_check_scrolls_rather_than_overlapping(self, qapp, window):
        """Its tables are pinned, so the deficit has to go somewhere."""
        self._run_language_check(qapp, window, (1000, 700))
        page = window.tab_widgets["Language Check"]
        areas = page.findChildren(ui.QScrollArea)
        assert areas, "a view stacking pinned tables must be scrollable"
        area = areas[0]
        assert area.widgetResizable()
        assert area.widget().height() >= area.viewport().height()

    def test_pinned_tables_show_every_row_they_claim(self, qapp, window):
        """
        fit_table_height measures real section sizes; arithmetic on
        TABLE_ROW_HEIGHT under-measures when the font forces taller rows, and
        the last rows then sit behind the bottom edge.
        """
        self._run_language_check(qapp, window, (1280, 900))
        for table, cap in ((window.lang_detail_table, None),
                           (window.lang_breakdown_table, 6)):
            model = table.model()
            assert model is not None
            shown = model.rowCount() if cap is None \
                else min(model.rowCount(), cap)
            header = table.verticalHeader()
            needed = sum(header.sectionSize(r) for r in range(shown))
            assert table.viewport().height() >= needed, \
                f"{shown} rows need {needed}px, viewport is " \
                f"{table.viewport().height()}px"

    def test_fit_table_height_survives_an_empty_model(self, qapp, window):
        """A zero-row model must not collapse the table to nothing."""
        table = QTableView()
        ui.configure_table(table)
        ui.set_table_model(table, ui.QStandardItemModel(0, 2))
        ui.fit_table_height(table)
        assert table.height() > 0


class TestFindingsTextIsSelectable:
    """
    The tables were copyable but the summary cards were not, so the one sentence
    a user actually wants in a ticket - "672 of 6,913 analyzed entries (9.7%)
    contain non-English text" - had to be retyped. Same rule as the tables:
    read-only means non-editable, not non-extractable.
    """

    def _labels(self, page):
        return page.findChildren(ui.QLabel)

    def test_every_label_in_every_view_is_selectable(self, qapp, window):
        missed = {}
        for row, name in _view_rows(window):
            window.nav_list.setCurrentRow(row)
            _settle(qapp, window)
            page = window.tab_widgets[name]
            bad = [label.text()[:40] for label in self._labels(page)
                   if not (label.textInteractionFlags()
                           & Qt.TextInteractionFlag.TextSelectableByMouse)]
            if bad:
                missed[name] = bad
        assert not missed, f"unselectable text: {missed}"

    def test_labels_carry_copy_actions(self, qapp, window):
        window.show_view("Logic Checks")
        _settle(qapp, window)
        page = window.tab_widgets["Logic Checks"]
        for label in self._labels(page):
            assert [a.text() for a in label.actions()] == \
                ["Copy", "Copy for report"]

    def test_labels_stay_out_of_the_tab_order(self, qapp, window):
        """
        The keyboard flag would raise focusPolicy to StrongFocus and put ~70
        labels between the user and the next control.
        """
        window.show_view("Logic Checks")
        _settle(qapp, window)
        for label in self._labels(window.tab_widgets["Logic Checks"]):
            assert not (label.textInteractionFlags()
                        & Qt.TextInteractionFlag.TextSelectableByKeyboard)
            assert label.focusPolicy() != Qt.FocusPolicy.StrongFocus

    def test_copy_with_no_selection_takes_the_whole_finding(self, qapp, window):
        window.show_view("Logic Checks")
        _settle(qapp, window)
        label = next(l for l in self._labels(window.tab_widgets["Logic Checks"])
                     if l.text())
        qapp.clipboard().clear()
        ui._copy_label(label, with_context=False)
        assert qapp.clipboard().text() == label.text()

    def test_copy_prefers_the_selection(self, qapp, window):
        window.show_view("Logic Checks")
        _settle(qapp, window)
        label = next(l for l in self._labels(window.tab_widgets["Logic Checks"])
                     if len(l.text()) > 12)
        label.setSelection(0, 8)
        qapp.clipboard().clear()
        ui._copy_label(label, with_context=False)
        assert qapp.clipboard().text() == label.text()[:8]

    def test_a_multi_line_selection_keeps_its_line_breaks(self, qapp, window):
        """Qt hands back U+2029 for a break inside selected label text."""
        label = ui.QLabel("first line\nsecond line")
        ui.make_text_selectable(label)
        label.setSelection(0, len(label.text()))
        assert " " in label.selectedText()
        qapp.clipboard().clear()
        ui._copy_label(label, with_context=False)
        assert qapp.clipboard().text() == "first line\nsecond line"

    def test_markup_is_copied_as_plain_text(self, qapp, window):
        label = ui.QLabel("<b>Bold</b> and <i>italic</i>")
        label.setTextFormat(Qt.TextFormat.RichText)
        ui.make_text_selectable(label)
        qapp.clipboard().clear()
        ui._copy_label(label, with_context=False)
        assert qapp.clipboard().text() == "Bold and italic"

    def test_an_empty_label_does_not_clobber_the_clipboard(self, qapp, window):
        label = ui.QLabel("   ")
        ui.make_text_selectable(label)
        qapp.clipboard().setText("untouched")
        ui._copy_label(label, with_context=False)
        assert qapp.clipboard().text() == "untouched"

    def test_report_copy_carries_the_provenance(self, qapp, window):
        window.show_view("Logic Checks")
        _settle(qapp, window)
        label = next(l for l in self._labels(window.tab_widgets["Logic Checks"])
                     if l.text())
        qapp.clipboard().clear()
        ui._copy_label(label, with_context=True)
        text = qapp.clipboard().text()
        assert "tickets.xlsx" in text
        assert "Generated:" in text
        assert label.text() in text

    def test_the_language_summary_names_its_column_and_mode(self, qapp, window):
        """A bare "9.7%" in a ticket cannot be traced to a column or a mode."""
        window.show_view("Language Check")
        _settle(qapp, window)
        combo = window.lang_mode_combo
        combo.setCurrentIndex(combo.findData("rules_only"))
        window.lang_col_combo.setCurrentText("description")
        window.run_language_check()
        _settle(qapp, window)

        qapp.clipboard().clear()
        ui._copy_label(window.lang_summary_label, with_context=True)
        text = qapp.clipboard().text()
        assert "Column checked: description" in text
        assert "Detection method:" in text
        assert "Rows analyzed:" in text

    def test_installation_is_idempotent(self, qapp, window):
        """show_view runs on every visit; actions must not accumulate."""
        for _ in range(3):
            window.show_view("Null Analysis")
            _settle(qapp, window)
        for label in self._labels(window.tab_widgets["Null Analysis"]):
            assert len(label.actions()) == 2

    def test_help_dialog_text_is_selectable(self, qapp, window, monkeypatch):
        captured = {}
        monkeypatch.setattr(
            ui.QDialog, "exec",
            lambda self: captured.setdefault("labels", self.findChildren(ui.QLabel)))
        window.show_user_guide_dialog()
        assert captured["labels"], "the guide must build a label to check"
        assert all(
            label.textInteractionFlags()
            & Qt.TextInteractionFlag.TextSelectableByMouse
            for label in captured["labels"])

    def test_help_dialog_links_still_work(self, qapp, window, monkeypatch):
        """Selection must not cost the link handling the guide relies on."""
        captured = {}
        monkeypatch.setattr(
            ui.QDialog, "exec",
            lambda self: captured.setdefault(
                "linked", [l for l in self.findChildren(ui.QLabel)
                           if l.openExternalLinks()]))
        window.show_user_guide_dialog()
        for label in captured["linked"]:
            assert (label.textInteractionFlags()
                    & Qt.TextInteractionFlag.LinksAccessibleByMouse)


class TestPandasModel:
    def test_sort_reorders(self):
        model = ui.PandasModel(pd.DataFrame({"v": [3, 1, 2]}))
        model.sort(0, Qt.SortOrder.AscendingOrder)
        assert [model.data(model.index(r, 0)) for r in range(3)] == ["1", "2", "3"]
        model.sort(0, Qt.SortOrder.DescendingOrder)
        assert [model.data(model.index(r, 0)) for r in range(3)] == ["3", "2", "1"]

    def test_nan_renders_blank_not_the_string_nan(self):
        model = ui.PandasModel(pd.DataFrame({"v": [1.0, None]}))
        assert model.data(model.index(1, 0)) == ""

    def test_mixed_types_sort_without_raising(self):
        model = ui.PandasModel(pd.DataFrame({"v": [1, "b", None, 2.5]}))
        model.sort(0, Qt.SortOrder.AscendingOrder)   # must not raise

    def test_is_read_only(self):
        model = ui.PandasModel(pd.DataFrame({"v": [1]}))
        assert not (model.flags(model.index(0, 0)) & Qt.ItemFlag.ItemIsEditable)


class TestWindowStatePersistence:
    def test_geometry_round_trips(self, qapp, tmp_path, monkeypatch, ticket_df):
        cfg = str(tmp_path / "geom.json")
        monkeypatch.setattr(ui, "ConfigManager", lambda *a, **k: ConfigManager(cfg))

        first = ui.SanityCheckApp()
        first.resize(1180, 742)
        qapp.processEvents()
        first.close()
        qapp.processEvents()

        second = ui.SanityCheckApp()
        assert (second.width(), second.height()) == (1180, 742)

    def test_offscreen_geometry_is_ignored(self, qapp, tmp_path, monkeypatch):
        cfg = str(tmp_path / "offscreen.json")
        mgr = ConfigManager(cfg)
        mgr.config["window_geometry"] = {"x": -30000, "y": -30000, "w": 800, "h": 600}
        mgr.save()
        monkeypatch.setattr(ui, "ConfigManager", lambda *a, **k: ConfigManager(cfg))

        win = ui.SanityCheckApp()
        assert win.x() > -10000 and win.y() > -10000, \
            "a saved position on a detached monitor would be unreachable"

    def _window_with_last_view(self, tmp_path, monkeypatch, saved):
        cfg = str(tmp_path / "lastview.json")
        mgr = ConfigManager(cfg)
        mgr.config["last_view"] = saved
        mgr.save()
        monkeypatch.setattr(ui, "ConfigManager", lambda *a, **k: ConfigManager(cfg))
        return ui.SanityCheckApp()

    def test_the_last_view_is_restored(self, qapp, tmp_path, monkeypatch):
        """
        closeEvent has always written last_view; nothing read it, so the key was
        dead and _restore_window_state's docstring claimed a restore it did not
        perform. Restoring it selects the nav row, which is what makes the
        sidebar highlight and the page header agree.
        """
        win = self._window_with_last_view(tmp_path, monkeypatch, "Data Profile")
        assert win.current_view == "Data Profile"
        row = win.nav_list.currentRow()
        assert win.nav_list.item(row).data(Qt.ItemDataRole.UserRole) == "Data Profile"

    def test_the_first_load_lands_on_the_restored_view(self, qapp, tmp_path,
                                                       monkeypatch, ticket_df):
        """
        load_sheet only jumps to the first row while current_view is None. That
        is the mechanism the restore relies on, so it is worth pinning: without
        it the session silently reopens on Overview every time.
        """
        win = self._window_with_last_view(tmp_path, monkeypatch, "Data Profile")
        win.df = ticket_df
        win.analyzer = SanityAnalyzer(ticket_df)
        win.filepath = str(tmp_path / "tickets.xlsx")
        win._set_content_enabled(True)
        win.refresh_current_tab()
        assert win.current_view == "Data Profile"

    @pytest.mark.parametrize("saved", ["A View That Was Renamed", "Settings"])
    def test_an_unshowable_saved_view_falls_back(self, qapp, tmp_path,
                                                 monkeypatch, saved):
        """
        Two ways a saved name stops being showable: the view was renamed or
        removed, or it is a NON_NAV_VIEWS entry like Settings that has no stack
        page at all. Both must degrade to the default rather than leave the
        window on a blank selection.
        """
        win = self._window_with_last_view(tmp_path, monkeypatch, saved)
        assert win.current_view is None


class TestNonNavViewsAreNotPages:
    """
    Settings is registered in view_defs but placed in no NAV_GROUPS group, so
    setup_views never builds it a stack page - it is reached as a dialog. That
    made show_view's `name not in self.view_defs` guard the wrong test: the name
    passed it and then raised KeyError on the tab_widgets lookup. It only fired
    with a file open, because the empty-state return happens first, which is
    exactly the kind of conditional crash that survives a manual click-through.
    """

    def test_settings_is_registered_but_has_no_page(self, window):
        assert "Settings" in window.view_defs
        assert "Settings" in ui.NON_NAV_VIEWS
        assert "Settings" not in window.tab_widgets, \
            "a non-nav view must not occupy a stack page"

    def test_showing_a_non_nav_view_is_a_no_op_not_a_crash(self, window):
        window.show_view("Null Analysis")
        window.show_view("Settings")
        assert window.current_view == "Null Analysis", \
            "show_view must not half-switch to a view it cannot display"

    def test_showing_an_unregistered_view_is_a_no_op(self, window):
        window.show_view("Column Check")
        window.show_view("No Such View")
        assert window.current_view == "Column Check"


class TestOverviewWaitsToBeAsked:
    """
    The findings register runs every check in the app - ~1.2s at 100k rows and
    ~3.7s at 300k - and Overview is the view a file open lands on, so running it
    automatically made every open pay for the whole register whether or not
    anyone read it. It now waits for the button.

    The failure mode being guarded is not slowness, it is a false clean bill of
    health: a row of zeros on "Failing" is a claim about data nothing has looked
    at yet.
    """

    def test_opening_a_file_does_not_run_the_register(self, qapp, window):
        window._select_view("Overview")
        _settle(qapp, window, rounds=8)
        assert window._overview_started is False
        assert window._overview_frame is None, \
            "the register was built without anyone asking for it"

    def test_the_prompt_is_shown_and_the_table_is_not(self, qapp, window):
        window._select_view("Overview")
        _settle(qapp, window)
        assert window.overview_prompt.isVisibleTo(window.overview_table.parent())
        assert not window.overview_table.isVisibleTo(
            window.overview_table.parent()), \
            "a window of blank rows is a worse answer than saying nothing ran"

    def test_no_tile_shows_a_zero_before_the_checks_run(self, qapp, window):
        """
        The whole point of the gate. A zero on "Failing" beside an unchecked
        file is the most misleading thing this view could say.
        """
        window._select_view("Overview")
        _settle(qapp, window)
        for label, (_tile, value) in window.overview_kpis.items():
            assert value.text() == "—", \
                f"{label} shows {value.text()!r} before any check has run"

    def test_the_button_runs_it(self, qapp, window):
        window._select_view("Overview")
        _settle(qapp, window)
        window.btn_overview_run.click()
        _settle(qapp, window, rounds=8)

        assert window._overview_started is True
        assert window._overview_frame is not None
        assert window.overview_table.model().rowCount() > 0
        assert window.overview_table.isVisibleTo(
            window.overview_table.parent())
        assert not window.overview_prompt.isVisibleTo(
            window.overview_table.parent())

    def test_the_file_facts_are_shown_without_running_anything(self, qapp,
                                                               window):
        """
        Rows, fields and the column map are known from loading, so withholding
        them would make the view emptier than it needs to be.
        """
        window._select_view("Overview")
        _settle(qapp, window)
        assert f"{len(window.df):,} rows" in window.overview_facts.text()

    def test_a_new_sheet_asks_again(self, qapp, window, tmp_path, ticket_df):
        """
        Having run the checks on one sheet says nothing about the next, and
        carrying the flag over would put back the whole cost of a file open.
        """
        window._select_view("Overview")
        _settle(qapp, window)
        window.btn_overview_run.click()
        _settle(qapp, window, rounds=8)
        assert window._overview_started is True

        csv = tmp_path / "second.csv"
        ticket_df.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window, rounds=8)

        assert window._overview_started is False, \
            "the new sheet inherited the previous sheet's consent"
        assert window._overview_frame is None

    def test_revisiting_after_running_keeps_the_register(self, qapp, window):
        """
        The gate is per sheet, not per visit: having asked once, walking away
        and coming back must not throw the answer away.
        """
        window._select_view("Overview")
        _settle(qapp, window)
        window.btn_overview_run.click()
        _settle(qapp, window, rounds=8)

        window._select_view("Data Profile")
        _settle(qapp, window)
        window._select_view("Overview")
        _settle(qapp, window, rounds=8)

        assert window._overview_frame is not None
        assert window.overview_table.model().rowCount() > 0

    def test_a_failed_run_returns_to_the_prompt(self, qapp, window, mocker):
        """
        Otherwise the busy "..." stays for ever with the retry button hidden
        behind it, and the only way out is reloading the sheet.
        """
        window._select_view("Overview")
        _settle(qapp, window)
        mocker.patch.object(ui.ReportGenerator, "_build_findings_df",
                            side_effect=RuntimeError("boom"))
        window.btn_overview_run.click()
        _settle(qapp, window, rounds=8)

        assert window._overview_started is False
        assert window.overview_prompt.isVisibleTo(window.overview_table.parent())
        assert window.btn_overview_run.text() == "Try again"
        for _label, (_tile, value) in window.overview_kpis.items():
            assert value.text() != "...", \
                "a failed run left the tiles claiming to still be working"


class TestReportView:
    """
    The Report view is the one place a result is legitimately plain text - the
    text *is* the artifact the user came for. It is behind a button for the same
    reason Overview's register is: both generators run every check in the app,
    so arriving here must not start one.

    The failure modes guarded here are all silent. A document pane that is
    editable loses the report to a stray keystroke; a pane holding text before
    anyone asked is a report about checks that never ran; and a copy or a save
    offered with nothing behind it writes an empty file.
    """

    def _pane(self, window):
        return window.report_text

    # -- registration ------------------------------------------------------

    def test_the_view_is_registered_and_placed(self, window):
        assert "Report" in dict(window.tab_defs)
        assert "Report" in window.tab_widgets
        assert "Report" in window.view_indices
        assert ui.VIEW_SUBTITLES.get("Report"), "a view needs its one-line subtitle"
        placed = {name for _caption, names in ui.NAV_GROUPS for name in names}
        assert "Report" in placed, "an unplaced view lands in a trailing MORE group"
        assert "Report" in [n for _stage, names in ui.REVIEW_STAGES for n in names]

    def test_it_is_reachable_from_the_sidebar(self, qapp, window):
        row = next(r for r, name in _view_rows(window) if name == "Report")
        window.nav_list.setCurrentRow(row)
        _settle(qapp, window)
        assert window.stack.currentIndex() == window.view_indices["Report"]
        assert window.view_title.text() == "Report"

    def test_the_refresh_is_wired(self, qapp, window):
        """
        Without a refresh_map entry the view builds and then never populates -
        which looks exactly like a view with nothing to say.
        """
        window._select_view("Report")
        _settle(qapp, window)
        assert window.report_prompt.isVisibleTo(window.tab_widgets["Report"]), \
            "refresh_report never ran, so the idle state was never painted"

    # -- the prompt gate ---------------------------------------------------

    def test_arriving_does_not_write_a_report(self, qapp, window):
        window._select_view("Report")
        _settle(qapp, window, rounds=8)
        assert window._report_started is False
        assert window._report_text is None
        assert self._pane(window).toPlainText() == "", \
            "a document appeared without anyone asking for it"

    def test_the_prompt_is_shown_and_the_document_is_not(self, qapp, window):
        window._select_view("Report")
        _settle(qapp, window)
        page = window.tab_widgets["Report"]
        assert window.report_prompt.isVisibleTo(page)
        assert not self._pane(window).isVisibleTo(page), \
            "an empty pane reads as a report that found nothing to say"
        assert not any(button.isVisibleTo(page)
                       for button in window.report_tabs.values()), \
            "a flavour strip over two documents that do not exist yet"

    def test_the_actions_are_unavailable_before_a_report_exists(self, qapp,
                                                                window):
        window._select_view("Report")
        _settle(qapp, window)
        page = window.tab_widgets["Report"]
        for button in (window.btn_report_copy, window.btn_report_save):
            assert not (button.isVisibleTo(page) and button.isEnabled()), \
                "an action offered with nothing behind it"

    def test_the_button_writes_it(self, qapp, window):
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)

        page = window.tab_widgets["Report"]
        assert window._report_started is True
        assert window._report_text
        assert "SANITY CHECK REPORT" in self._pane(window).toPlainText()
        assert self._pane(window).isVisibleTo(page)
        assert not window.report_prompt.isVisibleTo(page)
        assert window.btn_report_copy.isEnabled()
        assert window.btn_report_save.isEnabled()

    def test_a_new_sheet_asks_again(self, qapp, window, tmp_path, ticket_df):
        """Having written a report for one sheet says nothing about the next."""
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)
        assert window._report_started is True

        csv = tmp_path / "second.csv"
        ticket_df.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window, rounds=8)

        assert window._report_started is False, \
            "the new sheet inherited the previous sheet's consent"

    def test_a_failed_run_returns_to_the_prompt(self, qapp, window, mocker):
        """
        Otherwise the note says it is working for ever with the retry button
        hidden behind the pane, and the only way out is reloading the sheet.
        """
        window._select_view("Report")
        _settle(qapp, window)
        mocker.patch.object(ui.ReportGenerator, "generate_text_report",
                            side_effect=RuntimeError("boom"))
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)

        page = window.tab_widgets["Report"]
        assert window._report_started is False
        assert window.report_prompt.isVisibleTo(page)
        assert window.btn_report_run.text() == "Try again"
        assert window._report_text is None

    # -- the document itself -----------------------------------------------

    def test_the_document_is_read_only(self, qapp, window):
        """Read-only means non-editable; it is still selectable and copyable."""
        window._select_view("Report")
        _settle(qapp, window)
        pane = self._pane(window)
        assert isinstance(pane, ui.QPlainTextEdit)
        assert pane.isReadOnly()

    def test_the_document_is_never_html(self, qapp, window):
        """
        setHtml on a results pane is the idiom this codebase removed. The report
        is plain text and has to reach the clipboard as written.
        """
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)
        text = self._pane(window).toPlainText()
        for banned in ("<p>", "<br", "<table", "</"):
            assert banned not in text

    def test_switching_flavour_regenerates(self, qapp, window):
        """
        Two documents built by different code, not two views of one frame - so
        there is nothing to filter and a stale one would be the worst of both.
        """
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)
        full = self._pane(window).toPlainText()

        window.report_tabs[ui.SanityCheckApp.REPORT_SUMMARY].click()
        _settle(qapp, window, rounds=8)
        summary = self._pane(window).toPlainText()

        assert summary != full
        assert summary.splitlines()[0].startswith("Subject: ")
        assert window._report_flavour_shown == \
            ui.SanityCheckApp.REPORT_SUMMARY

    # -- the two actions ---------------------------------------------------

    def test_copy_puts_the_report_on_the_clipboard_without_a_dialog(
            self, qapp, window, monkeypatch):
        shown = []
        monkeypatch.setattr(ui.QMessageBox, "information",
                            staticmethod(lambda *a, **k: shown.append(a)))
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)

        qapp.clipboard().clear()
        window.btn_report_copy.click()
        assert qapp.clipboard().text() == window._report_text
        assert not shown, "a modal on every copy is noise; use the status bar"

    def test_save_writes_the_text_with_a_bom(self, qapp, window, tmp_path,
                                            monkeypatch):
        """
        utf-8-sig: without the BOM Notepad and Excel mis-decode non-ASCII on a
        double-click and mangle exactly the non-English text this tool finds.
        """
        out = tmp_path / "written.txt"
        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: (str(out), "")))
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)

        window.btn_report_save.click()
        assert out.exists()
        assert out.read_bytes().startswith(b"\xef\xbb\xbf")
        assert out.read_text(encoding="utf-8-sig") == window._report_text

    def test_a_cancelled_save_writes_nothing(self, qapp, window, monkeypatch):
        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: ("", "")))
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)
        window.btn_report_save.click()   # must not raise

    def test_an_unwritable_path_is_reported_not_raised(self, qapp, window,
                                                      tmp_path, monkeypatch):
        """A failure goes to the status bar and the log, never to a traceback."""
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(tmp_path / "nope" / "x.txt"), "")))
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)

        seen = []
        monkeypatch.setattr(window, "update_status", seen.append)
        window.btn_report_save.click()
        assert any("not save" in message for message in seen)

    def test_the_note_carries_the_provenance(self, qapp, window):
        """A pasted sentence has to name the flavour and the rows behind it."""
        window._select_view("Report")
        _settle(qapp, window)
        window.btn_report_run.click()
        _settle(qapp, window, rounds=8)

        qapp.clipboard().clear()
        ui._copy_label(window.report_note, with_context=True)
        text = qapp.clipboard().text()
        assert "tickets.xlsx" in text
        assert "Report: " in text
        assert "Rows analyzed: " in text


class TestNavBadges:
    """
    A badge is a count of failing checks, and the rules around it are all
    documented as deliberate. None of them was tested, and every failure mode is
    silent - nothing on screen contradicts a wrong badge.
    """

    def test_a_badge_is_absent_rather_than_zero(self, window):
        """
        A grey 0 beside every clean view reads as a score, and there is
        deliberately no score in this app.
        """
        window.set_nav_badge("Null Analysis", 0)
        item = window._nav_item("Null Analysis")
        assert item.data(ui.NAV_ROLE_BADGE) is None

        window.set_nav_badge("Null Analysis", None)
        assert item.data(ui.NAV_ROLE_BADGE) is None

    def test_a_count_is_thousands_separated(self, window):
        window.set_nav_badge("Null Analysis", 1234)
        assert window._nav_item("Null Analysis").data(ui.NAV_ROLE_BADGE) == "1,234"

    def test_the_badge_lives_in_item_data_not_the_text(self, window):
        """
        Anything in the text would sort, elide and copy as part of the view
        name - the reason NavDelegate exists at all.
        """
        window.set_nav_badge("Logic Checks", 7)
        item = window._nav_item("Logic Checks")
        assert item.text() == "Logic Checks"
        assert item.data(Qt.ItemDataRole.UserRole) == "Logic Checks"
        assert item.data(ui.NAV_ROLE_BADGE) == "7"

    def test_clear_nav_badges_clears_every_row(self, window):
        window.set_nav_badge("Null Analysis", 12)
        window.set_nav_badge("Logic Checks", 3)
        window.clear_nav_badges()
        assert all(window._nav_item(name).data(ui.NAV_ROLE_BADGE) is None
                   for name in ("Null Analysis", "Logic Checks"))

    def test_a_badge_does_not_survive_a_sheet_change(self, qapp, window,
                                                    tmp_path, ticket_df):
        """
        A badge counted on one sheet must not sit beside another.

        Driven through load_sheet rather than by calling clear_nav_badges
        directly, because the invariant that matters is that the *load path*
        clears them - a test on the method alone still passes with the call
        deleted from on_loaded, which is the actual regression.

        Starting on a view that is **not** Overview is load-bearing, and is the
        whole reason this test bites. _apply_overview_badges clears the badges
        itself before writing its own, so with Overview on screen the load-path
        call is masked and deleting it changes nothing observable. It is only
        when the refresh after a load goes to some other view - nobody having
        opened Overview for the new sheet yet - that on_loaded is the one thing
        standing between the reader and sheet A's counts beside sheet B.

        The probe is a sentinel no real count could produce.
        """
        window.show_view("Data Profile")
        assert window.current_view == "Data Profile"

        csv = tmp_path / "other_sheet.csv"
        ticket_df.to_csv(csv, index=False)
        window.filepath = str(csv)

        sentinel = "9,999"
        for name, _ in window.tab_defs:
            window.set_nav_badge(name, 9999)

        window.load_sheet(None)
        _settle(qapp, window)

        assert window.current_view == "Data Profile", \
            "the load must not have navigated to Overview, or this proves nothing"
        survivors = sorted(
            name for name, _ in window.tab_defs
            if window._nav_item(name) is not None
            and window._nav_item(name).data(ui.NAV_ROLE_BADGE) == sentinel)
        assert not survivors, \
            f"a previous sheet's badge survived on: {survivors}"

    def test_fresh_badges_replace_the_previous_sheets(self, qapp, window,
                                                     tmp_path, ticket_df):
        """
        The other half: where the new sheet *does* have findings, the badge must
        be its own count. Asserting an empty sidebar would be wrong here - the
        register legitimately re-badges what it finds.

        The register is the only writer of badges and it no longer runs on load,
        so it is started explicitly. Leaving that out would make this test pass
        for the wrong reason - an empty sidebar satisfies the sentinel check on
        its own, which is exactly what the second assertion below is for.
        """
        csv = tmp_path / "reload.csv"
        ticket_df.to_csv(csv, index=False)
        window.filepath = str(csv)

        for name, _ in window.tab_defs:
            window.set_nav_badge(name, 9999)

        window.load_sheet(None)
        _settle(qapp, window)

        window._select_view("Overview")
        _settle(qapp, window)
        window.btn_overview_run.click()
        _settle(qapp, window, rounds=8)

        badges = {name: window._nav_item(name).data(ui.NAV_ROLE_BADGE)
                  for name, _ in window.tab_defs
                  if window._nav_item(name) is not None}
        assert "9,999" not in badges.values()
        # Not vacuous: this fixture reliably fails a null and a logic check, so
        # if nothing at all were written the assertion above would prove nothing.
        assert any(badges.values()), \
            f"the register wrote no badge at all: {badges}"

    def test_clearing_keeps_the_row_index(self, window):
        """The index is separate data and must not be collateral damage."""
        item = window._nav_item("Null Analysis")
        before = item.data(ui.NAV_ROLE_INDEX)
        window.set_nav_badge("Null Analysis", 4)
        window.clear_nav_badges()
        assert item.data(ui.NAV_ROLE_INDEX) == before
        assert before is not None

    def test_a_badge_on_an_unknown_view_is_ignored(self, window):
        window.set_nav_badge("No Such View", 5)      # must not raise


class TestFindingsModelColouring:
    """
    RESULT_STATES is described as "the same vocabulary reporter.py sorts by, so
    the screen and the workbook cannot disagree". The reporter half is well
    covered; the GUI half had no test, so drift would show up as an uncoloured
    or - worse - a green "Not checked".
    """

    def _brush(self, model, row, column_name):
        column = list(model._df.columns).index(column_name)
        index = model.index(row, column)
        return model.data(index, Qt.ItemDataRole.ForegroundRole)

    def _model(self, results):
        frame = pd.DataFrame({"Check": [f"c{i}" for i in range(len(results))],
                              "Result": results})
        return ui.FindingsModel(frame)

    @pytest.mark.parametrize("result,token", [
        ("Fail", "COLOR_ERROR"),
        ("Review", "COLOR_WARNING"),
        ("Not checked", "COLOR_WARNING"),
        ("Pass", "COLOR_SUCCESS"),
    ])
    def test_each_outcome_takes_its_state_colour(self, result, token):
        brush = self._brush(self._model([result]), 0, "Result")
        assert brush is not None, f"{result!r} lost its colour"
        assert brush.color().name() == getattr(ui, token).lower()

    def test_not_checked_is_never_green(self):
        """
        The one rule CLAUDE.md states "here as everywhere": a check that could
        not run must never be coloured like one that passed.
        """
        not_checked = self._brush(self._model(["Not checked"]), 0, "Result")
        passed = self._brush(self._model(["Pass"]), 0, "Result")
        assert not_checked.color().name() != passed.color().name()
        assert not_checked.color().name() == ui.COLOR_WARNING.lower()

    def test_an_unknown_outcome_is_left_uncoloured(self):
        """A reporter rename must not silently pick up another state's colour."""
        assert self._brush(self._model(["Not run"]), 0, "Result") is None

    def test_only_the_outcome_column_is_tinted(self):
        model = self._model(["Fail"])
        assert self._brush(model, 0, "Check") is None

    def test_the_severity_vocabulary_agrees_with_the_result_one(self):
        """
        Both dicts encode the same three meanings; the shared rule is that only
        Pass is ok and Not checked is warn in each.
        """
        for vocabulary in (ui.RESULT_STATES, ui.SEVERITY_STATES):
            assert vocabulary.get("Pass") == "ok"
            assert vocabulary.get("Not checked") == "warn"
            assert set(vocabulary.values()) <= {"ok", "warn", "error"}

    def test_every_state_has_a_colour(self):
        for state in set(ui.RESULT_STATES.values()) | set(ui.SEVERITY_STATES.values()):
            assert state in ui._STATE_COLOURS

    def test_a_cell_carries_its_full_text_as_a_tooltip(self):
        """The Detail column is elided, so the tooltip is how it stays readable."""
        model = self._model(["Fail"])
        column = list(model._df.columns).index("Check")
        assert model.data(model.index(0, column),
                          Qt.ItemDataRole.ToolTipRole) == "c0"


class TestFindingsLeadSomewhere:
    """
    The register is only useful if a row leads somewhere. The fallback is
    `.get(area, "Column Check")`, so an Area the reporter adds later lands on
    the wrong view *silently* - which is what makes this worth pinning.
    """

    def test_every_area_the_reporter_emits_is_mapped(self, window):
        from core.reporter import ReportGenerator
        register = ReportGenerator(window.df, window.analyzer,
                                   window.filepath)._build_findings_df()
        areas = set(register["Area"].astype(str))
        unmapped = areas - set(ui.SanityCheckApp.OVERVIEW_AREA_VIEWS)
        assert not unmapped, f"these areas fall back to Column Check: {unmapped}"

    def test_every_mapped_view_actually_exists(self, window):
        registered = {name for name, _ in window.tab_defs}
        for area, view in ui.SanityCheckApp.OVERVIEW_AREA_VIEWS.items():
            assert view in registered, f"{area} maps to missing view {view}"

    def test_a_date_format_finding_goes_to_logic_checks(self, window):
        """
        Filed under Structure by the reporter, corrected in Logic Checks - the
        documented reason Check is consulted before Area.
        """
        assert window._view_for_finding(
            "Structure", "Date format of opened_at") == "Logic Checks"
        assert window._view_for_finding("Structure", "Required columns") == \
            "Column Check"

    def test_an_unknown_area_still_lands_on_a_real_view(self, window):
        target = window._view_for_finding("Something New", "whatever")
        assert target in {name for name, _ in window.tab_defs}


class TestExtractView:
    """
    The extract exists because Excel hangs on the filter its users need. Its one
    real advantage over doing it by hand is the live count, so most of what is
    worth pinning here is that the count tracks the selection and tells the
    truth about what will come out.
    """

    def _show(self, qapp, window):
        window.show_view("Extract")
        _settle(qapp, window)

    def _load(self, qapp, window, frame, tmp_path, name="src.csv"):
        """Swap the data the way the app does, so per-sheet state is reset too."""
        csv = tmp_path / name
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)

    def _month_rows(self, frame, period):
        return int((frame["opened_at"].dt.to_period("M").astype(str)
                    == period).sum())

    # -- the selection UI -------------------------------------------------

    def test_the_view_is_registered_and_populates(self, qapp, window):
        assert "Extract" in dict(window.tab_defs)
        assert "Extract" in [name for _i, name in _view_rows(window)]
        self._show(qapp, window)
        assert window.extract_col_list.count() == len(window.df.columns)
        assert window.extract_from_combo.count() > 0

    def test_columns_start_from_the_detected_standard_set(self, qapp, window):
        """Not a guess: they are the columns every other check already uses."""
        self._show(qapp, window)
        assert set(window._extract_selected_columns()) == \
            set(window._extract_standard_columns())
        assert "mostly_empty" not in window._extract_selected_columns()

    def test_every_column_is_offered_not_only_the_standard_ones(self, qapp, window):
        self._show(qapp, window)
        # The name lives in item data; the label also carries the fill share, so
        # reading a name off the text would return "state   (100% filled)".
        offered = [window.extract_col_list.item(i).data(Qt.ItemDataRole.UserRole)
                   for i in range(window.extract_col_list.count())]
        assert offered == [str(c) for c in window.df.columns], \
            "column order must follow the source, not the tick order"

    def test_the_label_shows_how_full_each_column_is(self, qapp, window):
        """It is the guide to what is worth keeping on a wide export."""
        self._show(qapp, window)
        labels = {window.extract_col_list.item(i).data(Qt.ItemDataRole.UserRole):
                  window.extract_col_list.item(i).text()
                  for i in range(window.extract_col_list.count())}
        assert "100% filled" in labels["number"]
        assert "25% filled" in labels["mostly_empty"], \
            "2 values in 8 rows"

    def test_the_filter_narrows_the_list_without_losing_ticks(self, qapp, window):
        self._show(qapp, window)
        window._extract_select_standard()
        before = window._extract_selected_columns()

        window.extract_col_filter.setText("descr")
        _settle(qapp, window)
        shown = [item.data(Qt.ItemDataRole.UserRole)
                 for item in window._extract_visible_column_items()]
        assert shown and all("descr" in name for name in shown)
        assert window._extract_selected_columns() == before, \
            "filtering is a view, not a selection"

        window.extract_col_filter.clear()
        _settle(qapp, window)
        assert len(window._extract_visible_column_items()) == \
            len(window.df.columns)

    def test_a_preset_acts_on_what_is_shown(self, qapp, window):
        """
        Narrow to "desc", press All, and the columns chosen elsewhere stay
        chosen - otherwise the filter would be a trap.
        """
        self._show(qapp, window)
        window._extract_set_checked({"number"})
        window.extract_col_filter.setText("descr")
        _settle(qapp, window)
        window._extract_select_all()
        _settle(qapp, window)

        chosen = window._extract_selected_columns()
        assert "number" in chosen, "a hidden tick must survive a preset"
        assert "description" in chosen and "short_description" in chosen
        assert "state" not in chosen

    def test_non_empty_drops_the_columns_nobody_filled(self, qapp, window,
                                                       ticket_df, tmp_path):
        frame = ticket_df.copy()
        frame["never_used"] = None
        frame["blank_text"] = ""
        self._load(qapp, window, frame, tmp_path)
        self._show(qapp, window)

        window._extract_select_non_empty()
        _settle(qapp, window)
        chosen = window._extract_selected_columns()
        assert "never_used" not in chosen
        assert "blank_text" not in chosen, \
            "a sheet full of empty strings is not data"
        assert "number" in chosen and "mostly_empty" in chosen

    def test_the_count_follows_the_range(self, qapp, window):
        self._show(qapp, window)
        window._extract_select_all()
        full = window._extract_plan.rows

        one_month = window.extract_from_combo.itemData(0)
        window.extract_from_combo.setCurrentIndex(0)
        window.extract_to_combo.setCurrentIndex(0)
        _settle(qapp, window)

        assert window._extract_plan.rows < full
        assert window._extract_plan.rows == self._month_rows(window.df, one_month)

    def test_the_closing_month_is_kept_whole(self, qapp, window):
        """
        Writing the range as `<= start_of(To)` keeps only the 1st of that month.
        The fixture dates all fall on the 15th, so that bug reads as zero rows.
        """
        self._show(qapp, window)
        last = window.extract_to_combo.count() - 1
        period = window.extract_to_combo.itemData(last)
        window.extract_from_combo.setCurrentIndex(last)
        window.extract_to_combo.setCurrentIndex(last)
        _settle(qapp, window)
        assert window._extract_plan.rows == self._month_rows(window.df, period) > 0

    def test_the_count_follows_the_columns(self, qapp, window):
        self._show(qapp, window)
        window._extract_select_all()
        assert window._extract_plan.columns == len(window.df.columns)
        window._extract_select_standard()
        assert window._extract_plan.columns == len(window._extract_standard_columns())

    def test_the_column_title_counts_the_ticks(self, qapp, window):
        self._show(qapp, window)
        window._extract_select_none()
        assert f"0 of {len(window.df.columns)}" in window.extract_cols_title.text()
        window._extract_select_all()
        assert (f"{len(window.df.columns)} of {len(window.df.columns)}"
                in window.extract_cols_title.text())

    def test_an_empty_selection_is_refused_not_written_empty(self, qapp, window):
        self._show(qapp, window)
        window._extract_select_none()
        assert not window.btn_extract.isEnabled()
        assert "No columns" in window.extract_summary.text()

    def test_every_date_column_can_be_filtered_by(self, qapp, window, tmp_path):
        """
        "The tickets updated in March" is as reasonable as "opened in March",
        and a wide export carries several date columns beyond the detected pair.
        """
        frame = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(12)],
            "state": ["Closed"] * 12,
            "opened_at": pd.date_range("2025-01-01", periods=12, freq="ME"),
            "resolved_at": pd.date_range("2025-01-05", periods=12, freq="ME"),
            "sys_updated_on": pd.date_range("2025-02-01", periods=12, freq="ME"),
            "due_date": pd.date_range("2025-03-01", periods=12, freq="ME"),
            "short_description": ["A long enough description here"] * 12,
        })
        window.df = frame
        window.analyzer = SanityAnalyzer(frame)
        window.tab_widgets["Extract"] = window.tab_widgets["Extract"]
        window.show_view("Extract")
        window.refresh_extract()
        _settle(qapp, window)

        offered = [window.extract_date_combo.itemData(i)
                   for i in range(window.extract_date_combo.count())]
        assert "sys_updated_on" in offered
        assert "due_date" in offered
        assert None in offered, "the no-filter escape must stay"

        window.extract_date_combo.setCurrentIndex(
            window.extract_date_combo.findData("sys_updated_on"))
        _settle(qapp, window)
        assert window._extract_plan.rows > 0
        assert window._extract_dates() is not None

    def test_the_detected_columns_lead_the_list(self, qapp, window):
        """They are what every other view means by a date."""
        window.show_view("Extract")
        _settle(qapp, window)
        offered = [window.extract_date_combo.itemData(i)
                   for i in range(window.extract_date_combo.count())]
        assert offered[0] == window.analyzer.created_column

    def test_switching_the_date_column_repopulates_the_months(self, qapp, window):
        """resolved_at has its own months; keeping the old ones would mislead."""
        self._show(qapp, window)
        index = window.extract_date_combo.findData("resolved_at")
        assert index >= 0
        window.extract_date_combo.setCurrentIndex(index)
        _settle(qapp, window)
        assert window.extract_from_combo.count() > 0
        assert window._extract_plan.rows > 0

    def test_no_date_filter_keeps_every_row(self, qapp, window):
        self._show(qapp, window)
        window.extract_date_combo.setCurrentIndex(
            window.extract_date_combo.findData(None))
        _settle(qapp, window)
        assert window._extract_plan.rows == len(window.df)
        assert not window.extract_undated_check.isVisibleTo(
            window.extract_undated_check.parentWidget()), \
            "with no date filter nothing is being dropped, so the box misleads"

    # -- honesty ----------------------------------------------------------

    def test_undated_rows_are_named_with_their_count(self, qapp, window,
                                                     ticket_df, tmp_path):
        frame = ticket_df.copy()
        frame.loc[frame.index[:3], "opened_at"] = pd.NaT
        self._load(qapp, window, frame, tmp_path)
        self._show(qapp, window)
        assert "3 rows with no date" in window.extract_undated_check.text()

        before = window._extract_plan.rows
        window.extract_undated_check.setChecked(True)
        _settle(qapp, window)
        assert window._extract_plan.rows == before + 3, \
            "rows the filter cannot place must be surfaced, never dropped silently"

    def test_a_large_selection_says_excel_will_struggle(self, qapp, window,
                                                       monkeypatch):
        monkeypatch.setattr(ui.extract_core, "EXCEL_SLOW_CELLS", 1)
        self._show(qapp, window)
        window._extract_select_all()
        assert "slow" in window.extract_summary.text().lower()
        assert window.extract_result_card.property("severity") == "warn"

    def test_over_the_row_ceiling_blocks_xlsx_but_not_csv(self, qapp, window,
                                                         monkeypatch):
        monkeypatch.setattr(ui.extract_core, "DETAIL_MAX_DATA_ROWS", 2)
        self._show(qapp, window)
        window._extract_select_all()
        assert not window.btn_extract.isEnabled()
        assert window.extract_result_card.property("severity") == "error"

        window.extract_format_combo.setCurrentIndex(
            window.extract_format_combo.findData(ui.extract_core.FORMAT_CSV))
        _settle(qapp, window)
        assert window.btn_extract.isEnabled(), \
            "CSV has no row ceiling, so it must stay offered"

    def test_an_ambiguous_date_format_is_surfaced(self, qapp, window, tmp_path):
        frame = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(6)],
            "opened_at": ["01/02/2024", "03/04/2024", "05/06/2024",
                          "07/08/2024", "09/10/2024", "11/12/2024"],
            "short_description": ["A long enough description here"] * 6,
        })
        self._load(qapp, window, frame, tmp_path, name="ambiguous.csv")
        self._show(qapp, window)
        assert window.extract_format_warning.isVisibleTo(
            window.extract_format_warning.parentWidget()), \
            "month buckets built on a guessed date order are simply wrong"
        assert "guess" in window.extract_format_warning.text().lower()

    # -- writing ----------------------------------------------------------

    def _prepare(self, tmp_path, monkeypatch, filename="extract.xlsx"):
        out = tmp_path / filename
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k:
                         shown.append((title, text))))
        return out, shown

    def test_the_extract_is_written(self, qapp, window, tmp_path, monkeypatch):
        out, shown = self._prepare(tmp_path, monkeypatch)
        self._show(qapp, window)
        window.run_extract()
        _settle(qapp, window)
        assert out.exists()
        assert shown, "the user must be told where it went"
        assert str(out) in shown[-1][1]

    def test_the_button_is_gated_during_the_write(self, qapp, window):
        self._show(qapp, window)
        window._begin_extract()
        assert not window.btn_extract.isEnabled()
        window._finish_extract()
        assert window.btn_extract.isEnabled()

    def test_progress_reaches_the_status_bar(self, qapp, window, tmp_path,
                                            monkeypatch):
        self._prepare(tmp_path, monkeypatch)
        seen = []
        monkeypatch.setattr(window, "update_status", seen.append)
        self._show(qapp, window)
        window.run_extract()
        _settle(qapp, window)
        assert seen

    def test_the_dialog_says_undated_rows_were_left_out(self, qapp, window,
                                                        tmp_path, monkeypatch,
                                                        ticket_df):
        """A row that silently vanished is the failure this guards."""
        frame = ticket_df.copy()
        frame.loc[frame.index[:2], "opened_at"] = pd.NaT
        self._load(qapp, window, frame, tmp_path)
        _out, shown = self._prepare(tmp_path, monkeypatch)
        self._show(qapp, window)
        window.run_extract()
        _settle(qapp, window)
        assert "2 rows with no date were left out" in shown[-1][1]

    def test_the_selection_is_remembered(self, qapp, window, tmp_path,
                                         monkeypatch):
        self._prepare(tmp_path, monkeypatch)
        self._show(qapp, window)
        window._extract_select_all()
        window.extract_format_combo.setCurrentIndex(
            window.extract_format_combo.findData(ui.extract_core.FORMAT_CSV))
        window.run_extract()
        _settle(qapp, window)

        saved = ConfigManager(window.config_mgr.config_file).config
        assert saved["extract_columns"] == [str(c) for c in window.df.columns]
        assert saved["extract_format"] == ui.extract_core.FORMAT_CSV
        assert saved["extract_date_column"] == "opened_at"

    def test_a_remembered_selection_is_restored(self, qapp, window, ticket_df,
                                                tmp_path):
        window.config["extract_columns"] = ["number", "state"]
        self._load(qapp, window, ticket_df, tmp_path)
        self._show(qapp, window)
        assert window._extract_selected_columns() == ["number", "state"]

    def test_a_remembered_column_missing_from_this_file_is_ignored(
            self, qapp, window, ticket_df, tmp_path):
        """A different export shares the setting but not the columns."""
        window.config["extract_columns"] = ["number", "not_in_this_file"]
        self._load(qapp, window, ticket_df, tmp_path)
        self._show(qapp, window)
        assert window._extract_selected_columns() == ["number"]

    def test_a_hand_made_selection_survives_leaving_the_view(self, qapp, window):
        """
        refresh_* runs on every visit, so restoring the remembered set there
        would undo the ticks the user just made on the way past another view.
        """
        self._show(qapp, window)
        window._extract_set_checked(["number", "priority"])
        window.show_view("Raw Data")
        _settle(qapp, window)
        self._show(qapp, window)
        assert window._extract_selected_columns() == ["number", "priority"]

    def test_a_hand_made_selection_does_not_carry_to_the_next_sheet(
            self, qapp, window, ticket_df, tmp_path):
        """It described the previous sheet, and may not even fit this one."""
        self._show(qapp, window)
        window._extract_set_checked(["number", "priority"])
        self._load(qapp, window, ticket_df, tmp_path)
        self._show(qapp, window)
        assert set(window._extract_selected_columns()) == \
            set(window._extract_standard_columns())



class TestExtractMonthPicker:
    """
    The people using this name their ranges "Feb25-Jul26", and an export pulled
    on the 12th has a closing month that looks like a collapse in volume unless
    something says otherwise.
    """

    def _load(self, qapp, window, first, last, tmp_path, name="span.csv"):
        """A frame whose opened_at runs from `first` to `last` inclusive."""
        stamps = pd.date_range(first, last, freq="12h")
        n = len(stamps)
        frame = pd.DataFrame({
            "number": [f"INC{i:06d}" for i in range(n)],
            "state": ["Closed", "New"] * (n // 2) + ["New"] * (n % 2),
            "priority": ["P1", "P2", "P3", "P2"] * (n // 4) + ["P1"] * (n % 4),
            "opened_at": stamps,
            "short_description": ["Password reset for the user account"] * n,
        })
        csv = tmp_path / name
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)
        return frame

    def _alert_visible(self, window):
        row = window.extract_partial_row
        return row.isVisibleTo(row.parentWidget())

    def _trim_offered(self, window):
        """
        isVisibleTo, not isVisible: the window is never shown in these tests, so
        isVisible() is False for every child and an assertion that the button is
        hidden would pass without testing anything.
        """
        button = window.btn_extract_trim
        return button.isVisibleTo(button.parentWidget())

    # -- MMM-YY labels ----------------------------------------------------

    def test_months_are_offered_as_mmm_yy(self, qapp, window, tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-30", tmp_path)
        combo = window.extract_from_combo
        shown = [combo.itemText(i).split("  ")[0] for i in range(combo.count())]
        assert shown == ["Feb-25", "Mar-25", "Apr-25"]

    def test_the_iso_period_stays_the_item_data(self, qapp, window, tmp_path):
        """Every consumer downstream keys off it - filenames included."""
        self._load(qapp, window, "2025-02-01", "2025-04-30", tmp_path)
        combo = window.extract_from_combo
        assert [combo.itemData(i) for i in range(combo.count())] == \
            ["2025-02", "2025-03", "2025-04"]

    def test_the_row_count_is_still_shown(self, qapp, window, tmp_path):
        """It is the thing Excel cannot tell you without filtering first."""
        self._load(qapp, window, "2025-02-01", "2025-04-30", tmp_path)
        assert "(" in window.extract_from_combo.itemText(0)

    def test_a_complete_month_carries_the_iso_code_as_its_tooltip(
            self, qapp, window, tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-30", tmp_path)
        combo = window.extract_from_combo
        tip = combo.itemData(combo.findData("2025-03"),
                             Qt.ItemDataRole.ToolTipRole)
        assert tip == "2025-03"

    # -- the incomplete-month alert ---------------------------------------

    def test_a_complete_span_raises_no_alert(self, qapp, window, tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-30", tmp_path)
        assert not self._alert_visible(window)

    def test_an_export_taken_mid_month_is_flagged(self, qapp, window, tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        assert self._alert_visible(window)
        text = window.extract_partial_warning.text()
        assert "Apr-25" in text
        assert "day 12 of 30" in text, "the shortfall must be in days, not adjectives"

    def test_the_partial_month_is_marked_where_it_is_chosen(
            self, qapp, window, tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        combo = window.extract_to_combo
        april = combo.findData("2025-04")
        assert "part month" in combo.itemText(april)
        assert "part month" not in combo.itemText(combo.findData("2025-03"))
        tip = combo.itemData(april, Qt.ItemDataRole.ToolTipRole)
        assert "stops on day 12" in tip

    def test_a_weekday_only_queue_is_not_accused(self, qapp, window, tmp_path):
        """
        Judged on the span of days present, not how many days have rows - a
        queue with nothing at weekends is missing eight or nine days a month
        and is perfectly complete.
        """
        stamps = pd.bdate_range("2025-02-01", "2025-04-30")
        frame = pd.DataFrame({
            "number": [f"INC{i:05d}" for i in range(len(stamps))],
            "opened_at": stamps,
            "state": ["Closed"] * len(stamps),
            "short_description": ["Password reset for the account"] * len(stamps),
        })
        csv = tmp_path / "weekdays.csv"
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)
        assert not self._alert_visible(window)

    def test_a_month_outside_the_range_is_not_reported(self, qapp, window,
                                                       tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        combo = window.extract_to_combo
        combo.setCurrentIndex(combo.findData("2025-03"))
        _settle(qapp, window)
        assert not self._alert_visible(window)

    def test_no_date_filter_means_no_alert(self, qapp, window, tmp_path):
        """Nothing is being filtered, so there is no month to exclude."""
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        window.extract_date_combo.setCurrentIndex(
            window.extract_date_combo.findData(None))
        _settle(qapp, window)
        assert not self._alert_visible(window)

    # -- the exclude action -----------------------------------------------

    def test_exclude_narrows_past_both_partial_ends(self, qapp, window,
                                                    tmp_path):
        self._load(qapp, window, "2025-01-19", "2025-04-12", tmp_path)
        assert self._trim_offered(window)
        before = window._extract_plan.rows

        window._extract_trim_partial_months()
        _settle(qapp, window)

        assert window.extract_from_combo.currentData() == "2025-02"
        assert window.extract_to_combo.currentData() == "2025-03"
        assert window._extract_plan.rows < before
        assert not self._alert_visible(window), "the alert must clear once acted on"

    def test_exclude_moves_only_the_end_that_is_partial(self, qapp, window,
                                                        tmp_path):
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        window._extract_trim_partial_months()
        _settle(qapp, window)
        assert window.extract_from_combo.currentData() == "2025-02", \
            "a complete opening month must not be discarded"
        assert window.extract_to_combo.currentData() == "2025-03"

    def test_exclude_is_a_no_op_once_the_range_is_clean(self, qapp, window,
                                                        tmp_path):
        self._load(qapp, window, "2025-01-19", "2025-04-12", tmp_path)
        window._extract_trim_partial_months()
        _settle(qapp, window)
        window._extract_trim_partial_months()
        _settle(qapp, window)
        assert window.extract_from_combo.currentData() == "2025-02"
        assert window.extract_to_combo.currentData() == "2025-03"

    def test_a_single_partial_month_is_not_offered_a_dead_button(
            self, qapp, window, tmp_path):
        """There is nothing to narrow to, and a button that does nothing is
        worse than no button."""
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        april = window.extract_from_combo.findData("2025-04")
        window.extract_from_combo.setCurrentIndex(april)
        window.extract_to_combo.setCurrentIndex(april)
        _settle(qapp, window)

        assert self._alert_visible(window), "it is still a partial month"
        assert not self._trim_offered(window)
        assert "only month selected" in window.extract_partial_warning.text()

        window._extract_trim_partial_months()
        _settle(qapp, window)
        assert window.extract_from_combo.currentData() == "2025-04"
        assert window.extract_to_combo.currentData() == "2025-04"

    def test_an_interior_gap_is_named_but_not_trimmable(self, qapp, window,
                                                        tmp_path):
        """Excluding it would put a hole in the extract, not shorten it."""
        days = [d for d in pd.date_range("2025-02-01", "2025-04-30")
                if not (d.month == 3 and 8 <= d.day <= 22)]
        frame = pd.DataFrame({
            "number": [f"INC{i:05d}" for i in range(len(days))],
            "opened_at": pd.to_datetime(days),
            "state": ["Closed"] * len(days),
            "short_description": ["Password reset for the account"] * len(days),
        })
        csv = tmp_path / "gap.csv"
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)

        assert self._alert_visible(window)
        text = window.extract_partial_warning.text()
        assert "Mar-25" in text and "stretch with no rows" in text
        assert not self._trim_offered(window)
        assert "cannot be excluded by shortening" in text

    # -- disclosure in the output ------------------------------------------

    def test_the_dialog_names_the_range_and_the_partial_month(
            self, qapp, window, tmp_path, monkeypatch):
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        out = tmp_path / "extract.xlsx"
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k:
                         shown.append((title, text))))
        window.run_extract()
        _settle(qapp, window)

        assert out.exists()
        text = shown[-1][1]
        assert "Feb-25 to Apr-25" in text
        assert "Apr-25 is not a whole month" in text

    def test_the_workbook_records_the_partial_month(self, qapp, window,
                                                    tmp_path, monkeypatch):
        """The caveat has to travel with the file, not stay on the screen."""
        from openpyxl import load_workbook

        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        out = tmp_path / "recorded.xlsx"
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(out), "")))
        monkeypatch.setattr(ui.QMessageBox, "information",
                            staticmethod(lambda *a, **k: None))
        window.run_extract()
        _settle(qapp, window)

        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["Incomplete months included"] == "Apr-25"
        assert "Feb-25 to Apr-25" in str(rows["Months included"])

    def test_a_csv_extract_gets_its_provenance_beside_it(
            self, qapp, window, tmp_path, monkeypatch):
        self._load(qapp, window, "2025-02-01", "2025-04-12", tmp_path)
        out = tmp_path / "extract.csv"
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k:
                         shown.append((title, text))))
        window.extract_format_combo.setCurrentIndex(
            window.extract_format_combo.findData(ui.extract_core.FORMAT_CSV))
        _settle(qapp, window)
        window.run_extract()
        _settle(qapp, window)

        sidecar = tmp_path / "extract_ExtractInfo.txt"
        assert sidecar.exists()
        text = sidecar.read_text(encoding="utf-8-sig")
        assert "Date column used: opened_at" in text
        assert "Incomplete months included: Apr-25" in text
        assert sidecar.name in shown[-1][1], \
            "the user has to be told the note exists, or it is not disclosure"


class TestExtractIndividualMonths:
    """
    A range cannot express "February, May and July", which is a real request.
    The two modes sit side by side: the range combos, or a tick list.
    """

    def _load(self, qapp, window, first, last, tmp_path, name="span.csv"):
        stamps = pd.date_range(first, last, freq="12h")
        n = len(stamps)
        frame = pd.DataFrame({
            "number": [f"INC{i:06d}" for i in range(n)],
            "state": ["Closed", "New"] * (n // 2) + ["New"] * (n % 2),
            "opened_at": stamps,
            "short_description": ["Password reset for the user account"] * n,
        })
        csv = tmp_path / name
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)
        return frame

    def _pick_mode(self, qapp, window):
        combo = window.extract_month_mode_combo
        combo.setCurrentIndex(combo.findData(ui.EXTRACT_MONTHS_PICK))
        _settle(qapp, window)

    def _shown(self, widget):
        return widget.isVisibleTo(widget.parentWidget())

    # -- the two modes -----------------------------------------------------

    def test_range_is_the_default_and_the_list_is_hidden(self, qapp, window,
                                                         tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        assert window._extract_month_mode() == ui.EXTRACT_MONTHS_RANGE
        assert self._shown(window.extract_range_row)
        assert not self._shown(window.extract_months_panel)

    def test_picking_swaps_the_control(self, qapp, window, tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        assert self._shown(window.extract_months_panel)
        assert not self._shown(window.extract_range_row)

    def test_the_mode_is_read_from_item_data_not_display_text(self, qapp, window,
                                                             tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        combo = window.extract_month_mode_combo
        assert sorted(combo.itemData(i) for i in range(combo.count())) == \
            sorted([ui.EXTRACT_MONTHS_RANGE, ui.EXTRACT_MONTHS_PICK])

    def test_the_list_offers_every_month_in_mmm_yy(self, qapp, window, tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        widget = window.extract_month_list
        labels = [widget.item(i).text().split("  ")[0]
                  for i in range(widget.count())]
        assert labels == ["Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26",
                          "Jun-26"]
        assert [widget.item(i).data(Qt.ItemDataRole.UserRole)
                for i in range(widget.count())] == \
            ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]

    # -- picking -----------------------------------------------------------

    def test_a_non_contiguous_pick_is_counted_correctly(self, qapp, window,
                                                        tmp_path):
        frame = self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02", "2026-05"])
        _settle(qapp, window)

        months = frame["opened_at"].dt.to_period("M").astype(str)
        expected = int(months.isin(["2026-02", "2026-05"]).sum())
        assert window._extract_plan.rows == expected
        assert [m for m, _c in window._extract_plan.months] == \
            ["2026-02", "2026-05"]

    def test_the_summary_names_the_months_not_a_range(self, qapp, window,
                                                      tmp_path):
        """Calling this "Feb-26 to May-26" would claim March and April."""
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02", "2026-05"])
        _settle(qapp, window)
        assert "Feb-26 and May-26" in window.extract_summary.text()
        assert "Feb-26 to May-26" not in window.extract_summary.text()

    def test_no_months_ticked_is_refused(self, qapp, window, tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_months_none()
        _settle(qapp, window)
        assert not window.btn_extract.isEnabled()
        assert "No months ticked" in window.extract_summary.text()

    def test_the_presets_tick_and_untick_everything(self, qapp, window,
                                                    tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_months_none()
        assert window._extract_selected_months() == []
        window._extract_months_all()
        assert len(window._extract_selected_months()) == 6

    def test_whole_months_only_drops_the_part_month(self, qapp, window,
                                                    tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-12", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_months_complete()
        _settle(qapp, window)
        assert "2026-06" not in window._extract_selected_months()
        assert len(window._extract_selected_months()) == 5
        assert not self._shown(window.extract_partial_row)

    def test_switching_to_pick_keeps_the_range_it_came_from(self, qapp, window,
                                                            tmp_path):
        """
        Ticking everything instead would silently widen the selection, which is
        the opposite of what reaching for a finer control means.
        """
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        window.extract_from_combo.setCurrentIndex(
            window.extract_from_combo.findData("2026-03"))
        window.extract_to_combo.setCurrentIndex(
            window.extract_to_combo.findData("2026-04"))
        _settle(qapp, window)
        before = window._extract_plan.rows

        self._pick_mode(qapp, window)
        assert window._extract_selected_months() == ["2026-03", "2026-04"]
        assert window._extract_plan.rows == before

    def test_a_hand_made_pick_survives_leaving_the_view(self, qapp, window,
                                                        tmp_path):
        """refresh_* runs on every visit; it must not re-tick everything."""
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02", "2026-05"])
        window.show_view("Raw Data")
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)
        assert window._extract_selected_months() == ["2026-02", "2026-05"]

    def test_a_pick_does_not_carry_to_the_next_sheet(self, qapp, window,
                                                     tmp_path):
        """Those months describe the previous file, which this one may lack."""
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02"])
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path,
                   name="second.csv")
        assert len(window._extract_selected_months()) == 6

    # -- the part-month alert in pick mode ---------------------------------

    def test_a_part_month_is_flagged_when_ticked(self, qapp, window, tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-12", tmp_path)
        self._pick_mode(qapp, window)
        assert self._shown(window.extract_partial_row)
        text = window.extract_partial_warning.text()
        assert "Jun-26" in text and "day 12 of 30" in text
        assert "in this selection" in text, \
            "it was not a range, so calling it one would be wrong"

    def test_unticking_the_part_month_clears_the_alert(self, qapp, window,
                                                       tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-12", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02", "2026-03"])
        _settle(qapp, window)
        assert not self._shown(window.extract_partial_row)

    def test_exclude_unticks_rather_than_narrowing(self, qapp, window, tmp_path):
        self._load(qapp, window, "2026-01-01", "2026-06-12", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_trim_partial_months()
        _settle(qapp, window)
        assert window._extract_selected_months() == [
            "2026-01", "2026-02", "2026-03", "2026-04", "2026-05"]
        assert not self._shown(window.extract_partial_row)

    def test_exclude_can_drop_a_month_in_the_middle(self, qapp, window,
                                                    tmp_path):
        """
        The range mode cannot reach an interior partial month; ticking is not
        restricted to a span, so here it can.
        """
        days = [d for d in pd.date_range("2026-01-01", "2026-04-30")
                if not (d.month == 3 and 5 <= d.day <= 25)]
        frame = pd.DataFrame({
            "number": [f"INC{i:05d}" for i in range(len(days))],
            "opened_at": pd.to_datetime(days),
            "state": ["Closed"] * len(days),
            "short_description": ["Password reset for the account"] * len(days),
        })
        csv = tmp_path / "hole.csv"
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)

        self._pick_mode(qapp, window)
        assert self._shown(window.extract_partial_row)
        assert self._shown(window.btn_extract_trim)
        window._extract_trim_partial_months()
        _settle(qapp, window)
        assert "2026-03" not in window._extract_selected_months()

    def test_exclude_refuses_to_empty_the_selection(self, qapp, window,
                                                    tmp_path):
        """That is not what exclude means, and it would leave no extract."""
        self._load(qapp, window, "2026-06-01", "2026-06-12", tmp_path)
        self._pick_mode(qapp, window)
        assert window._extract_selected_months() == ["2026-06"]
        window._extract_trim_partial_months()
        _settle(qapp, window)
        assert window._extract_selected_months() == ["2026-06"]

    # -- writing -----------------------------------------------------------

    def test_the_written_file_holds_only_the_picked_months(
            self, qapp, window, tmp_path, monkeypatch):
        from openpyxl import load_workbook

        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02", "2026-05"])
        _settle(qapp, window)

        out = tmp_path / "picked.xlsx"
        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k:
                         shown.append((title, text))))
        window.run_extract()
        _settle(qapp, window)

        book = load_workbook(out)
        info = {r[0].value: r[1].value for r in
                book["Extract Info"].iter_rows(min_row=2,
                                               max_row=book["Extract Info"].max_row)}
        assert info["Months included"] == "Feb-26 and May-26"
        assert info["Months, in full"] == "Feb-26, May-26"

        # Loaded from CSV, so the date column arrives as text - read the month
        # off whichever form the cell is in rather than assuming a datetime.
        seen = set()
        for row in book["Extract"].iter_rows(min_row=2, values_only=True):
            for cell in row:
                if hasattr(cell, "year"):
                    seen.add(f"{cell.year}-{cell.month:02d}")
                    break
                if isinstance(cell, str) and cell[:4].isdigit():
                    seen.add(cell[:7])
                    break
        assert seen == {"2026-02", "2026-05"}
        assert "Feb-26 and May-26" in shown[-1][1]

    def test_the_filename_does_not_claim_the_months_between(
            self, qapp, window, tmp_path, monkeypatch):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        window._extract_set_months(["2026-02", "2026-05"])
        _settle(qapp, window)

        captured = {}

        def save(parent, title, name, filt):
            captured["name"] = os.path.basename(name)
            return (str(tmp_path / "out.xlsx"), filt)

        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName", staticmethod(save))
        monkeypatch.setattr(ui.QMessageBox, "information",
                            staticmethod(lambda *a, **k: None))
        window.run_extract()
        _settle(qapp, window)
        assert "2_months_2026-02_to_2026-05" in captured["name"]

    def test_the_mode_is_remembered(self, qapp, window, tmp_path, monkeypatch):
        self._load(qapp, window, "2026-01-01", "2026-06-30", tmp_path)
        self._pick_mode(qapp, window)
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(tmp_path / "m.xlsx"), "")))
        monkeypatch.setattr(ui.QMessageBox, "information",
                            staticmethod(lambda *a, **k: None))
        window.run_extract()
        _settle(qapp, window)

        saved = ConfigManager(window.config_mgr.config_file).config
        assert saved["extract_month_mode"] == ui.EXTRACT_MONTHS_PICK
        assert "extract_months" not in saved, \
            "the months describe this file's span, not the next one's"


class TestExtractPreview:
    """
    Without it the only way to find a wrong column or a mis-read date was to
    wait out a 25-second write and open the file.
    """

    def _show(self, qapp, window):
        window.show_view("Extract")
        _settle(qapp, window)

    def test_the_preview_holds_the_real_first_rows(self, qapp, window):
        self._show(qapp, window)
        model = window.extract_preview_table.model()
        assert model is not None and model.rowCount() > 0

        head = ui.extract_core.build_extract(
            window.df, window._extract_dates(),
            start=window.extract_from_combo.currentData(),
            end=window.extract_to_combo.currentData(),
            columns=window._extract_selected_columns(),
            limit=model.rowCount())
        assert [model.headerData(c, Qt.Orientation.Horizontal)
                for c in range(model.columnCount())] == \
            [str(c) for c in head.columns]
        assert model.data(model.index(0, 0)) == str(head.iloc[0, 0])

    def test_the_title_says_how_many_of_how_many(self, qapp, window):
        self._show(qapp, window)
        title = window.extract_preview_title.text()
        assert "first" in title
        assert f"{window._extract_plan.rows:,} rows" in title

    def test_the_preview_follows_the_column_choice(self, qapp, window):
        self._show(qapp, window)
        window._extract_set_checked({"number", "state"})
        _settle(qapp, window)
        model = window.extract_preview_table.model()
        assert [model.headerData(c, Qt.Orientation.Horizontal)
                for c in range(model.columnCount())] == ["number", "state"]

    def test_an_empty_selection_empties_the_preview(self, qapp, window):
        self._show(qapp, window)
        window._extract_select_none()
        _settle(qapp, window)
        model = window.extract_preview_table.model()
        assert model is None or model.rowCount() == 0

    def test_the_preview_is_copyable_like_any_table(self, qapp, window):
        self._show(qapp, window)
        assert "Copy" in [a.text() for a in
                          window.extract_preview_table.actions()]


class TestExtractCancel:
    """
    A .xlsx write is tens of seconds at scale. Cancelling must not leave a
    part-written file where the user asked for their extract - a short workbook
    that opens is worse than no workbook at all.
    """

    def test_cancelling_deletes_the_part_file(self, qapp, window, tmp_path,
                                              monkeypatch):
        out = tmp_path / "cancelled.xlsx"
        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k:
                         shown.append((title, text))))
        errors = []
        monkeypatch.setattr(window, "show_error", errors.append)

        # Cancel on the first progress message, whenever that lands.
        monkeypatch.setattr(ui.extract_core, "DETAIL_PROGRESS_EVERY", 1)
        real = ui.extract_core.write_extract

        def cancelling(path, frame, **kwargs):
            progress = kwargs.get("progress")

            def relay(message):
                if progress:
                    progress(message)
                window._cancel_extract()

            kwargs["progress"] = relay
            return real(path, frame, **kwargs)

        monkeypatch.setattr(ui.extract_core, "write_extract", cancelling)

        window.show_view("Extract")
        _settle(qapp, window)
        window.run_extract()
        _settle(qapp, window)

        assert not out.exists(), "a part-written extract must not be left behind"
        assert shown and shown[-1][0] == "Extract cancelled"
        assert not errors, "the user asked for this; it is not an error"

    def test_the_cancel_button_only_shows_while_writing(self, qapp, window):
        window.show_view("Extract")
        _settle(qapp, window)
        button = window.btn_extract_cancel
        assert not button.isVisibleTo(button.parentWidget())
        window._begin_extract()
        assert button.isVisibleTo(button.parentWidget())
        window._finish_extract()
        assert not button.isVisibleTo(button.parentWidget())

    def test_core_raises_a_distinct_error_so_it_is_not_a_success(self):
        """A cancelled write must not be mistaken for a finished one."""
        assert issubclass(ui.extract_core.ExtractCancelled, Exception)
        assert not issubclass(ui.extract_core.ExtractCancelled,
                              ui.extract_core.ExtractTooLarge)


class TestExtractReveal:
    def test_the_button_appears_only_after_a_write(self, qapp, window, tmp_path,
                                                   monkeypatch):
        window.show_view("Extract")
        _settle(qapp, window)
        button = window.btn_extract_reveal
        assert not button.isVisibleTo(button.parentWidget())

        out = tmp_path / "done.xlsx"
        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: (str(out), "")))
        monkeypatch.setattr(ui.QMessageBox, "information",
                            staticmethod(lambda *a, **k: None))
        window.run_extract()
        _settle(qapp, window)
        assert button.isVisibleTo(button.parentWidget())

    def test_it_opens_the_folder_the_file_went_to(self, qapp, window, tmp_path,
                                                 monkeypatch):
        opened = []
        monkeypatch.setattr(ui.QDesktopServices, "openUrl",
                            staticmethod(lambda url: opened.append(url.toLocalFile())))
        window._extract_last_path = str(tmp_path / "somewhere.xlsx")
        window.show_view("Extract")
        _settle(qapp, window)
        window._reveal_extract()
        assert opened and os.path.normpath(opened[0]) == \
            os.path.normpath(str(tmp_path))

    def test_a_missing_folder_is_reported_not_opened(self, qapp, window,
                                                    tmp_path, monkeypatch):
        opened = []
        monkeypatch.setattr(ui.QDesktopServices, "openUrl",
                            staticmethod(lambda url: opened.append(url)))
        errors = []
        monkeypatch.setattr(window, "show_error", errors.append)
        window._extract_last_path = str(tmp_path / "gone" / "x.xlsx")
        window._reveal_extract()
        assert not opened
        assert errors


class TestVersionIntegrity:
    """
    The About dialog verifies a SHA-256 of author+name+version. Bumping the
    version without regenerating that hash makes the dialog refuse to open,
    which is the kind of thing nobody notices until a user asks what version
    they have.
    """

    def test_the_about_dialog_passes_its_own_check(self, qapp, window,
                                                   monkeypatch):
        refused = []
        monkeypatch.setattr(
            ui.QMessageBox, "critical",
            staticmethod(lambda parent, title, text, *a, **k:
                         refused.append((title, text))))
        captured = {}
        monkeypatch.setattr(ui.QDialog, "exec", lambda self: captured.setdefault(
            "labels", [w.text() for w in self.findChildren(ui.QLabel)]))
        window.show_about_dialog()
        assert not refused, f"integrity check refused: {refused}"

    def test_the_dialog_states_the_version(self, qapp, window, monkeypatch):
        captured = {}
        monkeypatch.setattr(ui.QDialog, "exec", lambda self: captured.setdefault(
            "text", " ".join(w.text() for w in self.findChildren(ui.QLabel))))
        monkeypatch.setattr(ui.QMessageBox, "critical",
                            staticmethod(lambda *a, **k: None))
        window.show_about_dialog()
        assert "2.2.1" in captured.get("text", "")


class TestExtractExcludeCancelled:
    """
    Rows that leave the extract for a reason the file itself cannot show have to
    be named on screen and recorded in the output - the rule the undated rows
    and the part-months already follow.
    """

    def _load(self, qapp, window, states, tmp_path, name="states.csv",
              undated=0):
        stamps = pd.date_range("2026-01-01", periods=len(states), freq="D")
        frame = pd.DataFrame({
            "number": [f"INC{i:05d}" for i in range(len(states))],
            "state": states,
            "opened_at": stamps,
            "short_description": ["Password reset for the account"] * len(states),
        })
        if undated:
            frame.loc[frame.index[:undated], "opened_at"] = pd.NaT
        csv = tmp_path / name
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)
        return frame

    def _box(self, window):
        return window.extract_cancelled_check

    def _offered(self, window):
        box = self._box(window)
        return box.isVisibleTo(box.parentWidget())

    # -- when it is offered -------------------------------------------------

    def test_it_is_offered_when_there_are_cancellations(self, qapp, window,
                                                        tmp_path):
        self._load(qapp, window, ["Closed", "Cancelled", "New"] * 4, tmp_path)
        assert self._offered(window)
        assert not self._box(window).isChecked(), "opt in, not out"

    def test_it_is_hidden_when_nothing_is_cancelled(self, qapp, window,
                                                    tmp_path):
        """A box that removes nothing implies rows are being dropped."""
        self._load(qapp, window, ["Closed", "New", "Resolved"] * 4, tmp_path)
        assert not self._offered(window)

    def test_it_is_hidden_when_the_file_has_no_state_column(self, qapp, window,
                                                            tmp_path):
        frame = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(6)],
            "opened_at": pd.date_range("2026-01-01", periods=6),
            "short_description": ["A long enough description here"] * 6,
        })
        csv = tmp_path / "nostate.csv"
        frame.to_csv(csv, index=False)
        window.filepath = str(csv)
        window.load_sheet(None)
        _settle(qapp, window)
        window.show_view("Extract")
        _settle(qapp, window)
        assert not self._offered(window)

    # -- what it says ------------------------------------------------------

    def test_the_label_names_the_values_it_matched(self, qapp, window, tmp_path):
        """
        "Exclude cancelled" is a claim about someone else's vocabulary; the only
        way to check it is to see which values were caught.
        """
        self._load(qapp, window,
                   ["Cancelled", "canceled", "Closed Cancelled", "Closed"] * 3,
                   tmp_path)
        label = self._box(window).text()
        assert "Cancelled" in label and "canceled" in label
        assert "in this file" in label, "the count's scope must be stated"
        tip = self._box(window).toolTip()
        assert "cancel" in tip
        assert "withdrawn" in tip.lower(), \
            "the vocabulary it does not cover has to be visible"

    def test_the_summary_says_how_many_left_this_selection(self, qapp, window,
                                                           tmp_path):
        self._load(qapp, window, ["Closed", "Cancelled"] * 8, tmp_path)
        before = window._extract_plan.rows
        self._box(window).setChecked(True)
        _settle(qapp, window)

        plan = window._extract_plan
        assert plan.rows == before - plan.excluded_rows
        assert f"{plan.excluded_rows:,} cancelled left out" in \
            window.extract_summary.text()

    # -- what it does ------------------------------------------------------

    def test_ticking_it_drops_the_rows(self, qapp, window, tmp_path):
        frame = self._load(qapp, window, ["Closed", "Cancelled", "New"] * 5,
                           tmp_path)
        self._box(window).setChecked(True)
        _settle(qapp, window)
        expected = int((frame["state"] != "Cancelled").sum())
        assert window._extract_plan.rows == expected

    def test_the_preview_stops_showing_them(self, qapp, window, tmp_path):
        self._load(qapp, window, ["Cancelled", "Closed"] * 8, tmp_path)
        self._box(window).setChecked(True)
        _settle(qapp, window)
        model = window.extract_preview_table.model()
        seen = {model.data(model.index(r, 1)) for r in range(model.rowCount())}
        assert "Cancelled" not in seen

    def test_untitcking_puts_them_back(self, qapp, window, tmp_path):
        self._load(qapp, window, ["Closed", "Cancelled"] * 8, tmp_path)
        full = window._extract_plan.rows
        self._box(window).setChecked(True)
        _settle(qapp, window)
        self._box(window).setChecked(False)
        _settle(qapp, window)
        assert window._extract_plan.rows == full
        assert window._extract_plan.excluded_rows == 0

    def test_the_undated_count_is_what_ticking_would_add(self, qapp, window,
                                                         tmp_path):
        """
        Some undated rows are cancelled, so with the exclusion on they would not
        be added. The number beside the box must mean the same thing either way.
        """
        self._load(qapp, window, ["Cancelled", "Closed"] * 8, tmp_path,
                   undated=4)
        self._box(window).setChecked(True)
        _settle(qapp, window)
        text = window.extract_undated_check.text()
        assert f"{window._extract_plan.undated_rows:,} rows with no date" in text

    # -- disclosure --------------------------------------------------------

    def test_the_dialog_and_the_file_both_record_it(self, qapp, window,
                                                    tmp_path, monkeypatch):
        from openpyxl import load_workbook

        self._load(qapp, window, ["Closed", "Cancelled", "canceled"] * 4,
                   tmp_path)
        self._box(window).setChecked(True)
        _settle(qapp, window)
        excluded = window._extract_plan.excluded_rows

        out = tmp_path / "pruned.xlsx"
        monkeypatch.setattr(ui.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: (str(out), "")))
        shown = []
        monkeypatch.setattr(
            ui.QMessageBox, "information",
            staticmethod(lambda parent, title, text, *a, **k:
                         shown.append((title, text))))
        window.run_extract()
        _settle(qapp, window)

        assert f"{excluded:,} cancelled tickets were left out" in shown[-1][1]
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert "Cancelled" in str(rows["States excluded"])
        assert rows["Rows excluded by state"] == excluded

    def test_the_choice_is_remembered(self, qapp, window, tmp_path,
                                     monkeypatch):
        self._load(qapp, window, ["Closed", "Cancelled"] * 6, tmp_path)
        self._box(window).setChecked(True)
        _settle(qapp, window)
        monkeypatch.setattr(
            ui.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(tmp_path / "r.xlsx"), "")))
        monkeypatch.setattr(ui.QMessageBox, "information",
                            staticmethod(lambda *a, **k: None))
        window.run_extract()
        _settle(qapp, window)

        saved = ConfigManager(window.config_mgr.config_file).config
        assert saved["extract_exclude_cancelled"] is True

    def test_a_file_without_cancellations_does_not_inherit_the_setting(
            self, qapp, window, tmp_path):
        """
        The saved preference must not leave a hidden box ticked on a file it
        cannot apply to - the plan would claim an exclusion that never happened.
        """
        window.config["extract_exclude_cancelled"] = True
        self._load(qapp, window, ["Closed", "New"] * 6, tmp_path)
        assert not self._offered(window)
        assert not self._box(window).isChecked()
        assert window._extract_plan.excluded_rows == 0
