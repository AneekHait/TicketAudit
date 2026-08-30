"""
Tests for core/extract.py — pulling a workable subset out of a large export.

Two failure modes drive most of these. The first is the closing-month
off-by-one: writing the upper bound as `<= start_of(end)` keeps only the 1st of
the final month, which looks plausible and quietly loses four weeks of data. The
second is losing rows whose date does not parse — they satisfy neither the range
nor its complement, so a plain mask drops them with no trace.
"""
import codecs

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.analyzer import SanityAnalyzer
from core.extract import (
    EXCEL_SLOW_CELLS, FORMAT_CSV, FORMAT_XLSX, MIN_INTERIOR_GAP_DAYS,
    MIN_PARTIAL_EDGE_DAYS, MONTH_ABBR, ExtractPlan, ExtractTooLarge,
    build_extract, info_sidecar_path, month_bounds, month_counts,
    is_contiguous, month_coverage, month_label, months_label, partial_months,
    plan_extract, range_label, suggest_filename, write_extract,
)


def _frame(dates, extra_columns=0):
    """A frame whose opened_at is exactly the dates given."""
    n = len(dates)
    data = {
        "number": [f"INC{i:05d}" for i in range(n)],
        "state": ["Closed"] * n,
        "priority": ["P1", "P2"] * (n // 2) + ["P1"] * (n % 2),
        "opened_at": pd.to_datetime(dates),
        "resolved_at": pd.to_datetime(dates),
        "short_description": ["Login issue for the user account"] * n,
    }
    for i in range(extra_columns):
        data[f"custom_{i}"] = [f"v{j % 5}" for j in range(n)]
    return pd.DataFrame(data)


@pytest.fixture
def spanning_frame():
    """Three years of monthly data, one row per month, plus two undated rows."""
    months = pd.period_range("2024-01", "2026-12", freq="M")
    dates = [p.start_time + pd.Timedelta(days=5) for p in months]
    frame = _frame(dates, extra_columns=4)
    undated = frame.head(2).copy()
    undated["opened_at"] = pd.NaT
    return pd.concat([frame, undated], ignore_index=True)


def _dates(frame):
    return frame["opened_at"]


class TestMonthBounds:
    def test_the_closing_month_is_whole(self):
        """`<= start_of(end)` would keep only the 1st of the final month."""
        lower, upper = month_bounds("2025-02", "2026-07")
        assert lower == pd.Timestamp("2025-02-01")
        assert upper == pd.Timestamp("2026-08-01")

    def test_a_single_month_range(self):
        lower, upper = month_bounds("2025-03", "2025-03")
        assert lower == pd.Timestamp("2025-03-01")
        assert upper == pd.Timestamp("2025-04-01")

    def test_a_reversed_range_is_normalised(self):
        """Picking From after To is a slip, not a request for no rows."""
        assert month_bounds("2026-07", "2025-02") == month_bounds("2025-02",
                                                                  "2026-07")

    def test_it_spans_a_year_boundary(self):
        lower, upper = month_bounds("2024-12", "2025-01")
        assert lower == pd.Timestamp("2024-12-01")
        assert upper == pd.Timestamp("2025-02-01")


class TestMonthCounts:
    def test_counts_ascend_by_month(self, spanning_frame):
        counts = month_counts(_dates(spanning_frame))
        assert counts[0][0] == "2024-01"
        assert counts[-1][0] == "2026-12"
        assert [m for m, _ in counts] == sorted(m for m, _ in counts)

    def test_undated_rows_are_not_a_month(self, spanning_frame):
        counts = month_counts(_dates(spanning_frame))
        assert sum(c for _, c in counts) == len(spanning_frame) - 2

    def test_no_date_column(self):
        assert month_counts(None) == []

    def test_all_dates_unparseable(self):
        assert month_counts(pd.Series([pd.NaT, pd.NaT])) == []


class TestPlanExtract:
    def test_the_closing_month_is_included(self, spanning_frame):
        """The bug this feature is most likely to ship."""
        plan = plan_extract(spanning_frame, _dates(spanning_frame),
                            start="2025-02", end="2026-07",
                            columns=["number", "opened_at"])
        assert [m for m, _ in plan.months][-1] == "2026-07"
        assert plan.rows == 18, "Feb-2025..Jul-2026 inclusive is 18 months"

    def test_it_agrees_with_build_extract(self, spanning_frame):
        kwargs = dict(start="2025-02", end="2026-07",
                      columns=["number", "state", "opened_at"])
        plan = plan_extract(spanning_frame, _dates(spanning_frame), **kwargs)
        frame = build_extract(spanning_frame, _dates(spanning_frame), **kwargs)
        assert plan.rows == len(frame)
        assert plan.columns == len(frame.columns)

    def test_undated_rows_are_counted_but_excluded_by_default(self, spanning_frame):
        plan = plan_extract(spanning_frame, _dates(spanning_frame),
                            start="2024-01", end="2026-12",
                            columns=["number"])
        assert plan.undated_rows == 2
        assert plan.undated_included is False
        assert plan.rows == len(spanning_frame) - 2

    def test_undated_rows_can_be_included(self, spanning_frame):
        plan = plan_extract(spanning_frame, _dates(spanning_frame),
                            start="2024-01", end="2026-12",
                            columns=["number"], include_undated=True)
        assert plan.rows == len(spanning_frame)

    def test_cells_is_rows_times_columns(self, spanning_frame):
        plan = plan_extract(spanning_frame, _dates(spanning_frame),
                            start="2025-01", end="2025-12",
                            columns=["number", "state", "priority"])
        assert plan.cells == plan.rows * 3

    def test_a_range_outside_the_data_yields_nothing(self, spanning_frame):
        plan = plan_extract(spanning_frame, _dates(spanning_frame),
                            start="2030-01", end="2030-06",
                            columns=["number"])
        assert plan.rows == 0
        assert plan.months == []

    def test_no_date_column_leaves_every_row(self, spanning_frame):
        plan = plan_extract(spanning_frame, None, start="2025-01", end="2025-02",
                            columns=["number", "state"])
        assert plan.rows == len(spanning_frame)
        assert plan.months == []

    def test_the_slow_flag_tracks_the_cell_count(self):
        small = ExtractPlan(10, 5, 0, False, [])
        big = ExtractPlan(EXCEL_SLOW_CELLS, 2, 0, False, [])
        assert small.may_be_slow_in_excel is False
        assert big.may_be_slow_in_excel is True

    def test_the_row_ceiling_flag(self):
        assert ExtractPlan(1_000, 5, 0, False, []).exceeds_excel_rows is False
        assert ExtractPlan(2_000_000, 5, 0, False, []).exceeds_excel_rows is True


class TestBuildExtract:
    def test_columns_come_back_in_source_order(self, spanning_frame):
        """Tick order is not a meaningful order."""
        frame = build_extract(spanning_frame, _dates(spanning_frame),
                              start="2025-01", end="2025-06",
                              columns=["short_description", "number", "state"])
        assert list(frame.columns) == ["number", "state", "short_description"]

    def test_an_unknown_column_is_ignored(self, spanning_frame):
        frame = build_extract(spanning_frame, _dates(spanning_frame),
                              start="2025-01", end="2025-06",
                              columns=["number", "not_a_column"])
        assert list(frame.columns) == ["number"]

    def test_an_empty_selection_is_refused(self, spanning_frame):
        with pytest.raises(ValueError, match="no columns"):
            build_extract(spanning_frame, _dates(spanning_frame),
                          start="2025-01", end="2025-06", columns=[])

    def test_dtypes_survive_the_slice(self, spanning_frame):
        """Categories are kept; converting to str would double the memory."""
        frame = spanning_frame.copy()
        frame["state"] = frame["state"].astype("category")
        out = build_extract(frame, _dates(frame), start="2025-01", end="2025-06",
                            columns=["number", "state"])
        assert isinstance(out["state"].dtype, pd.CategoricalDtype)

    def test_the_index_is_the_source_index(self, spanning_frame):
        """So a row can still be traced back to the original file."""
        out = build_extract(spanning_frame, _dates(spanning_frame),
                            start="2025-02", end="2025-02", columns=["number"])
        assert out.index.isin(spanning_frame.index).all()


class TestWriteExtract:
    def _subset(self, spanning_frame, **kwargs):
        options = dict(start="2025-01", end="2025-06",
                       columns=["number", "state", "opened_at"])
        options.update(kwargs)
        return build_extract(spanning_frame, _dates(spanning_frame), **options)

    def test_xlsx_round_trips(self, spanning_frame, tmp_path):
        frame = self._subset(spanning_frame)
        out = tmp_path / "x.xlsx"
        written = write_extract(str(out), frame, fmt=FORMAT_XLSX,
                                date_column="opened_at", start="2025-01",
                                end="2025-06")
        assert written == len(frame)
        sheet = load_workbook(out)["Extract"]
        assert sheet.max_row == len(frame) + 1
        assert [c.value for c in sheet[1]] == list(frame.columns)

    def test_the_header_is_frozen(self, spanning_frame, tmp_path):
        """In write-only mode this only works if set before the first append."""
        out = tmp_path / "x.xlsx"
        write_extract(str(out), self._subset(spanning_frame))
        assert load_workbook(out)["Extract"].freeze_panes == "A2"

    def test_column_widths_are_set(self, spanning_frame, tmp_path):
        out = tmp_path / "x.xlsx"
        write_extract(str(out), self._subset(spanning_frame))
        sheet = load_workbook(out)["Extract"]
        assert sheet.column_dimensions["A"].width

    def test_it_carries_no_table_or_formatting(self, spanning_frame, tmp_path):
        """Lean on purpose: Tables and conditional formats cost Excel time."""
        out = tmp_path / "x.xlsx"
        write_extract(str(out), self._subset(spanning_frame))
        sheet = load_workbook(out)["Extract"]
        assert len(sheet.tables) == 0
        assert len(sheet.conditional_formatting._cf_rules) == 0

    def test_provenance_is_recorded_on_its_own_sheet(self, spanning_frame,
                                                    tmp_path):
        """
        "September's tickets" means different rows depending on whether September
        came from the opened or the resolved date.
        """
        out = tmp_path / "x.xlsx"
        write_extract(str(out), self._subset(spanning_frame),
                      date_column="opened_at", start="2025-01", end="2025-06",
                      source_path="ServiceNow_Q1.xlsx")
        info = load_workbook(out)["Extract Info"]
        text = "\n".join(str(c.value) for row in info.iter_rows()
                         for c in row if c.value is not None)
        assert "opened_at" in text
        assert "2025-01 to 2025-06" in text
        assert "ServiceNow_Q1.xlsx" in text

    def test_provenance_states_whether_undated_rows_are_in(self, spanning_frame,
                                                           tmp_path):
        frame = self._subset(spanning_frame, start="2024-01", end="2026-12",
                             include_undated=True)
        plan = plan_extract(spanning_frame, _dates(spanning_frame),
                            start="2024-01", end="2026-12",
                            columns=["number", "state", "opened_at"],
                            include_undated=True)
        out = tmp_path / "x.xlsx"
        write_extract(str(out), frame, plan=plan)
        info = load_workbook(out)["Extract Info"]
        rows = {str(r[0].value): str(r[1].value) for r in info.iter_rows()}
        assert rows["Undated rows included"] == "Yes"
        assert rows["Rows with no date"] == "2"

    def test_csv_round_trips_with_a_bom(self, spanning_frame, tmp_path):
        frame = self._subset(spanning_frame)
        out = tmp_path / "x.csv"
        written = write_extract(str(out), frame, fmt=FORMAT_CSV)
        assert written == len(frame)
        with open(out, "rb") as handle:
            assert handle.read(3) == codecs.BOM_UTF8
        assert len(pd.read_csv(out)) == len(frame)

    def test_an_empty_extract_is_refused(self, spanning_frame, tmp_path):
        """Better than handing over a file with only a header row."""
        frame = self._subset(spanning_frame, start="2030-01", end="2030-02")
        with pytest.raises(ValueError, match="empty"):
            write_extract(str(tmp_path / "x.xlsx"), frame)

    def test_too_many_rows_for_xlsx_is_refused_not_truncated(self, tmp_path,
                                                             monkeypatch):
        """
        The user chose the range, so silently dropping rows off the end of their
        own selection is worse than saying it will not fit.
        """
        import core.extract as module
        monkeypatch.setattr(module, "DETAIL_MAX_DATA_ROWS", 3)
        frame = _frame(pd.date_range("2025-01-01", periods=10))
        with pytest.raises(ExtractTooLarge, match="CSV"):
            write_extract(str(tmp_path / "x.xlsx"), frame, fmt=FORMAT_XLSX)

    def test_csv_has_no_row_ceiling(self, tmp_path, monkeypatch):
        import core.extract as module
        monkeypatch.setattr(module, "DETAIL_MAX_DATA_ROWS", 3)
        frame = _frame(pd.date_range("2025-01-01", periods=10))
        assert write_extract(str(tmp_path / "x.csv"), frame,
                             fmt=FORMAT_CSV) == 10

    def test_progress_is_reported(self, tmp_path, monkeypatch):
        import core.extract as module
        monkeypatch.setattr(module, "DETAIL_PROGRESS_EVERY", 5)
        seen = []
        frame = _frame(pd.date_range("2025-01-01", periods=20))
        write_extract(str(tmp_path / "x.xlsx"), frame, progress=seen.append)
        assert any("Writing extract" in message for message in seen)
        assert any("Finalising" in message for message in seen)


class TestWriterTraps:
    """
    Each reproduced against openpyxl during the report work; most fail silently.
    pd.NA is the worst: it kills the row generator, and save() then emits a
    truncated workbook that openpyxl itself cannot reopen.
    """

    def _write(self, frame, tmp_path, fmt=FORMAT_XLSX):
        out = tmp_path / f"x.{fmt}"
        rows = write_extract(str(out), frame, fmt=fmt)
        return rows, out

    def test_pandas_na_does_not_truncate_the_file(self, tmp_path):
        frame = pd.DataFrame({
            "text": pd.array(["a", None, "c"], dtype="string"),
            "num": pd.array([1, None, 3], dtype="Int64"),
        })
        rows, out = self._write(frame, tmp_path)
        assert rows == 3
        assert load_workbook(out)["Extract"].max_row == 4

    def test_control_characters_are_stripped(self, tmp_path):
        frame = pd.DataFrame({"text": ["clean", "bad\x00char\x07here"]})
        _rows, out = self._write(frame, tmp_path)
        assert load_workbook(out)["Extract"]["A3"].value == "badcharhere"

    def test_a_leading_equals_stays_text(self, tmp_path):
        """A write-only cell cannot be re-typed after assignment."""
        frame = pd.DataFrame({"text": ["=SUM(A1:A9)", "plain"]})
        _rows, out = self._write(frame, tmp_path)
        cell = load_workbook(out)["Extract"]["A2"]
        assert cell.data_type == "s"
        assert cell.value == "=SUM(A1:A9)"

    def test_over_long_strings_are_truncated_for_xlsx(self, tmp_path):
        from core.excel_pivot import EXCEL_MAX_CELL_CHARS
        frame = pd.DataFrame({"text": ["x" * (EXCEL_MAX_CELL_CHARS + 500)]})
        _rows, out = self._write(frame, tmp_path)
        assert len(load_workbook(out)["Extract"]["A2"].value) == \
            EXCEL_MAX_CELL_CHARS

    def test_csv_keeps_long_text_whole(self, tmp_path):
        """CSV has no cell limit, so truncating would be loss for nothing."""
        from core.excel_pivot import EXCEL_MAX_CELL_CHARS
        long_value = "x" * (EXCEL_MAX_CELL_CHARS + 500)
        frame = pd.DataFrame({"text": [long_value]})
        _rows, out = self._write(frame, tmp_path, fmt=FORMAT_CSV)
        assert len(pd.read_csv(out).loc[0, "text"]) == len(long_value)

    def test_booleans_read_as_words(self, tmp_path):
        frame = pd.DataFrame({"made_sla": [True, False]})
        _rows, out = self._write(frame, tmp_path)
        sheet = load_workbook(out)["Extract"]
        assert sheet["A2"].value == "Yes"
        assert sheet["A3"].value == "No"

    def test_timezone_aware_dates_keep_wall_clock_time(self, tmp_path):
        frame = pd.DataFrame({
            "when": pd.to_datetime(["2024-01-01T08:30:00+05:30"])})
        _rows, out = self._write(frame, tmp_path)
        value = load_workbook(out)["Extract"]["A2"].value
        assert value.hour == 8
        assert value.tzinfo is None


class TestSuggestFilename:
    def test_it_records_the_range_and_the_date_column(self):
        name = suggest_filename("C:/exports/ServiceNow Q3.xlsx", "opened_at",
                                "2025-02", "2026-07", "xlsx")
        assert name == "ServiceNow Q3_opened_at_2025-02_to_2026-07.xlsx"

    def test_an_awkward_column_name_is_made_safe(self):
        name = suggest_filename("t.xlsx", "sys created/on", "2025-01",
                                "2025-01", "csv")
        assert "/" not in name
        assert name.endswith(".csv")

    def test_it_works_without_a_source_path(self):
        assert suggest_filename(None, "opened_at", "2025-01", "2025-02",
                                "xlsx").startswith("extract_")


class TestMonthLabel:
    """
    The people using this name their ranges "Feb25-Jul26", so the picker says
    Feb-25. The ISO code stays the identifier everywhere a machine reads it.
    """

    def test_every_month_abbreviates(self):
        got = [month_label(f"2025-{m:02d}") for m in range(1, 13)]
        assert got == [f"{name}-25" for name in MONTH_ABBR]

    def test_two_digit_year_comes_from_the_period(self):
        assert month_label("2026-07") == "Jul-26"
        assert month_label("1999-12") == "Dec-99"

    def test_names_do_not_come_from_the_locale(self):
        """
        strftime("%b") follows the C locale, so the same month would render
        differently on a non-English Windows install and one label would mean
        two things across two machines.
        """
        import inspect
        from core import extract
        source = inspect.getsource(extract.month_label)
        assert "%b" not in source and "strftime" not in source

    def test_malformed_input_is_returned_not_guessed(self):
        for value in ("", "nonsense", "2025", None):
            assert month_label(value) == value

    def test_range_label_reads_as_a_range(self):
        assert range_label("2025-02", "2026-07") == "Feb-25 to Jul-26"
        assert range_label("2025-02", "2025-02") == "Feb-25"
        assert range_label("", "") == "all dates"


class TestMonthCoverage:
    """
    An export taken mid-month makes that month look like a collapse in volume.
    Coverage judges the *span* of days present, not how many days have rows, so
    a weekday-only queue is not accused of being incomplete.
    """

    def _cov(self, dates):
        series = pd.Series(pd.to_datetime(dates))
        return {c.period: c for c in month_coverage(series)}

    def test_a_whole_month_is_complete(self):
        cov = self._cov(pd.date_range("2025-03-01", "2025-03-31"))
        assert not cov["2025-03"].partial

    def test_weekdays_only_is_complete(self):
        """The false alarm a distinct-day count would raise on real data."""
        cov = self._cov(pd.bdate_range("2025-01-01", "2025-03-31"))
        assert [c.partial for c in cov.values()] == [False, False, False]

    def test_an_export_taken_mid_month_truncates_the_last_month(self):
        cov = self._cov(pd.date_range("2026-06-01", "2026-08-12"))
        assert not cov["2026-06"].partial
        assert not cov["2026-07"].partial
        august = cov["2026-08"]
        assert august.partial and august.truncated_end
        assert august.last_day == 12 and august.days_in_month == 31
        assert august.missing_end == 19
        assert "stops on day 12 of 31" in august.describe()

    def test_data_starting_mid_month_truncates_the_first_month(self):
        march = self._cov(pd.date_range("2025-03-18", "2025-04-30"))["2025-03"]
        assert march.partial and march.truncated_start
        assert march.first_day == 18 and march.missing_start == 17
        assert "starts on day 18" in march.describe()

    def test_a_quiet_day_at_the_edge_is_not_a_partial_month(self):
        cov = self._cov(pd.date_range("2025-01-02", "2025-01-30"))
        assert not cov["2025-01"].partial

    def test_the_edge_threshold_is_where_it_says(self):
        inside = self._cov(pd.date_range(
            f"2025-01-{MIN_PARTIAL_EDGE_DAYS:02d}", "2025-01-31"))["2025-01"]
        assert inside.missing_start == MIN_PARTIAL_EDGE_DAYS - 1
        assert not inside.partial
        beyond = self._cov(pd.date_range(
            f"2025-01-{MIN_PARTIAL_EDGE_DAYS + 1:02d}", "2025-01-31"))["2025-01"]
        assert beyond.missing_start == MIN_PARTIAL_EDGE_DAYS
        assert beyond.partial

    def test_a_week_long_hole_in_the_middle_is_reported(self):
        days = [d for d in pd.date_range("2025-02-01", "2025-02-28")
                if not (8 <= d.day <= 19)]
        february = self._cov(days)["2025-02"]
        assert february.partial and february.has_interior_gap
        assert february.longest_gap == 12
        assert not february.truncated_start and not february.truncated_end
        assert "12-day stretch" in february.describe()

    def test_a_weekend_shaped_gap_is_not_a_hole(self):
        gap = MIN_INTERIOR_GAP_DAYS - 1
        days = [d for d in pd.date_range("2025-04-01", "2025-04-30")
                if not (10 <= d.day < 10 + gap)]
        april = self._cov(days)["2025-04"]
        assert april.longest_gap == gap
        assert not april.partial

    def test_a_single_day_of_data_is_partial(self):
        may = self._cov(["2025-05-14"] * 5)["2025-05"]
        assert may.partial
        assert may.first_day == may.last_day == 14
        assert "day 14 to day 14" in may.describe()

    def test_rows_are_counted_per_month(self):
        cov = self._cov(["2025-01-05"] * 3 + ["2025-02-06"] * 7)
        assert cov["2025-01"].rows == 3
        assert cov["2025-02"].rows == 7

    def test_february_length_follows_the_year(self):
        leap = self._cov(pd.date_range("2024-02-01", "2024-02-29"))["2024-02"]
        assert leap.days_in_month == 29
        assert not leap.partial

    def test_no_dates_is_not_an_error(self):
        assert month_coverage(None) == []
        assert month_coverage(pd.Series([], dtype="datetime64[ns]")) == []
        assert month_coverage(pd.Series([pd.NaT, pd.NaT])) == []

    def test_undated_rows_are_ignored_not_treated_as_a_gap(self):
        dates = list(pd.date_range("2025-06-01", "2025-06-30")) + [pd.NaT] * 5
        june = self._cov(dates)["2025-06"]
        assert june.rows == 30
        assert not june.partial


class TestPartialMonths:
    def _coverage(self):
        span = pd.date_range("2025-01-19", "2026-08-12")
        return month_coverage(pd.Series(span))

    def test_both_truncated_ends_are_found(self):
        found = [c.period for c in partial_months(self._coverage())]
        assert found == ["2025-01", "2026-08"]

    def test_a_range_inside_the_data_is_clean(self):
        assert partial_months(self._coverage(), "2025-02", "2026-07") == []

    def test_the_range_bounds_are_inclusive(self):
        found = [c.period for c in partial_months(self._coverage(),
                                                  "2026-08", "2026-08")]
        assert found == ["2026-08"]


class TestIncompleteMonthDisclosure:
    """
    A caveat that stays on screen is lost the moment the file is shared, so the
    partial months travel with the extract - the same rule as the undated rows.
    """

    def _subset(self, frame):
        return build_extract(frame, _dates(frame), start="2025-01",
                             end="2025-02", columns=["number", "opened_at"])

    def test_the_workbook_records_them(self, spanning_frame, tmp_path):
        subset = self._subset(spanning_frame)
        out = tmp_path / "x.xlsx"
        write_extract(str(out), subset, fmt=FORMAT_XLSX, date_column="opened_at",
                      start="2025-01", end="2025-02", incomplete=["2025-01"])
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["Incomplete months included"] == "Jan-25"
        assert "Jan-25 to Feb-25" in str(rows["Months included"])
        assert "2025-01 to 2025-02" in str(rows["Months included"])

    def test_none_is_stated_rather_than_left_blank(self, spanning_frame, tmp_path):
        subset = self._subset(spanning_frame)
        out = tmp_path / "clean.xlsx"
        write_extract(str(out), subset, fmt=FORMAT_XLSX, date_column="opened_at",
                      start="2025-01", end="2025-02")
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["Incomplete months included"] == "none"

    def test_the_csv_gets_a_sidecar_rather_than_a_broken_header(
            self, spanning_frame, tmp_path):
        """
        Provenance inside the CSV would break read_csv and any re-import, which
        is why the .xlsx puts it on a second sheet. The promise that the date
        column travels with the extract still has to be kept.
        """
        subset = self._subset(spanning_frame)
        out = tmp_path / "x.csv"
        write_extract(str(out), subset, fmt=FORMAT_CSV, date_column="opened_at",
                      start="2025-01", end="2025-02", incomplete=["2025-01"],
                      source_path="Book1.xlsx")

        reread = pd.read_csv(out)
        assert len(reread) == len(subset)
        assert list(reread.columns) == ["number", "opened_at"]

        sidecar = tmp_path / "x_ExtractInfo.txt"
        assert sidecar.exists()
        assert str(sidecar) == info_sidecar_path(str(out))
        text = sidecar.read_text(encoding="utf-8-sig")
        assert "Date column used: opened_at" in text
        assert "Incomplete months included: Jan-25" in text
        assert "Source file: Book1.xlsx" in text

    def test_a_failed_sidecar_does_not_lose_the_extract(
            self, spanning_frame, tmp_path, monkeypatch):
        """The data is written and valid; the note beside it is not worth it."""
        subset = self._subset(spanning_frame)
        out = tmp_path / "y.csv"
        real_open = open

        def refuse(path, *args, **kwargs):
            if str(path).endswith("_ExtractInfo.txt"):
                raise OSError("read-only folder")
            return real_open(path, *args, **kwargs)

        monkeypatch.setattr("builtins.open", refuse)
        written = write_extract(str(out), subset, fmt=FORMAT_CSV,
                                date_column="opened_at", start="2025-01",
                                end="2025-02")
        assert written == len(subset)
        assert out.exists()


class TestPickedMonths:
    """
    "February, May and July" is a real request and is not a range. A picked set
    is filtered by membership, and must never be *described* as a range - that
    would claim the months in between.
    """

    def _frame(self):
        stamps = pd.date_range("2026-01-01", "2026-08-31", freq="6h")
        n = len(stamps)
        frame = pd.DataFrame({
            "number": [f"INC{i:05d}" for i in range(n)],
            "opened_at": stamps,
            "state": ["Closed"] * n,
        })
        frame.loc[frame.sample(9, random_state=1).index, "opened_at"] = pd.NaT
        return frame

    def test_only_the_picked_months_come_out(self):
        frame = self._frame()
        picked = ["2026-02", "2026-05", "2026-07"]
        built = build_extract(frame, _dates(frame), months=picked,
                             columns=["number", "opened_at"])
        got = sorted(built["opened_at"].dt.to_period("M").astype(str).unique())
        assert got == picked

    def test_the_plan_agrees_with_the_build(self):
        frame = self._frame()
        picked = ["2026-02", "2026-05", "2026-07"]
        plan = plan_extract(frame, _dates(frame), months=picked,
                            columns=["number", "opened_at"])
        built = build_extract(frame, _dates(frame), months=picked,
                             columns=["number", "opened_at"])
        assert plan.rows == len(built)
        assert [m for m, _count in plan.months] == picked

    def test_a_month_between_two_picked_ones_is_excluded(self):
        """The whole point: a range cannot express this."""
        frame = self._frame()
        built = build_extract(frame, _dates(frame),
                              months=["2026-02", "2026-04"],
                              columns=["number", "opened_at"])
        months = set(built["opened_at"].dt.to_period("M").astype(str))
        assert "2026-03" not in months

    def test_undated_rows_still_need_asking_for(self):
        frame = self._frame()
        picked = ["2026-02"]
        without = plan_extract(frame, _dates(frame), months=picked,
                              columns=["number"])
        with_them = plan_extract(frame, _dates(frame), months=picked,
                                columns=["number"], include_undated=True)
        assert without.undated_rows == 9
        assert with_them.rows == without.rows + 9

    def test_an_empty_selection_produces_nothing(self):
        frame = self._frame()
        assert plan_extract(frame, _dates(frame), months=[],
                            columns=["number"]).rows == 0
        assert build_extract(frame, _dates(frame), months=[],
                             columns=["number"]).empty

    def test_a_month_the_file_lacks_is_simply_absent(self):
        frame = self._frame()
        plan = plan_extract(frame, _dates(frame),
                            months=["2026-02", "2030-01"], columns=["number"])
        assert [m for m, _count in plan.months] == ["2026-02"]

    def test_months_take_precedence_over_a_range(self):
        """Both are accepted at the call site; the explicit set wins."""
        frame = self._frame()
        plan = plan_extract(frame, _dates(frame), start="2026-01",
                            end="2026-08", months=["2026-03"],
                            columns=["number"])
        assert [m for m, _count in plan.months] == ["2026-03"]

    def test_a_picked_set_is_never_described_as_a_range(self):
        assert months_label(["2026-02", "2026-05", "2026-07"]) == \
            "Feb-26, May-26 and Jul-26"
        assert not is_contiguous(["2026-02", "2026-05"])

    def test_a_contiguous_pick_reads_as_a_range(self):
        assert is_contiguous(["2026-02", "2026-03", "2026-04"])
        assert months_label(["2026-02", "2026-03", "2026-04"]) == \
            "Feb-26 to Apr-26"

    def test_many_months_are_summarised_with_their_span(self):
        many = ["2025-01", "2025-04", "2025-07", "2025-10", "2026-01"]
        assert months_label(many) == \
            "5 selected months between Jan-25 and Jan-26"

    def test_order_does_not_matter_to_the_label(self):
        assert months_label(["2026-07", "2026-02", "2026-05"]) == \
            "Feb-26, May-26 and Jul-26"

    def test_no_months_is_said_not_left_blank(self):
        assert months_label([]) == "no months"

    def test_the_filename_does_not_claim_the_months_between(self):
        """
        "2026-02_to_2026-07" for a three-month pick would name four months the
        file does not contain.
        """
        name = suggest_filename("Book1.xlsx", "opened_at", "", "", "csv",
                                months=["2026-02", "2026-05", "2026-07"])
        assert name == "Book1_opened_at_3_months_2026-02_to_2026-07.csv"

    def test_a_contiguous_pick_keeps_the_plain_range_filename(self):
        name = suggest_filename("Book1.xlsx", "opened_at", "", "", "csv",
                                months=["2026-02", "2026-03"])
        assert name == "Book1_opened_at_2026-02_to_2026-03.csv"

    def test_a_single_picked_month_names_itself(self):
        name = suggest_filename("Book1.xlsx", "opened_at", "", "", "csv",
                                months=["2026-02"])
        assert name == "Book1_opened_at_2026-02.csv"

    def test_the_output_lists_the_months_in_full(self, tmp_path):
        """
        A summarised label is fine on screen; the file has to be able to say
        exactly which months it holds.
        """
        frame = self._frame()
        picked = ["2026-02", "2026-05", "2026-07"]
        subset = build_extract(frame, _dates(frame), months=picked,
                               columns=["number", "opened_at"])
        out = tmp_path / "picked.xlsx"
        write_extract(str(out), subset, fmt=FORMAT_XLSX,
                      date_column="opened_at", months=picked)
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["Months included"] == "Feb-26, May-26 and Jul-26"
        assert rows["Months, in full"] == "Feb-26, May-26, Jul-26"

    def test_a_range_write_says_so(self, spanning_frame, tmp_path):
        subset = build_extract(spanning_frame, _dates(spanning_frame),
                               start="2025-01", end="2025-03",
                               columns=["number", "opened_at"])
        out = tmp_path / "ranged.xlsx"
        write_extract(str(out), subset, fmt=FORMAT_XLSX,
                      date_column="opened_at", start="2025-01", end="2025-03")
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["Months, in full"] == "a continuous range"


class TestCancelledStates:
    """
    A cancelled ticket was never worked, so it flatters resolution times and
    pads volume counts. Which values mean cancelled is a claim about someone
    else's vocabulary, so the values matched are reported rather than assumed.
    """

    def _frame(self, states):
        n = len(states)
        return pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": states,
            "opened_at": pd.date_range("2025-01-01", periods=n, freq="D"),
            "short_description": ["A long enough description here"] * n,
        })

    def test_every_spelling_is_caught(self):
        frame = self._frame(["Cancelled", "canceled", "CANCELLED ",
                             "Closed Cancelled", "Cancelled by user",
                             "Cancelling", "Closed"])
        found = SanityAnalyzer(frame).cancelled_states()
        assert found["rows"] == 6
        assert "Closed" not in found["values"]

    def test_words_that_only_arguably_mean_cancelled_are_kept(self):
        """
        Whether withdrawn or rejected counts is a judgement about a vocabulary
        this tool cannot see. Keeping them is the conservative error: the rows
        stay, and the label says what was matched.
        """
        frame = self._frame(["Withdrawn", "Rejected", "Abandoned", "Void",
                             "Cancelled"])
        found = SanityAnalyzer(frame).cancelled_states()
        assert found["values"] == ["Cancelled"]
        assert found["rows"] == 1

    def test_no_state_column_is_none_not_empty(self):
        """
        None means "no way to tell", an empty mask means "nothing to exclude",
        and a caller has to be able to distinguish them.
        """
        frame = pd.DataFrame({"number": ["INC0001"], "value": [3]})
        found = SanityAnalyzer(frame).cancelled_states()
        assert found["mask"] is None
        assert found["state_col"] is None

    def test_a_file_with_no_cancellations_reports_an_empty_mask(self):
        frame = self._frame(["Closed", "New", "Resolved"])
        found = SanityAnalyzer(frame).cancelled_states()
        assert found["mask"] is not None
        assert found["rows"] == 0
        assert found["values"] == []

    def test_nulls_do_not_count_as_cancelled(self):
        frame = self._frame(["Cancelled", None, "Closed"])
        found = SanityAnalyzer(frame).cancelled_states()
        assert found["rows"] == 1

    def test_the_values_are_stable_between_runs(self):
        """The label names them, so an order that shifts reads as a change."""
        frame = self._frame(["canceled", "Cancelled", "Closed Cancelled"])
        first = SanityAnalyzer(frame).cancelled_states()["values"]
        second = SanityAnalyzer(frame).cancelled_states()["values"]
        assert first == second == sorted(first)

    def test_a_categorical_state_column_works(self):
        """The dtype is an optimisation and must not change what is found."""
        frame = self._frame(["Cancelled", "Closed", "canceled"])
        frame["state"] = frame["state"].astype("category")
        assert SanityAnalyzer(frame).cancelled_states()["rows"] == 2


class TestExcludeRows:
    """`exclude` is a plain row mask: extract.py never learns what a state is."""

    def _frame(self):
        stamps = pd.date_range("2026-01-01", "2026-04-30", freq="12h")
        n = len(stamps)
        frame = pd.DataFrame({
            "number": range(n),
            "opened_at": stamps,
            "state": (["Closed", "Cancelled"] * (n // 2 + 1))[:n],
        })
        frame.loc[frame.index[:6], "opened_at"] = pd.NaT
        return frame

    def _cancelled(self, frame):
        return frame["state"] == "Cancelled"

    def test_excluded_rows_are_gone_from_the_output(self):
        frame = self._frame()
        built = build_extract(frame, _dates(frame), start="2026-01",
                             end="2026-04", columns=["number", "state"],
                             exclude=self._cancelled(frame))
        assert "Cancelled" not in set(built["state"])

    def test_the_plan_counts_what_it_dropped(self):
        frame = self._frame()
        plain = plan_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                             columns=["number"])
        pruned = plan_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                              columns=["number"],
                              exclude=self._cancelled(frame))
        assert pruned.excluded_rows == plain.rows - pruned.rows
        assert plain.excluded_rows == 0

    def test_the_plan_agrees_with_the_build(self):
        frame = self._frame()
        plan = plan_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                            columns=["number"], exclude=self._cancelled(frame))
        built = build_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                              columns=["number"], exclude=self._cancelled(frame))
        assert plan.rows == len(built)

    def test_month_counts_lose_the_excluded_rows_too(self):
        """A month total that still counted them would not match the file."""
        frame = self._frame()
        plan = plan_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                            columns=["number"], exclude=self._cancelled(frame))
        assert sum(count for _m, count in plan.months) == \
            plan.rows - (plan.undated_rows if plan.undated_included else 0)

    def test_undated_rows_report_what_ticking_would_add(self):
        """
        Not the file's undated count: some of those are excluded, so ticking
        the box would not add them. The number must not change meaning with
        the flag.
        """
        frame = self._frame()
        cancelled = self._cancelled(frame)
        undated_and_kept = int((frame["opened_at"].isna() & ~cancelled).sum())
        for include in (False, True):
            plan = plan_extract(frame, _dates(frame), start="2026-01",
                                end="2026-04", columns=["number"],
                                exclude=cancelled, include_undated=include)
            assert plan.undated_rows == undated_and_kept

    def test_including_undated_still_excludes_them(self):
        frame = self._frame()
        built = build_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                              columns=["number", "state", "opened_at"],
                              exclude=self._cancelled(frame),
                              include_undated=True)
        assert "Cancelled" not in set(built["state"])
        assert built["opened_at"].isna().any(), "undated rows were asked for"

    def test_it_works_with_no_date_column_at_all(self):
        frame = self._frame()
        plan = plan_extract(frame, None, columns=["number"],
                            exclude=self._cancelled(frame))
        built = build_extract(frame, None, columns=["number", "state"],
                              exclude=self._cancelled(frame))
        assert plan.rows == len(built) == int((~self._cancelled(frame)).sum())
        assert plan.excluded_rows == int(self._cancelled(frame).sum())

    def test_it_works_with_a_picked_month_set(self):
        frame = self._frame()
        plan = plan_extract(frame, _dates(frame), months=["2026-02", "2026-04"],
                            columns=["number"], exclude=self._cancelled(frame))
        built = build_extract(frame, _dates(frame), months=["2026-02", "2026-04"],
                              columns=["number", "state"],
                              exclude=self._cancelled(frame))
        assert plan.rows == len(built)
        assert "Cancelled" not in set(built["state"])

    def test_the_preview_limit_applies_after_the_exclusion(self):
        frame = self._frame()
        head = build_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                             columns=["number", "state"],
                             exclude=self._cancelled(frame), limit=5)
        assert len(head) == 5
        assert "Cancelled" not in set(head["state"])

    def test_no_exclusion_changes_nothing(self):
        frame = self._frame()
        with_none = plan_extract(frame, _dates(frame), start="2026-01",
                                 end="2026-04", columns=["number"], exclude=None)
        assert with_none.excluded_rows == 0

    def test_the_output_names_the_states_it_dropped(self, tmp_path):
        """The rows are simply absent, so the file has to say why."""
        frame = self._frame()
        subset = build_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                               columns=["number", "state"],
                               exclude=self._cancelled(frame))
        plan = plan_extract(frame, _dates(frame), start="2026-01", end="2026-04",
                            columns=["number", "state"],
                            exclude=self._cancelled(frame))
        out = tmp_path / "pruned.xlsx"
        write_extract(str(out), subset, fmt=FORMAT_XLSX, plan=plan,
                      date_column="opened_at", start="2026-01", end="2026-04",
                      excluded_states=["Cancelled"])
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["States excluded"] == "Cancelled"
        assert rows["Rows excluded by state"] == plan.excluded_rows

    def test_none_is_stated_when_nothing_was_excluded(self, spanning_frame,
                                                     tmp_path):
        subset = build_extract(spanning_frame, _dates(spanning_frame),
                               start="2025-01", end="2025-02",
                               columns=["number", "opened_at"])
        out = tmp_path / "whole.xlsx"
        write_extract(str(out), subset, fmt=FORMAT_XLSX,
                      date_column="opened_at", start="2025-01", end="2025-02")
        info = load_workbook(out)["Extract Info"]
        rows = {r[0].value: r[1].value
                for r in info.iter_rows(min_row=2, max_row=info.max_row)}
        assert rows["States excluded"] == "none"
        assert rows["Rows excluded by state"] == 0
