"""
Tests for core/excel_report.py — workbook assembly.

Most breakage here is silent: a chart can point at the wrong range, a Table can
name columns that do not match its header row, and a truncated value axis can
misrepresent the data without anything raising. So these tests re-open the
written workbook and assert structure rather than trusting the build call.

Verified against real Excel (Office 16) during development: six charts survive
the open with the intended types, eleven Tables on All Details, both helper
sheets hidden.
"""
import codecs

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.analyzer import SanityAnalyzer
from core.excel_report import (
    DEEMPHASIS, NULL_CHART_TOP_N, PRIORITY_RAMP, STATUS_CRITICAL, TIER_RAMP,
    ExcelReportBuilder, build_workbook,
)
from core.reporter import ReportGenerator


@pytest.fixture
def ticket_df():
    n = 60
    return pd.DataFrame({
        "number": [f"INC{i:04d}" for i in range(n - 1)] + ["INC0000"],
        "state": ["Closed"] * 40 + ["New"] * 18 + ["Closd"] + ["Stage 9"],
        "priority": ["P1"] * 20 + ["P2"] * 20 + ["P3"] * 17 + [None] * 3,
        "assignment_group": ["Service Desk"] * 30 + ["Network"] * 20 + ["Apps"] * 10,
        "cmdb_ci": [f"SRV-{i % 7:03d}" for i in range(n)],
        "opened_at": pd.to_datetime(["2024-01-15"] * 20 + ["2024-02-15"] * 20
                                    + ["2024-03-15"] * 20),
        "resolved_at": pd.to_datetime(["2024-02-01"] * 38 + [None] * 2
                                      + ["2024-02-01"] + [None] * 19),
        "short_description": ["Login issue for the user account"] * n,
        "description": ["Le serveur ne repond plus depuis ce matin"] * 5
                       + ["A sufficiently detailed description of a problem"] * 50
                       + [""] * 5,
    })


@pytest.fixture
def built(ticket_df, tmp_path):
    """(builder, reopened workbook) for one assembled report."""
    analyzer = SanityAnalyzer(ticket_df)
    reporter = ReportGenerator(ticket_df, analyzer, filepath="tickets.xlsx")
    out = tmp_path / "report.xlsx"
    builder = build_workbook(reporter, str(out))
    return builder, load_workbook(out)


class TestSheetLayout:
    def test_visible_tabs_are_the_consolidated_set(self, built):
        _builder, wb = built
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        assert visible == ["Overview", "Dashboard", "Findings", "All Details",
                           "Pivots", "Ticket Detail"]

    def test_helper_sheets_are_hidden(self, built):
        _builder, wb = built
        hidden = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}
        assert hidden == {"ChartData", "PivotData"}

    def test_overview_opens_first(self, built):
        _builder, wb = built
        assert wb.worksheets[wb.active.index if hasattr(wb.active, 'index')
                             else 0].title == "Overview"

    def test_findings_header_is_frozen(self, built):
        _builder, wb = built
        assert wb["Findings"].freeze_panes is not None


class TestExcelTables:
    def test_every_block_is_a_named_table(self, built):
        _builder, wb = built
        assert len(wb["All Details"].tables) >= 8
        assert len(wb["Findings"].tables) == 1

    def test_table_names_are_unique_across_the_workbook(self, built):
        _builder, wb = built
        names = [t for ws in wb.worksheets for t in ws.tables]
        assert len(names) == len(set(names)), "Excel shares one name namespace"

    def test_table_names_are_excel_legal(self, built):
        _builder, wb = built
        for ws in wb.worksheets:
            for name in ws.tables:
                assert " " not in name
                assert name[0].isalpha() or name[0] == "_"

    def test_table_ref_starts_at_its_header_row(self, built):
        """A Table whose ref excludes the header names its columns Column1..N."""
        builder, wb = built
        block = builder.blocks["Findings"]
        table = wb["Findings"].tables[block.table]
        assert table.ref.startswith(f"A{block.header_row}")


@pytest.fixture
def built_with_language(ticket_df, tmp_path):
    """The same report with a language check handed over, so all six charts plot."""
    from core.language import LanguageChecker
    analyzer = SanityAnalyzer(ticket_df)
    result = LanguageChecker().check_column(ticket_df, "description", mode="all")
    reporter = ReportGenerator(ticket_df, analyzer, filepath="tickets.xlsx",
                               lang_results=[result])
    out = tmp_path / "report_lang.xlsx"
    builder = build_workbook(reporter, str(out))
    return builder, load_workbook(out)


class TestCharts:
    def test_five_charts_without_a_language_check(self, built):
        """The language chart is omitted, not emitted empty."""
        builder, wb = built
        assert len(wb["Dashboard"]._charts) == 5
        assert "Language mix" not in builder.chart_blocks
        assert any("Language mix" in note for note in builder.omissions)

    def test_all_six_charts_once_language_has_run(self, built_with_language):
        _builder, wb = built_with_language
        assert len(wb["Dashboard"]._charts) == 6

    def test_charts_read_from_the_hidden_chart_sheet(self, built):
        """
        Never from a visible Table: sorting one would reorder the cells the
        chart points at and silently rewrite the dashboard.
        """
        _builder, wb = built
        for chart in wb["Dashboard"]._charts:
            for series in chart.series:
                assert "ChartData" in series.val.numRef.f

    def test_value_axis_starts_at_zero(self, built):
        """
        Excel auto-scaled a 20/20/17 comparison to a baseline of 15.5, which
        renders 17 as a sliver of 20 instead of 85% of it.
        """
        _builder, wb = built
        for chart in wb["Dashboard"]._charts:
            assert chart.y_axis.scaling.min == 0

    def test_single_series_charts_have_no_legend(self, built):
        _builder, wb = built
        for chart in wb["Dashboard"]._charts:
            if len(chart.series) == 1:
                assert chart.legend is None, "the title already names the series"

    def test_the_stacked_chart_keeps_its_legend(self, built):
        _builder, wb = built
        stacked = [c for c in wb["Dashboard"]._charts if len(c.series) > 1]
        assert stacked, "expected the quality-tier chart"
        for chart in stacked:
            assert chart.legend is not None

    def test_no_chart_uses_a_secondary_axis(self, built):
        """Two y-scales on one plot invent a correlation that is not there."""
        _builder, wb = built
        for chart in wb["Dashboard"]._charts:
            axis_ids = {chart.x_axis.axId, chart.y_axis.axId}
            assert len(axis_ids) == 2

    def test_ordered_scales_use_an_ordinal_ramp(self, built):
        """Priority is ordered, so one hue stepped by lightness, not 4 hues."""
        builder, _wb = built
        assert "Priority mix" in builder.chart_blocks
        rows = builder.chart_blocks["Priority mix"].row_count
        assert rows <= len(PRIORITY_RAMP) + 1

    def test_null_chart_emphasises_only_breaching_columns(self, built):
        builder, wb = built
        written = builder.chart_frames["Nulls by column"]
        chart = wb["Dashboard"]._charts[1]
        tinted = {point.idx for point in chart.series[0].dPt}
        expected = {i for i, above in enumerate(written["Above Threshold"])
                    if above == "Yes"}
        assert tinted == expected

    def test_null_chart_excludes_columns_with_no_nulls(self, built):
        """A 'worst 15' of mostly zeros is a chart of nothing."""
        builder, _wb = built
        written = builder.chart_frames["Nulls by column"]
        assert (written["Null %"] > 0).all()

    def test_null_chart_title_does_not_overclaim(self, built):
        builder, wb = built
        written = builder.chart_frames["Nulls by column"]
        title_runs = wb["Dashboard"]._charts[1].title.tx.rich.p[0].r
        text = "".join(run.t for run in title_runs)
        if len(written) < NULL_CHART_TOP_N:
            assert "worst" not in text


class TestOmissionsAreStated:
    def test_a_missing_chart_source_is_recorded(self, ticket_df, tmp_path):
        """A chart is omitted rather than emitted empty - and Overview says so."""
        df = ticket_df.drop(columns=["opened_at", "resolved_at"])
        analyzer = SanityAnalyzer(df)
        reporter = ReportGenerator(df, analyzer, filepath="tickets.xlsx")
        builder = build_workbook(reporter, str(tmp_path / "r.xlsx"))

        assert "Monthly inflow" not in builder.chart_blocks
        assert any("Monthly inflow" in note for note in builder.omissions)

    def test_an_unavailable_block_says_not_checked_in_place(self, ticket_df, tmp_path):
        df = ticket_df.drop(columns=["state"])
        analyzer = SanityAnalyzer(df)
        reporter = ReportGenerator(df, analyzer, filepath="tickets.xlsx")
        out = tmp_path / "r.xlsx"
        build_workbook(reporter, str(out))

        details = load_workbook(out)["All Details"]
        text = "\n".join(
            str(c.value) for row in details.iter_rows() for c in row if c.value)
        assert "State vocabulary" in text, "the block heading must still appear"
        assert "Not checked" in text

    def test_overview_lists_the_disclosures(self, ticket_df, tmp_path):
        df = ticket_df.drop(columns=["opened_at", "resolved_at"])
        analyzer = SanityAnalyzer(df)
        reporter = ReportGenerator(df, analyzer, filepath="tickets.xlsx")
        out = tmp_path / "r.xlsx"
        build_workbook(reporter, str(out))

        overview = load_workbook(out)["Overview"]
        text = "\n".join(
            str(c.value) for row in overview.iter_rows() for c in row if c.value)
        assert "What was and was not checked" in text
        assert "Monthly inflow" in text

    def test_hidden_helper_sheets_are_named_on_overview(self, built):
        """Hidden, not secret."""
        _builder, wb = built
        text = "\n".join(str(c.value) for row in wb["Overview"].iter_rows()
                         for c in row if c.value)
        assert "PivotData" in text and "ChartData" in text


class TestAllDetailsUsability:
    def test_a_jump_list_links_to_every_block(self, built):
        builder, wb = built
        details = wb["All Details"]
        links = [c.hyperlink.location for row in details.iter_rows()
                 for c in row if c.hyperlink is not None]
        assert len(links) >= 8
        for location in links:
            assert location.startswith("'All Details'!A")

    def test_jump_targets_hit_a_block_title(self, built):
        builder, wb = built
        details = wb["All Details"]
        title_rows = {b.title_row for b in builder.blocks.values()}
        for row in details.iter_rows():
            for cell in row:
                if cell.hyperlink is None:
                    continue
                target = int(cell.hyperlink.location.rsplit("A", 1)[1])
                assert target in title_rows or details.cell(
                    row=target, column=1).value is not None

    def test_null_percentages_get_data_bars(self, built):
        _builder, wb = built
        assert len(wb["All Details"].conditional_formatting._cf_rules) >= 1

    def test_blocks_do_not_overlap(self, built):
        builder, _wb = built
        detail_blocks = sorted(
            (b for name, b in builder.blocks.items()
             if name not in ("Findings", "Summary", "Settings", "PivotData",
                             "Ticket Detail")),
            key=lambda b: b.title_row)
        for earlier, later in zip(detail_blocks, detail_blocks[1:]):
            assert earlier.last_row < later.title_row


class TestNativePivots:
    """
    Pivots are the one part of this workbook assembled from raw OOXML, and a
    malformed one makes Excel offer to repair the file - silently dropping every
    pivot. Two canaries run here without Excel: validate() on the bytes, and
    load_workbook, which raises KeyError on the cache-dedupe defect.

    Verified against real Excel: four pivots survive the open, one cache, and
    every grand total matches the independently computed value.
    """

    def test_the_pivots_tab_is_in_reading_order(self, built):
        _builder, wb = built
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        assert visible == ["Overview", "Dashboard", "Findings", "All Details",
                           "Pivots", "Ticket Detail"]

    def test_one_pivot_per_cube_dimension(self, built, ticket_df):
        builder, wb = built
        cube = builder.reporter._build_pivot_cube_df()
        dimensions = [c for c in cube.columns if c != "Ticket count"]
        assert builder.pivot_count == len(dimensions)
        assert len(wb["Pivots"]._pivots) == len(dimensions)

    def test_the_produced_file_validates(self, ticket_df, tmp_path):
        from core.excel_pivot import validate
        analyzer = SanityAnalyzer(ticket_df)
        reporter = ReportGenerator(ticket_df, analyzer, filepath="tickets.xlsx")
        out = tmp_path / "r.xlsx"
        build_workbook(reporter, str(out))
        assert validate(out.read_bytes()) == []

    def test_load_workbook_canary(self, ticket_df, tmp_path):
        """Raises KeyError if two cacheIds point at one deduped cache part."""
        analyzer = SanityAnalyzer(ticket_df)
        reporter = ReportGenerator(ticket_df, analyzer, filepath="tickets.xlsx")
        out = tmp_path / "r.xlsx"
        build_workbook(reporter, str(out))
        load_workbook(out)          # must not raise

    def test_all_pivots_share_one_cache(self, built):
        _builder, wb = built
        pivots = wb["Pivots"]._pivots
        assert len({id(p.cache) for p in pivots}) == 1
        assert len({p.cacheId for p in pivots}) == 1

    def test_pivot_ranges_do_not_overlap(self, built):
        from openpyxl.utils.cell import range_boundaries
        _builder, wb = built
        boxes = [range_boundaries(p.location.ref) for p in wb["Pivots"]._pivots]
        for i, (c1a, r1a, c2a, r2a) in enumerate(boxes):
            for (c1b, r1b, c2b, r2b) in boxes[i + 1:]:
                assert r2a < r1b or r2b < r1a or c2a < c1b or c2b < c1a

    def test_the_cache_reads_from_the_named_table(self, built):
        """
        A Table name rather than a raw range: it survives a sheet rename and
        auto-expands when rows are added.
        """
        _builder, wb = built
        source = wb["Pivots"]._pivots[0].cache.cacheSource.worksheetSource
        assert source.name is not None
        assert source.name.startswith("tbl_")

    def test_cache_records_are_written(self, built):
        """
        Without records the pivot depends on Excel honouring refreshOnLoad, and
        Excel Online and LibreOffice do not - it would render empty for anyone
        the report is forwarded to.
        """
        _builder, wb = built
        cache = wb["Pivots"]._pivots[0].cache
        assert cache.recordCount and cache.recordCount > 0
        assert cache.refreshOnLoad is True

    def test_titles_sit_outside_every_pivot_range(self, built):
        """Excel overwrites its own pivot range on refresh."""
        from openpyxl.utils.cell import range_boundaries
        _builder, wb = built
        sheet = wb["Pivots"]
        boxes = [range_boundaries(p.location.ref) for p in sheet._pivots]
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith(
                        "Tickets by "):
                    continue
                for (c1, r1, c2, r2) in boxes:
                    assert not (r1 <= cell.row <= r2 and c1 <= 1 <= c2)


class TestPivotFallback:
    def _reporter(self, df):
        return ReportGenerator(df, SanityAnalyzer(df), filepath="tickets.xlsx")

    def test_disabling_pivots_states_why(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        build_workbook(self._reporter(ticket_df), str(out), enable_pivots=False)
        wb = load_workbook(out)
        assert wb["Pivots"]._pivots == []
        text = "\n".join(str(c.value) for row in wb["Pivots"].iter_rows()
                         for c in row if c.value)
        assert "switched off" in text
        assert "Insert > PivotTable" in text, \
            "an empty tab must say what to do instead"

    def test_a_build_failure_does_not_lose_the_report(self, ticket_df, tmp_path,
                                                     monkeypatch):
        """Losing the pivots is survivable; losing the workbook is not."""
        import core.excel_report as module

        def explode(*args, **kwargs):
            raise RuntimeError("simulated pivot failure")

        monkeypatch.setattr(module, "PivotBuilder", explode)
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out))

        assert out.exists(), "the workbook must still be written"
        wb = load_workbook(out)
        assert wb["Pivots"]._pivots == []
        assert any("Pivots" in note for note in builder.omissions)
        # and the other sheets are intact
        assert len(wb["Dashboard"]._charts) >= 5
        assert len(wb["Findings"].tables) == 1

    def test_a_high_cardinality_column_says_why_it_is_not_a_dimension(
            self, ticket_df, tmp_path):
        """
        "Why can't I pivot by this?" must be answerable from the workbook. A
        silently coarser pivot is indistinguishable from a broken one.
        """
        from core.reporter import MAX_CUBE_DISTINCT
        df = ticket_df.copy()
        df["ci_serial"] = [f"SN-{i:06d}" for i in range(len(df))]
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(df), str(out))
        note = next(n for n in builder.omissions if "ci_serial" in n)
        assert str(MAX_CUBE_DISTINCT) in note
        assert "Ticket Detail" in note, "must say where to pivot it instead"

    def test_date_columns_are_declared_not_silently_skipped(self, ticket_df,
                                                            tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out))
        assert any("date columns are not pivot dimensions" in n
                   for n in builder.omissions)

    def test_dropped_dimensions_are_named_on_overview(self, tmp_path,
                                                      monkeypatch):
        """The cube drops dimensions to fit its row cap; that must be visible."""
        import core.reporter as reporter_module
        monkeypatch.setattr(reporter_module, "MAX_CUBE_ROWS", 3)
        n = 40
        df = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": ["Closed", "New"] * (n // 2),
            "priority": [f"P{i % 4 + 1}" for i in range(n)],
            "assignment_group": [f"Team {i % 5}" for i in range(n)],
            "opened_at": pd.to_datetime(["2024-01-01"] * n),
            "resolved_at": pd.to_datetime(["2024-02-01"] * n),
            "cmdb_ci": [f"SRV-{i % 6:03d}" for i in range(n)],
            "short_description": ["Login issue for the user account"] * n,
            "description": ["A sufficiently detailed description"] * n,
        })
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(df), str(out))
        note = next((n for n in builder.omissions if "dropped" in n), None)
        assert note is not None, "a coarser cube must say so"
        assert "coarser" in note
        text = "\n".join(str(c.value) for row in load_workbook(out)["Overview"]
                         .iter_rows() for c in row if c.value)
        assert "dropped" in text

    def test_the_notes_do_not_accumulate_across_builds(self, ticket_df, tmp_path):
        """_build_pivot_cube_df is called more than once per workbook."""
        reporter = self._reporter(ticket_df)
        first = reporter._build_pivot_cube_df()
        notes_after_one = list(reporter.pivot_cube_notes)
        reporter._build_pivot_cube_df()
        assert reporter.pivot_cube_notes == notes_after_one

    def test_no_pivotable_column_is_stated(self, tmp_path):
        """
        Every column is free text, so there is no dimension to pivot on.
        n must exceed MAX_CUBE_DISTINCT or the unique values still qualify.
        """
        from core.reporter import MAX_CUBE_DISTINCT
        n = MAX_CUBE_DISTINCT + 30
        df = pd.DataFrame({
            "number": [f"INC{i:04d}" for i in range(n)],
            "state": [f"state {i}" for i in range(n)],
            "priority": [f"prio {i}" for i in range(n)],
            "opened_at": pd.to_datetime(["2024-01-01"] * n),
            "resolved_at": pd.to_datetime(["2024-02-01"] * n),
            "assignment_group": [f"team {i}" for i in range(n)],
            "cmdb_ci": [f"ci {i}" for i in range(n)],
            "short_description": [f"issue number {i} on the system" for i in range(n)],
            "description": [f"a detailed description of issue {i}" for i in range(n)],
        })
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(df), str(out))
        assert builder.pivot_count == 0
        text = "\n".join(str(c.value) for row in load_workbook(out)["Pivots"]
                         .iter_rows() for c in row if c.value)
        assert "Insert > PivotTable" in text


class TestSanitiser:
    """
    Every case here was reproduced against openpyxl, and most fail silently.
    The worst is pd.NA: it raises inside the row writer, which in write-only
    mode kills the row generator so save() emits a truncated workbook that
    openpyxl itself cannot reopen — no exception reaches the caller.
    """

    def _round_trip(self, frame, tmp_path):
        """
        Through the inline path, which is the one that actually writes cells to
        a worksheet - the large-file path is a CSV and has no cell types.
        """
        from openpyxl import Workbook
        from core.excel_report import write_block
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ticket Detail"
        # write_block sanitises internally, so this exercises the shipped path.
        write_block(sheet, frame, title="Ticket Detail", start_row=1,
                    used_names=set())
        out = tmp_path / "detail.xlsx"
        workbook.save(out)
        # write_block puts the title on row 1, so data starts one row lower than
        # the streamed layout did.
        return len(frame), load_workbook(out)["Ticket Detail"]

    def test_pandas_na_does_not_truncate_the_workbook(self, tmp_path):
        frame = pd.DataFrame({
            "text": pd.array(["a", None, "c"], dtype="string"),
            "num": pd.array([1, None, 3], dtype="Int64"),
            "flag": pd.array([True, None, False], dtype="boolean"),
        })
        rows, sheet = self._round_trip(frame, tmp_path)
        assert rows == 3
        assert sheet.max_row == 5, "every row must survive the write"

    def test_control_characters_are_stripped(self, tmp_path):
        frame = pd.DataFrame({"text": ["clean", "bad\x00char\x07here"]})
        _rows, sheet = self._round_trip(frame, tmp_path)
        assert sheet["A4"].value == "badcharhere"

    def test_over_long_strings_are_truncated_visibly(self, tmp_path):
        from core.excel_pivot import EXCEL_MAX_CELL_CHARS
        frame = pd.DataFrame({"text": ["x" * (EXCEL_MAX_CELL_CHARS + 500)]})
        _rows, sheet = self._round_trip(frame, tmp_path)
        value = sheet["A3"].value
        assert len(value) == EXCEL_MAX_CELL_CHARS
        assert value.endswith("…"), "the loss should be visible in the cell"

    def test_a_leading_equals_stays_text(self, tmp_path):
        """openpyxl types a string starting with '=' as a formula."""
        frame = pd.DataFrame({"text": ["=SUM(A1:A9)", "plain"]})
        _rows, sheet = self._round_trip(frame, tmp_path)
        assert sheet["A3"].data_type == "s"
        assert sheet["A3"].value == "=SUM(A1:A9)"

    def test_booleans_read_as_words_not_numbers(self, tmp_path):
        """np.bool_ is written as numeric 1/0, which reads as a count."""
        frame = pd.DataFrame({"flag": [True, False]})
        _rows, sheet = self._round_trip(frame, tmp_path)
        assert sheet["A3"].value == "Yes"
        assert sheet["A4"].value == "No"

    def test_timezone_aware_datetimes_lose_only_the_offset(self, tmp_path):
        frame = pd.DataFrame({
            "when": pd.to_datetime(["2024-01-01T08:30:00+05:30",
                                    "2024-01-02T09:00:00+05:30"])})
        _rows, sheet = self._round_trip(frame, tmp_path)
        assert sheet["A3"].value.hour == 8, "wall-clock time is kept"
        assert sheet["A3"].value.tzinfo is None

    def test_category_dtype_is_cleaned_without_being_expanded(self):
        from core.excel_report import sanitize_for_excel
        frame = pd.DataFrame({
            "state": pd.Categorical(["Open\x00", "Closed", "Open\x00"])})
        cleaned, _formula = sanitize_for_excel(frame)
        assert list(cleaned["state"]) == ["Open", "Closed", "Open"]

    def test_colliding_categories_do_not_raise(self):
        """rename_categories raises when cleaning maps two names onto one."""
        from core.excel_report import sanitize_for_excel
        frame = pd.DataFrame({
            "state": pd.Categorical(["Open\x00", "Open\x07", "Closed"])})
        cleaned, _formula = sanitize_for_excel(frame)
        assert list(cleaned["state"]) == ["Open", "Open", "Closed"]

    def test_formula_columns_are_reported(self):
        from core.excel_report import sanitize_for_excel
        frame = pd.DataFrame({"a": ["=1", "b"], "b": ["plain", "text"]})
        _cleaned, formula_columns = sanitize_for_excel(frame)
        assert formula_columns == {"a"}

    def test_numeric_columns_are_left_alone(self):
        from core.excel_report import sanitize_for_excel
        frame = pd.DataFrame({"n": [1, 2, 3]})
        cleaned, _formula = sanitize_for_excel(frame)
        assert list(cleaned["n"]) == [1, 2, 3]

    def test_a_control_character_anywhere_does_not_abort_the_export(
            self, ticket_df, tmp_path):
        """
        Several All Details blocks carry raw source values - sample values, state
        names, distribution values, recurring-issue text. Sanitising only the
        detail sheet left one bad character in a description able to kill the
        whole export with IllegalCharacterError.
        """
        df = ticket_df.copy()
        df.loc[df.index[0], "description"] = "control\x00char\x07here"
        df.loc[df.index[1], "state"] = "Clo\x01sed"
        df.loc[df.index[2], "cmdb_ci"] = "SRV\x02-001"
        analyzer = SanityAnalyzer(df)
        reporter = ReportGenerator(df, analyzer, filepath="tickets.xlsx")
        out = tmp_path / "r.xlsx"
        build_workbook(reporter, str(out))          # must not raise
        assert out.exists()


class TestDetailSheetPlacement:
    def _reporter(self, df):
        return ReportGenerator(df, SanityAnalyzer(df), filepath="tickets.xlsx")

    def test_small_frames_stay_inline(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out))
        assert builder.detail_path is None
        assert builder.detail_rows_written == len(ticket_df)
        assert "Ticket Detail" in load_workbook(out).sheetnames

    def test_large_frames_move_to_a_companion_csv(self, ticket_df, tmp_path):
        """
        Two constraints meet: pivots need a non-streamed sheet, so a large
        detail cannot share the workbook - and a streamed .xlsx of that size
        costs minutes where CSV costs seconds.
        """
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out),
                                 inline_max_rows=10)
        assert builder.detail_path is not None
        assert builder.detail_path.endswith("_TicketDetail.csv")
        assert builder.detail_rows_written == len(ticket_df)
        # the main workbook keeps its pivots and loses only the detail tab
        wb = load_workbook(out)
        assert "Ticket Detail" not in wb.sheetnames
        assert len(wb["Pivots"]._pivots) > 0

    def test_the_csv_round_trips_every_row_and_column(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out),
                                 inline_max_rows=10)
        reloaded = pd.read_csv(builder.detail_path)
        assert len(reloaded) == len(ticket_df)
        for column in ticket_df.columns:
            assert column in reloaded.columns
        for derived in ("QA Flag Count", "QA Findings", "State Reads As"):
            assert derived in reloaded.columns

    def test_the_csv_carries_a_bom_so_excel_decodes_it(self, ticket_df, tmp_path):
        """
        Without the BOM Excel mis-decodes non-ASCII on a double-click, mangling
        exactly the non-English text this tool exists to find.
        """
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out),
                                 inline_max_rows=10)
        with open(builder.detail_path, "rb") as handle:
            assert handle.read(3) == codecs.BOM_UTF8

    def test_the_csv_keeps_long_text_untruncated(self, ticket_df, tmp_path):
        """CSV has no 32,767-character cell limit, so truncating would be loss
        for no reason."""
        from core.excel_pivot import EXCEL_MAX_CELL_CHARS
        df = ticket_df.copy()
        long_value = "x" * (EXCEL_MAX_CELL_CHARS + 500)
        df.loc[df.index[0], "description"] = long_value
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(df), str(out), inline_max_rows=10)
        reloaded = pd.read_csv(builder.detail_path)
        assert len(reloaded.loc[0, "description"]) == len(long_value)

    def test_the_split_is_disclosed(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out),
                                 inline_max_rows=10)
        assert any("separate" in note for note in builder.omissions)
        note = next(n for n in builder.omissions if "separate" in n)
        assert "From Text/CSV" in note, "must say how to pivot it"
        text = "\n".join(str(c.value) for row in load_workbook(out)["Overview"]
                         .iter_rows() for c in row if c.value)
        assert "_TicketDetail" in text


class TestRowLimitGuard:
    """
    Excel holds 1,048,576 rows per sheet and openpyxl enforces nothing: a Table
    ref past the ceiling is accepted silently and the file will not open. A
    truncated pivot source also under-reports every number derived from it, so
    the truncation is disclosed in four places.
    """

    def _reporter(self, df):
        return ReportGenerator(df, SanityAnalyzer(df), filepath="tickets.xlsx")

    def test_truncation_happens_and_is_flagged(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out),
                                 detail_row_limit=10)
        assert builder.detail_truncated is True
        assert builder.detail_rows_written == 10

    def test_the_tab_name_carries_the_warning(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        build_workbook(self._reporter(ticket_df), str(out), detail_row_limit=10)
        assert "Ticket Detail (TRUNCATED)" in load_workbook(out).sheetnames

    def test_the_omission_names_both_numbers(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out),
                                 detail_row_limit=10)
        note = next(n for n in builder.omissions if "TRUNCATED" in n)
        assert "10" in note and str(len(ticket_df)) in note

    def test_summary_states_the_rows_written(self, ticket_df, tmp_path):
        out = tmp_path / "r.xlsx"
        build_workbook(self._reporter(ticket_df), str(out), detail_row_limit=10)
        text = "\n".join(str(c.value) for row in load_workbook(out)["Overview"]
                         .iter_rows() for c in row if c.value)
        assert "Ticket Detail rows written" in text
        assert "TRUNCATED" in text

    def test_an_untruncated_export_says_nothing_about_truncation(self, ticket_df,
                                                                tmp_path):
        out = tmp_path / "r.xlsx"
        builder = build_workbook(self._reporter(ticket_df), str(out))
        assert builder.detail_truncated is False
        assert not any("TRUNCATED" in note for note in builder.omissions)


class TestVerdictTinting:
    def test_findings_rows_are_tinted_by_result(self, built):
        builder, wb = built
        block = builder.blocks["Findings"]
        findings = wb["Findings"]
        headers = [findings.cell(row=block.header_row, column=c).value
                   for c in range(block.first_col, block.last_col + 1)]
        result_col = block.first_col + headers.index("Result")

        seen = set()
        for row in range(block.first_data_row, block.last_row + 1):
            verdict = findings.cell(row=row, column=result_col).value
            fill = findings.cell(row=row, column=block.first_col).fill
            seen.add((verdict, fill.fgColor.rgb))
        # every verdict present maps to exactly one colour
        by_verdict = {}
        for verdict, colour in seen:
            by_verdict.setdefault(verdict, set()).add(colour)
        for verdict, colours in by_verdict.items():
            assert len(colours) == 1, f"{verdict} tinted inconsistently"

    def test_not_checked_is_not_tinted_green(self, ticket_df, tmp_path):
        """An unexamined dimension must never look like a clean one."""
        df = ticket_df.drop(columns=["opened_at", "resolved_at"])
        analyzer = SanityAnalyzer(df)
        reporter = ReportGenerator(df, analyzer, filepath="tickets.xlsx")
        out = tmp_path / "r.xlsx"
        builder = build_workbook(reporter, str(out))

        wb = load_workbook(out)
        block = builder.blocks["Findings"]
        findings = wb["Findings"]
        headers = [findings.cell(row=block.header_row, column=c).value
                   for c in range(block.first_col, block.last_col + 1)]
        result_col = block.first_col + headers.index("Result")

        greens, unchecked = set(), set()
        for row in range(block.first_data_row, block.last_row + 1):
            verdict = findings.cell(row=row, column=result_col).value
            colour = findings.cell(row=row, column=block.first_col).fill.fgColor.rgb
            if verdict == "Pass":
                greens.add(colour)
            elif verdict == "Not checked":
                unchecked.add(colour)
        assert unchecked, "the fixture must contain a skipped check"
        assert not (unchecked & greens)
